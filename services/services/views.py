import json
import mimetypes
import random
import re
import shutil
import zipfile
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, IntegerField, OuterRef, Q, Subquery
from django.db.models.functions import Lower
from django.core.serializers.json import DjangoJSONEncoder
from django.conf import settings
from django.http import HttpRequest
from django.http import FileResponse
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date

from .forms import (
    AcceptanceDocumentForm,
    AcceptanceEquipmentCreateForm,
    AddressForm,
    AdminUserCreateForm,
    BrandForm,
    CatalogAttachmentForm,
    ClientEquipmentForm,
    ConsumableForm,
    EquipmentCharacteristicTypeForm,
    GroupCreateForm,
    OrganizationContactForm,
    OrganizationForm,
    PartForm,
    ProductCategoryForm,
    ProductModelForm,
    RepairDocumentConsumableForm,
    RepairDocumentForm,
    RepairDocumentPartForm,
    RepairDocumentWorkForm,
    ShipmentDocumentForm,
    ServiceCenterContactForm,
    ServiceCenterForm,
    ServiceManForm,
    StatusDirectoryForm,
    WorkDirectoryForm,
)
from .models import (
    AcceptanceDocument,
    AcceptanceDocumentAttachment,
    AcceptanceDocumentEquipment,
    Address,
    Brand,
    ClientEquipment,
    Consumable,
    ConsumableAttachment,
    ConsumableCharacteristic,
    ConsumableCompatibility,
    EquipmentCharacteristicType,
    Organization,
    OrganizationAddress,
    OrganizationContact,
    Part,
    PartAttachment,
    PartCharacteristic,
    PartCompatibility,
    ProductCategory,
    ProductModel,
    ProductModelAttachment,
    ProductModelCharacteristic,
    RepairDocument,
    RepairDocumentAttachment,
    RepairDocumentConsumable,
    RepairDocumentPart,
    RepairDocumentWork,
    ShipmentDocument,
    ShipmentDocumentAttachment,
    ShipmentDocumentEquipment,
    ServiceExchangeLog,
    ServiceCenter,
    ServiceCenterAddress,
    ServiceCenterContact,
    ServiceMan,
    StatusDirectory,
    WorkDirectory,
    WorkDirectoryConsumable,
    WorkDirectoryPart,
)
from .theme_utils import (
    delete_uploaded_theme,
    get_active_theme,
    get_default_theme_key,
    get_theme_asset_path,
    get_theme_by_key,
    install_theme_from_zip,
    list_available_themes,
    set_active_theme,
    set_default_theme_key,
)

User = get_user_model()


def _first_form_error(form):
    for errors in form.errors.values():
        if errors:
            return errors[0]
    return "Проверьте корректность заполнения формы."


def _save_catalog_attachment(owner_field, owner, attachment_model, attachment_form):
    attachment = attachment_form.save(commit=False) if hasattr(attachment_form, "save") else None
    if attachment is not None:
        setattr(attachment, owner_field, owner)
        attachment.save()
        return attachment

    payload = attachment_form.cleaned_data
    return attachment_model.objects.create(
        **{
            owner_field: owner,
            "title": payload.get("title", ""),
            "file": payload["file"],
        }
    )


@login_required(login_url="login")
def style_settings(request):
    themes = list_available_themes()
    default_theme_key = get_default_theme_key(themes=themes)
    preview_theme_key = request.GET.get("preview_theme", "").strip()

    if request.method == "POST":
        action = request.POST.get("action", "").strip()

        if action == "upload_theme":
            uploaded_file = request.FILES.get("theme_zip")
            if not uploaded_file:
                messages.error(request, "Выберите ZIP-файл темы Bootstrap.")
            else:
                try:
                    theme = install_theme_from_zip(uploaded_file)
                except ValueError as error:
                    messages.error(request, str(error))
                else:
                    set_active_theme(request, theme["key"])
                    messages.success(request, f"Тема «{theme['name']}» загружена и активирована.")
            return redirect(reverse("style_settings"))

        if action == "select_theme":
            theme_key = request.POST.get("theme_key", "").strip()
            theme = get_theme_by_key(theme_key, themes=themes)
            if not theme:
                messages.error(request, "Выбранная тема не найдена.")
            else:
                set_active_theme(request, theme_key)
                messages.success(request, f"Активирована тема «{theme['name']}».")
            return redirect(reverse("style_settings"))

        if action == "set_default_theme":
            theme_key = request.POST.get("theme_key", "").strip()
            theme = get_theme_by_key(theme_key, themes=themes)
            if not theme:
                messages.error(request, "Тема для установки по умолчанию не найдена.")
            else:
                set_default_theme_key(theme_key)
                messages.success(request, f"Тема «{theme['name']}» назначена темой по умолчанию.")
            return redirect(reverse("style_settings"))

        if action == "delete_theme":
            theme_key = request.POST.get("theme_key", "").strip()
            theme = get_theme_by_key(theme_key, themes=themes)
            if not theme:
                messages.error(request, "Тема для удаления не найдена.")
                return redirect(reverse("style_settings"))
            try:
                delete_uploaded_theme(theme_key)
            except ValueError as error:
                messages.error(request, str(error))
            else:
                if request.session.get("active_theme_key") == theme_key:
                    request.session.pop("active_theme_key", None)
                    request.session.modified = True
                if default_theme_key == theme_key:
                    set_default_theme_key("builtin-light")
                messages.success(request, f"Тема «{theme['name']}» удалена.")
            return redirect(reverse("style_settings"))

    context = {
        "themes": themes,
        "active_theme": get_active_theme(request, themes=themes),
        "latest_uploaded_theme": next((theme for theme in themes if theme["source"] == "uploaded"), None),
        "default_theme_key": default_theme_key,
        "preview_theme_key": preview_theme_key,
        "preview_theme": get_theme_by_key(preview_theme_key, themes=themes),
    }
    return render(request, "style_settings.html", context)


@login_required(login_url="login")
def style_select(request, theme_key):
    themes = list_available_themes()
    theme = get_theme_by_key(theme_key, themes=themes)
    if not theme:
        messages.error(request, "Тема не найдена.")
        return redirect(reverse("style_settings"))

    set_active_theme(request, theme_key)
    messages.success(request, f"Активирована тема «{theme['name']}».")
    next_url = request.GET.get("next", "").strip()
    return redirect(next_url or reverse("style_settings"))


def theme_asset(request, theme_key, asset_path):
    try:
        file_path = get_theme_asset_path(theme_key, asset_path)
    except FileNotFoundError:
        return HttpResponse(status=404)

    content_type, _ = mimetypes.guess_type(str(file_path))
    return FileResponse(file_path.open("rb"), content_type=content_type or "application/octet-stream")


SERVICE_EXPORT_GROUPS = {
    "directories": {
        "label": "Справочники",
        "items": [
            {"key": "status_directory", "label": "Статусы", "model": StatusDirectory},
            {"key": "address", "label": "Адреса", "model": Address},
            {"key": "brand", "label": "Бренды", "model": Brand},
            {"key": "product_category", "label": "Категории товаров", "model": ProductCategory},
            {"key": "organization", "label": "Организации", "model": Organization},
            {"key": "service_center", "label": "Сервисные центры", "model": ServiceCenter},
            {"key": "serviceman", "label": "Сервисные инженеры", "model": ServiceMan},
            {"key": "product_model", "label": "Техника", "model": ProductModel},
            {"key": "equipment_characteristic_type", "label": "Типы характеристик техники", "model": EquipmentCharacteristicType},
            {"key": "product_model_characteristic", "label": "Характеристики техники", "model": ProductModelCharacteristic},
            {"key": "consumable", "label": "Расходные материалы", "model": Consumable},
            {"key": "consumable_characteristic", "label": "Характеристики расходных материалов", "model": ConsumableCharacteristic},
            {"key": "part", "label": "Запчасти", "model": Part},
            {"key": "part_characteristic", "label": "Характеристики запчастей", "model": PartCharacteristic},
            {"key": "work_directory", "label": "Работы", "model": WorkDirectory},
            {"key": "work_directory_consumable", "label": "Расходные материалы в работах", "model": WorkDirectoryConsumable},
            {"key": "work_directory_part", "label": "Запчасти в работах", "model": WorkDirectoryPart},
            {"key": "client_equipment", "label": "Техника клиентов", "model": ClientEquipment},
            {"key": "organization_address", "label": "Адреса организаций", "model": OrganizationAddress},
            {"key": "organization_contact", "label": "Контакты организаций", "model": OrganizationContact},
            {"key": "service_center_address", "label": "Адреса сервисных центров", "model": ServiceCenterAddress},
            {"key": "service_center_contact", "label": "Контакты сервисных центров", "model": ServiceCenterContact},
            {"key": "consumable_compatibility", "label": "Совместимость расходников", "model": ConsumableCompatibility},
            {"key": "part_compatibility", "label": "Совместимость запчастей", "model": PartCompatibility},
        ],
    },
    "documents": {
        "label": "Документы",
        "items": [
            {"key": "repair_document", "label": "Ремонт", "model": RepairDocument},
            {"key": "repair_document_work", "label": "Работы в ремонте", "model": RepairDocumentWork},
            {"key": "repair_document_consumable", "label": "Расходники в ремонте", "model": RepairDocumentConsumable},
            {"key": "repair_document_part", "label": "Запчасти в ремонте", "model": RepairDocumentPart},
            {"key": "acceptance_document", "label": "Приемка техники", "model": AcceptanceDocument},
            {"key": "acceptance_document_equipment", "label": "Техника в приемке", "model": AcceptanceDocumentEquipment},
            {"key": "shipment_document", "label": "Отгрузка техники", "model": ShipmentDocument},
            {"key": "shipment_document_equipment", "label": "Техника в отгрузке", "model": ShipmentDocumentEquipment},
        ],
    },
}

SERVICE_IMPORT_ORDER = [
    "status_directory",
    "address",
    "brand",
    "product_category",
    "organization",
    "service_center",
    "serviceman",
    "product_model",
    "product_model_characteristic",
    "consumable",
    "consumable_characteristic",
    "part",
    "part_characteristic",
    "work_directory",
    "work_directory_consumable",
    "work_directory_part",
    "client_equipment",
    "organization_address",
    "organization_contact",
    "service_center_address",
    "service_center_contact",
    "consumable_compatibility",
    "part_compatibility",
    "repair_document",
    "repair_document_work",
    "repair_document_consumable",
    "repair_document_part",
    "acceptance_document",
    "acceptance_document_equipment",
    "shipment_document",
    "shipment_document_equipment",
]

SERVICE_UI_ROW_ERROR_PREVIEW_LIMIT = 200
SERVICE_ARCHIVE_KEEP_FILES_DEFAULT = 30

SERVICE_CLEANUP_ORDER = list(reversed(SERVICE_IMPORT_ORDER))


def _service_item_map() -> dict:
    item_map = {}
    for group in SERVICE_EXPORT_GROUPS.values():
        for item in group["items"]:
            item_map[item["key"]] = item
    return item_map


def _service_selected_keys(group_key: str, request: HttpRequest) -> list[str]:
    group_config = SERVICE_EXPORT_GROUPS[group_key]
    all_keys = [item["key"] for item in group_config["items"]]
    mode = request.POST.get("selection_mode", "all")
    selected = request.POST.getlist("items") if mode == "specific" else all_keys
    selected = [key for key in selected if key in all_keys]
    return selected or all_keys


def _service_cleanup_selected_keys(request: HttpRequest) -> list[str]:
    all_keys = []
    for group_key in ("directories", "documents"):
        all_keys.extend([item["key"] for item in SERVICE_EXPORT_GROUPS[group_key]["items"]])

    mode = request.POST.get("selection_mode", "all")
    selected = request.POST.getlist("items") if mode == "specific" else all_keys
    selected = [key for key in selected if key in all_keys]
    return selected or all_keys


def _service_cleanup_rows(item_map: dict, selected_keys: list[str]) -> dict:
    summary = {
        "deleted": 0,
        "errors": 0,
        "datasets": {},
        "row_errors": [],
    }

    for key in SERVICE_CLEANUP_ORDER:
        if key not in selected_keys:
            continue

        item = item_map.get(key)
        if not item:
            continue

        model = item["model"]
        try:
            deleted_count, _ = model.objects.all().delete()
            summary["datasets"][key] = {
                "deleted": deleted_count,
                "errors": 0,
            }
            summary["deleted"] += deleted_count
        except Exception as exc:
            summary["datasets"][key] = {
                "deleted": 0,
                "errors": 1,
            }
            summary["errors"] += 1
            summary["row_errors"].append(
                {
                    "dataset": key,
                    "error": str(exc),
                }
            )

    return summary


def _archive_database_copy() -> dict:
    db_conf = settings.DATABASES.get("default", {})
    engine = db_conf.get("ENGINE", "")
    if engine != "django.db.backends.sqlite3":
        raise ValueError("Архивация поддерживается только для SQLite базы")

    db_name = db_conf.get("NAME")
    db_path = Path(db_name)
    if not db_path.exists():
        raise ValueError(f"Файл базы данных не найден: {db_path}")

    archive_dir = Path(settings.BASE_DIR) / "arxiv"
    archive_dir.mkdir(parents=True, exist_ok=True)
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    archive_file = archive_dir / f"data_db_{timestamp}.sqlite3"
    shutil.copy2(db_path, archive_file)

    keep_files = getattr(settings, "ARCHIVE_KEEP_FILES", SERVICE_ARCHIVE_KEEP_FILES_DEFAULT)
    try:
        keep_files = int(keep_files)
    except (TypeError, ValueError):
        keep_files = SERVICE_ARCHIVE_KEEP_FILES_DEFAULT
    keep_files = max(1, keep_files)

    archive_files = sorted(
        archive_dir.glob("data_db_*.sqlite3"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    deleted_old_archives = []
    if len(archive_files) > keep_files:
        for old_file in archive_files[keep_files:]:
            old_name = old_file.name
            old_file.unlink(missing_ok=True)
            deleted_old_archives.append(old_name)

    return {
        "source": str(db_path),
        "target": str(archive_file),
        "file_name": archive_file.name,
        "size_bytes": archive_file.stat().st_size,
        "keep_files": keep_files,
        "deleted_old_archives": deleted_old_archives,
    }


def _serialize_model_rows(model) -> list[dict]:
    fields = [field.attname for field in model._meta.concrete_fields]
    rows = []
    for obj in model.objects.all().order_by("pk"):
        row = {field_name: getattr(obj, field_name) for field_name in fields}
        rows.append(row)
    return rows


def _service_export_payload(group_key: str, selected_keys: list[str]) -> dict:
    item_map = _service_item_map()
    data = {}
    for key in selected_keys:
        model = item_map[key]["model"]
        data[key] = _serialize_model_rows(model)

    return {
        "manifest": {
            "version": 1,
            "group": group_key,
            "group_label": SERVICE_EXPORT_GROUPS[group_key]["label"],
            "generated_at": timezone.now().isoformat(),
            "datasets": selected_keys,
        },
        "data": data,
    }


def _service_parse_uploaded_payload(uploaded_file) -> dict:
    file_name = (uploaded_file.name or "").lower()

    if file_name.endswith(".zip"):
        with zipfile.ZipFile(uploaded_file) as archive:
            if "manifest.json" not in archive.namelist():
                raise ValueError("В архиве отсутствует manifest.json")

            manifest = json.loads(archive.read("manifest.json").decode("utf-8"))
            datasets = manifest.get("datasets") or []
            data = {}
            for key in datasets:
                data_file = f"data/{key}.json"
                if data_file in archive.namelist():
                    data[key] = json.loads(archive.read(data_file).decode("utf-8"))
                else:
                    data[key] = []

            return {
                "manifest": manifest,
                "data": data,
            }

    raw_payload = json.loads(uploaded_file.read().decode("utf-8"))
    if not isinstance(raw_payload, dict) or "data" not in raw_payload:
        raise ValueError("JSON должен содержать объект с полем data")
    return raw_payload


def _import_dataset_rows(item_map: dict, selected_keys: list[str], payload: dict) -> dict:
    rows_by_dataset = payload.get("data", {})
    summary = {
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "errors": 0,
        "datasets": {},
        "row_errors": [],
    }

    for key in SERVICE_IMPORT_ORDER:
        if key not in selected_keys:
            continue

        item = item_map.get(key)
        if not item:
            continue

        model = item["model"]
        rows = rows_by_dataset.get(key, []) or []
        created = 0
        updated = 0
        skipped = 0
        errors = 0

        for row_index, row in enumerate(rows, start=1):
            if not isinstance(row, dict):
                skipped += 1
                errors += 1
                summary["row_errors"].append(
                    {
                        "dataset": key,
                        "row": row_index,
                        "id": None,
                        "error": "Строка имеет неверный формат (ожидается объект JSON)",
                    }
                )
                continue

            payload_data = dict(row)
            pk = payload_data.pop("id", None)

            try:
                if pk is None:
                    model.objects.create(**payload_data)
                    created += 1
                    continue

                _, was_created = model.objects.update_or_create(
                    pk=pk,
                    defaults=payload_data,
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
            except Exception as exc:
                skipped += 1
                errors += 1
                summary["row_errors"].append(
                    {
                        "dataset": key,
                        "row": row_index,
                        "id": pk,
                        "error": str(exc),
                    }
                )

        summary["datasets"][key] = {
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
            "rows": len(rows),
        }
        summary["created"] += created
        summary["updated"] += updated
        summary["skipped"] += skipped
        summary["errors"] += errors

    return summary


def _create_service_exchange_log(
    request: HttpRequest,
    *,
    action: str,
    section: str,
    result_status: str,
    selected_keys: list[str],
    dry_run: bool = False,
    file_name: str = "",
    summary: dict | None = None,
    message: str = "",
):
    return ServiceExchangeLog.objects.create(
        user=request.user if request.user.is_authenticated else None,
        action=action,
        section=section,
        result_status=result_status,
        dry_run=dry_run,
        file_name=file_name,
        selected_keys=selected_keys,
        summary=summary or {},
        message=message,
    )


@login_required(login_url="login")
def service_exchange_log_download(request, log_id: int):
    log = get_object_or_404(ServiceExchangeLog, id=log_id)
    if not request.user.is_staff and log.user_id != request.user.id:
        messages.error(request, "Недостаточно прав для скачивания этого протокола.")
        return redirect(reverse("service_exchange"))

    payload = {
        "id": log.id,
        "created_at": log.created_at.isoformat(),
        "user": log.user.username if log.user_id else None,
        "action": log.action,
        "section": log.section,
        "result_status": log.result_status,
        "dry_run": log.dry_run,
        "file_name": log.file_name,
        "selected_keys": log.selected_keys,
        "summary": log.summary,
        "message": log.message,
    }
    response = HttpResponse(
        json.dumps(payload, ensure_ascii=False, indent=2, cls=DjangoJSONEncoder),
        content_type="application/json; charset=utf-8",
    )
    response["Content-Disposition"] = f'attachment; filename="service_exchange_log_{log.id}.json"'
    return response


@login_required(login_url="login")
def service_exchange(request):
    section = request.GET.get("section", request.POST.get("section", "directories")).strip()
    mode = request.GET.get("mode", request.POST.get("mode", "export")).strip()

    if section not in SERVICE_EXPORT_GROUPS:
        section = "directories"
    if mode not in {"export", "import", "cleanup", "archive"}:
        mode = "export"
    if mode == "cleanup" and not request.user.is_superuser:
        messages.error(request, "Пункт Очистка доступен только суперпользователю.")
        return redirect(f"{reverse('service_exchange')}?section=directories&mode=export")
    if mode == "archive" and not request.user.is_superuser:
        messages.error(request, "Пункт Архив базы доступен только суперпользователю.")
        return redirect(f"{reverse('service_exchange')}?section=directories&mode=export")

    group_config = SERVICE_EXPORT_GROUPS[section]
    selected_keys = [item["key"] for item in group_config["items"]]
    import_summary = None
    cleanup_summary = None
    archive_summary = None
    latest_log = None

    log_action_filter = request.GET.get("log_action", "").strip()
    log_status_filter = request.GET.get("log_status", "").strip()
    log_dry_run_filter = request.GET.get("log_dry_run", "").strip()
    log_user_filter = request.GET.get("log_user", "").strip()
    log_date_from = request.GET.get("log_date_from", "").strip()
    log_date_to = request.GET.get("log_date_to", "").strip()

    if request.method == "POST":
        action = request.POST.get("action", "")
        if action == "cleanup":
            selected_keys = _service_cleanup_selected_keys(request)
        else:
            selected_keys = _service_selected_keys(section, request)

        if action == "export":
            export_format = request.POST.get("export_format", "zip")
            payload = _service_export_payload(section, selected_keys)
            summary = {
                "datasets": {
                    key: {"rows": len(rows)}
                    for key, rows in payload["data"].items()
                }
            }
            _create_service_exchange_log(
                request,
                action=ServiceExchangeLog.Action.EXPORT,
                section=section,
                result_status=ServiceExchangeLog.ResultStatus.SUCCESS,
                selected_keys=selected_keys,
                dry_run=False,
                summary=summary,
                message=f"Р¤РѕСЂРјР°С‚: {export_format}",
            )

            if export_format == "json":
                body = json.dumps(payload, ensure_ascii=False, indent=2, cls=DjangoJSONEncoder)
                response = HttpResponse(body, content_type="application/json; charset=utf-8")
                response["Content-Disposition"] = f'attachment; filename="{section}_export.json"'
                return response

            output = BytesIO()
            with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                manifest = json.dumps(payload["manifest"], ensure_ascii=False, indent=2, cls=DjangoJSONEncoder)
                archive.writestr("manifest.json", manifest)
                for key, rows in payload["data"].items():
                    content = json.dumps(rows, ensure_ascii=False, indent=2, cls=DjangoJSONEncoder)
                    archive.writestr(f"data/{key}.json", content)

            output.seek(0)
            response = HttpResponse(output.read(), content_type="application/zip")
            response["Content-Disposition"] = f'attachment; filename="{section}_export.zip"'
            return response

        if action == "import":
            uploaded_file = request.FILES.get("import_file")
            dry_run = request.POST.get("dry_run") == "on"

            if not uploaded_file:
                messages.error(request, "Выберите файл для загрузки.")
                latest_log = _create_service_exchange_log(
                    request,
                    action=ServiceExchangeLog.Action.IMPORT,
                    section=section,
                    result_status=ServiceExchangeLog.ResultStatus.ERROR,
                    selected_keys=selected_keys,
                    dry_run=dry_run,
                    message="Р¤Р°Р№Р» РЅРµ РІС‹Р±СЂР°РЅ",
                )
            else:
                try:
                    payload = _service_parse_uploaded_payload(uploaded_file)
                    item_map = _service_item_map()
                    available_keys = set(payload.get("data", {}).keys())
                    selected_keys = [key for key in selected_keys if key in available_keys]
                    if not selected_keys:
                        messages.error(request, "В файле нет выбранных наборов данных.")
                        latest_log = _create_service_exchange_log(
                            request,
                            action=ServiceExchangeLog.Action.IMPORT,
                            section=section,
                            result_status=ServiceExchangeLog.ResultStatus.ERROR,
                            selected_keys=selected_keys,
                            dry_run=dry_run,
                            file_name=uploaded_file.name,
                            message="В файле нет выбранных наборов данных",
                        )
                    else:
                        with transaction.atomic():
                            import_summary = _import_dataset_rows(item_map, selected_keys, payload)
                            if dry_run:
                                transaction.set_rollback(True)

                        has_errors = import_summary.get("errors", 0) > 0

                        latest_log = _create_service_exchange_log(
                            request,
                            action=ServiceExchangeLog.Action.IMPORT,
                            section=section,
                            result_status=ServiceExchangeLog.ResultStatus.ERROR if has_errors else ServiceExchangeLog.ResultStatus.SUCCESS,
                            selected_keys=selected_keys,
                            dry_run=dry_run,
                            file_name=uploaded_file.name,
                            summary=import_summary,
                            message=(
                                f"Обнаружено ошибок: {import_summary.get('errors', 0)}"
                                if has_errors
                                else ("Проверка выполнена" if dry_run else "Загрузка выполнена")
                            ),
                        )

                        if dry_run:
                            messages.info(request, "Проверка выполнена. Изменения не сохранены (dry-run).")
                        elif not has_errors:
                            messages.success(request, "Загрузка выполнена успешно.")

                        if has_errors:
                            messages.warning(
                                request,
                                f"Загрузка завершена с ошибками: {import_summary.get('errors', 0)}. Подробности в протоколе.",
                            )
                except Exception as exc:
                    messages.error(request, f"Ошибка загрузки: {exc}")
                    latest_log = _create_service_exchange_log(
                        request,
                        action=ServiceExchangeLog.Action.IMPORT,
                        section=section,
                        result_status=ServiceExchangeLog.ResultStatus.ERROR,
                        selected_keys=selected_keys,
                        dry_run=dry_run,
                        file_name=uploaded_file.name if uploaded_file else "",
                        message=str(exc),
                    )

        if action == "cleanup":
            if not request.user.is_superuser:
                messages.error(request, "Пункт Очистка доступен только суперпользователю.")
            else:
                dry_run = request.POST.get("dry_run") == "on"
                confirmed = request.POST.get("confirm_cleanup") == "on"
                confirm_phrase = request.POST.get("confirm_phrase", "").strip()
                is_phrase_valid = confirm_phrase in {"ОЧИСТИТЬ", "OCHISTIT"}
                if not confirmed:
                    messages.error(request, "Подтвердите очистку флажком перед запуском.")
                    latest_log = _create_service_exchange_log(
                        request,
                        action=ServiceExchangeLog.Action.CLEANUP,
                        section=ServiceExchangeLog.Section.ALL,
                        result_status=ServiceExchangeLog.ResultStatus.ERROR,
                        selected_keys=selected_keys,
                        dry_run=dry_run,
                        message="Очистка не подтверждена",
                    )
                elif not is_phrase_valid:
                    messages.error(request, "Введите контрольную фразу ОЧИСТИТЬ для запуска очистки.")
                    latest_log = _create_service_exchange_log(
                        request,
                        action=ServiceExchangeLog.Action.CLEANUP,
                        section=ServiceExchangeLog.Section.ALL,
                        result_status=ServiceExchangeLog.ResultStatus.ERROR,
                        selected_keys=selected_keys,
                        dry_run=dry_run,
                        message="Неверная контрольная фраза",
                    )
                else:
                    try:
                        item_map = _service_item_map()
                        with transaction.atomic():
                            cleanup_summary = _service_cleanup_rows(item_map, selected_keys)
                            if dry_run:
                                transaction.set_rollback(True)

                        has_errors = cleanup_summary.get("errors", 0) > 0
                        latest_log = _create_service_exchange_log(
                            request,
                            action=ServiceExchangeLog.Action.CLEANUP,
                            section=ServiceExchangeLog.Section.ALL,
                            result_status=ServiceExchangeLog.ResultStatus.ERROR if has_errors else ServiceExchangeLog.ResultStatus.SUCCESS,
                            selected_keys=selected_keys,
                            dry_run=dry_run,
                            summary=cleanup_summary,
                            message=(
                                f"Обнаружено ошибок: {cleanup_summary.get('errors', 0)}"
                                if has_errors
                                else ("Проверка очистки выполнена" if dry_run else "Очистка выполнена")
                            ),
                        )

                        if dry_run:
                            messages.info(request, "Проверка очистки выполнена. Изменения не сохранены (dry-run).")
                        elif not has_errors:
                            messages.success(request, "Очистка выполнена успешно.")

                        if has_errors:
                            messages.warning(
                                request,
                                f"Очистка завершена с ошибками: {cleanup_summary.get('errors', 0)}. Подробности в протоколе.",
                            )
                    except Exception as exc:
                        messages.error(request, f"Ошибка очистки: {exc}")
                        latest_log = _create_service_exchange_log(
                            request,
                            action=ServiceExchangeLog.Action.CLEANUP,
                            section=ServiceExchangeLog.Section.ALL,
                            result_status=ServiceExchangeLog.ResultStatus.ERROR,
                            selected_keys=selected_keys,
                            dry_run=dry_run,
                            message=str(exc),
                        )

        if action == "archive":
            if not request.user.is_superuser:
                messages.error(request, "Пункт Архив базы доступен только суперпользователю.")
            else:
                try:
                    archive_summary = _archive_database_copy()
                    latest_log = _create_service_exchange_log(
                        request,
                        action=ServiceExchangeLog.Action.ARCHIVE,
                        section=ServiceExchangeLog.Section.ALL,
                        result_status=ServiceExchangeLog.ResultStatus.SUCCESS,
                        selected_keys=[],
                        dry_run=False,
                        file_name=archive_summary.get("target", ""),
                        summary=archive_summary,
                        message="Архив базы сохранен",
                    )
                    messages.success(request, f"Архив базы сохранен в папку arxiv: {archive_summary.get('file_name')}.")
                except Exception as exc:
                    messages.error(request, f"Ошибка архивации: {exc}")
                    latest_log = _create_service_exchange_log(
                        request,
                        action=ServiceExchangeLog.Action.ARCHIVE,
                        section=ServiceExchangeLog.Section.ALL,
                        result_status=ServiceExchangeLog.ResultStatus.ERROR,
                        selected_keys=[],
                        dry_run=False,
                        message=str(exc),
                    )

    logs_scope_qs = ServiceExchangeLog.objects.all()
    if mode == "cleanup":
        logs_scope_qs = logs_scope_qs.filter(action=ServiceExchangeLog.Action.CLEANUP)
    elif mode == "archive":
        logs_scope_qs = logs_scope_qs.filter(action=ServiceExchangeLog.Action.ARCHIVE)
    else:
        logs_scope_qs = logs_scope_qs.filter(section=section)
    if not request.user.is_staff:
        logs_scope_qs = logs_scope_qs.filter(user=request.user)

    recent_logs_qs = logs_scope_qs
    if log_action_filter in {
        ServiceExchangeLog.Action.EXPORT,
        ServiceExchangeLog.Action.IMPORT,
        ServiceExchangeLog.Action.CLEANUP,
        ServiceExchangeLog.Action.ARCHIVE,
    }:
        recent_logs_qs = recent_logs_qs.filter(action=log_action_filter)
    if log_status_filter in {ServiceExchangeLog.ResultStatus.SUCCESS, ServiceExchangeLog.ResultStatus.ERROR}:
        recent_logs_qs = recent_logs_qs.filter(result_status=log_status_filter)
    if log_dry_run_filter == "yes":
        recent_logs_qs = recent_logs_qs.filter(dry_run=True)
    elif log_dry_run_filter == "no":
        recent_logs_qs = recent_logs_qs.filter(dry_run=False)
    if log_user_filter:
        recent_logs_qs = recent_logs_qs.filter(user__username__icontains=log_user_filter)

    date_from_obj = parse_date(log_date_from) if log_date_from else None
    if date_from_obj:
        recent_logs_qs = recent_logs_qs.filter(created_at__date__gte=date_from_obj)

    date_to_obj = parse_date(log_date_to) if log_date_to else None
    if date_to_obj:
        recent_logs_qs = recent_logs_qs.filter(created_at__date__lte=date_to_obj)

    recent_logs = recent_logs_qs.select_related("user")[:50]
    ui_row_errors = (import_summary or {}).get("row_errors", [])
    ui_row_errors_preview = ui_row_errors[:SERVICE_UI_ROW_ERROR_PREVIEW_LIMIT]

    context = {
        "section": section,
        "mode": mode,
        "group_config": group_config,
        "cleanup_directories_items": SERVICE_EXPORT_GROUPS["directories"]["items"],
        "cleanup_documents_items": SERVICE_EXPORT_GROUPS["documents"]["items"],
        "selected_keys": selected_keys,
        "import_summary": import_summary,
        "cleanup_summary": cleanup_summary,
        "archive_summary": archive_summary,
        "ui_row_errors_preview_limit": SERVICE_UI_ROW_ERROR_PREVIEW_LIMIT,
        "ui_row_errors_preview": ui_row_errors_preview,
        "ui_row_errors_is_truncated": len(ui_row_errors) > SERVICE_UI_ROW_ERROR_PREVIEW_LIMIT,
        "latest_log": latest_log,
        "recent_logs": recent_logs,
        "log_action_filter": log_action_filter,
        "log_status_filter": log_status_filter,
        "log_dry_run_filter": log_dry_run_filter,
        "log_user_filter": log_user_filter,
        "log_date_from": log_date_from,
        "log_date_to": log_date_to,
    }
    return render(request, "service_exchange.html", context)


def _strip_tags(value: str) -> str:
    text = re.sub(r"<[^>]+>", "", value or "")
    text = text.replace("&nbsp;", " ")
    return re.sub(r"\s+", " ", text).strip()


def _normalize_search_term(value: str) -> str:
    return (value or "").strip().lower()


def _sanitize_filename_fragment(value: str) -> str:
    text = (value or "").strip()
    text = re.sub(r"[\\/:*?\"<>|]+", " ", text)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text).strip("._")
    return text[:40]


def _build_filtered_list_filename(prefix: str, fragments: list[tuple[str, str]]) -> str:
    cleaned = []
    for label, value in fragments:
        safe_label = _sanitize_filename_fragment(label)
        safe_value = _sanitize_filename_fragment(value)
        if safe_label and safe_value:
            cleaned.append(f"{safe_label}-{safe_value}")
    date_suffix = timezone.localdate().strftime("%Y-%m-%d")
    if cleaned:
        return f"{prefix}_{'__'.join(cleaned)[:130]}_{date_suffix}.xlsx"
    return f"{prefix}_list_{date_suffix}.xlsx"


def _set_download_filename(response: HttpResponse, filename: str) -> None:
    fallback = re.sub(r"[^A-Za-z0-9._-]+", "_", filename)
    fallback = re.sub(r"_+", "_", fallback).strip("._")
    if not fallback:
        fallback = "download.xlsx"
    encoded = quote(filename)
    response["Content-Disposition"] = f"attachment; filename=\"{fallback}\"; filename*=UTF-8''{encoded}"


def _paginate_report_queryset(request: HttpRequest, queryset, default_per_page: str = "10"):
    per_page = request.GET.get("per_page", default_per_page).strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = default_per_page

    if per_page == "all":
        return queryset, None, per_page

    paginator = Paginator(queryset, int(per_page))
    page_obj = paginator.get_page(request.GET.get("page", "1"))
    return page_obj.object_list, page_obj, per_page


def _parse_tags_value(raw_value: str) -> list[str]:
    result = []
    seen = set()
    for token in (raw_value or "").split(","):
        cleaned = token.strip()
        if not cleaned:
            continue
        lowered = cleaned.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        result.append(cleaned)
    return result


def _normalize_characteristic_input(c_type: EquipmentCharacteristicType, request: HttpRequest) -> tuple[str, str | None]:
    value_kind = c_type.value_kind

    if value_kind == EquipmentCharacteristicType.ValueKind.BOOLEAN:
        raw_bool = request.POST.get("characteristic_value_bool", "0").strip().lower()
        return ("1" if raw_bool in {"1", "true", "on", "yes"} else "0"), None

    if value_kind == EquipmentCharacteristicType.ValueKind.NUMBER:
        raw_number = request.POST.get("characteristic_value_number", "").strip().replace(",", ".")
        if not raw_number:
            return "", "Заполните числовое значение характеристики."
        try:
            parsed = Decimal(raw_number)
        except InvalidOperation:
            return "", "Введите корректное число для характеристики."
        normalized = f"{parsed:f}".rstrip("0").rstrip(".")
        return (normalized or "0"), None

    if value_kind == EquipmentCharacteristicType.ValueKind.TAGS:
        tags_raw = request.POST.get("characteristic_value_tags", "")
        tags = _parse_tags_value(tags_raw)
        if not tags:
            return "", "Добавьте хотя бы один элемент списка характеристик."
        return ", ".join(tags), None

    raw_text = request.POST.get("characteristic_value", "").strip()
    if not raw_text:
        return "", "Заполните значение характеристики."
    return raw_text, None


CHARACTERISTIC_DISPLAY_CODES = ("device_type", "color", "format_print", "speed_print", "weight", "dimensions")


def _annotate_characteristic_values(queryset, characteristic_model, owner_field_name: str, prefix: str = "char_"):
    annotations = {}
    for code in CHARACTERISTIC_DISPLAY_CODES:
        value_subquery = characteristic_model.objects.filter(
            **{owner_field_name: OuterRef("pk"), "characteristic_type__code": code}
        ).values("value")[:1]
        annotations[f"{prefix}{code}"] = Subquery(value_subquery)

    speed_subquery = characteristic_model.objects.filter(
        **{owner_field_name: OuterRef("pk"), "characteristic_type__code": "speed_print"}
    ).values("value")[:1]
    annotations[f"{prefix}speed_print_number"] = Subquery(speed_subquery, output_field=IntegerField())
    return queryset.annotate(**annotations)


def _extract_inn(inn_kpp_value: str) -> str:
    digits = re.sub(r"\D", "", inn_kpp_value or "")
    if len(digits) >= 12:
        return digits[:12]
    return digits[:10]


def _egrul_lookup_by_inn(inn: str) -> dict | None:
    payload = urlencode(
        {
            "query": inn,
            "region": "",
            "PreventChromeAutocomplete": "",
        }
    ).encode("utf-8")
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
    }

    search_request = Request("https://egrul.nalog.ru/", data=payload, headers=headers, method="POST")
    with urlopen(search_request, timeout=10) as response:
        search_data = json.loads(response.read().decode("utf-8"))

    token = search_data.get("t")
    if not token:
        return None

    result_url = f"https://egrul.nalog.ru/search-result/{token}?r={random.random()}"
    result_request = Request(result_url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(result_request, timeout=10) as response:
        result_data = json.loads(response.read().decode("utf-8"))

    rows = result_data.get("rows") or []
    if not rows:
        return None

    first = rows[0]
    address = ""
    for key in ("a", "address", "addr", "fullAddress", "АдресПолн"):
        address = _strip_tags(first.get(key, ""))
        if address:
            break

    ogrn = ""
    for key in ("o", "g", "ogrn", "РћР“Р Рќ"):
        ogrn = _strip_tags(first.get(key, ""))
        if ogrn:
            break

    return {
        "name": _strip_tags(first.get("n", "")),
        "ogrn": ogrn,
        "address": address,
    }


def _organizations_redirect_url(
    request: HttpRequest,
    edit_id: int | None = None,
    fragment: str | None = None,
) -> str:
    params = {}
    query = request.POST.get("q", request.GET.get("q", "")).strip()
    status = request.POST.get("status", request.GET.get("status", "")).strip()
    per_page = request.POST.get("per_page", request.GET.get("per_page", "")).strip().lower()
    address_query = request.POST.get("address_q", request.GET.get("address_q", "")).strip()

    if edit_id:
        params["edit"] = str(edit_id)
    if query:
        params["q"] = query
    if status:
        params["status"] = status
    if per_page in {"10", "30", "50", "100", "all"}:
        params["per_page"] = per_page
    if address_query:
        params["address_q"] = address_query

    base_url = reverse("organizations")
    if params:
        base_url = f"{base_url}?{urlencode(params)}"
    if fragment:
        base_url = f"{base_url}#{fragment}"
    return base_url


def _service_centers_redirect_url(
    request: HttpRequest,
    edit_id: int | None = None,
    fragment: str | None = None,
) -> str:
    params = {}
    query = request.POST.get("q", request.GET.get("q", "")).strip()
    status = request.POST.get("status", request.GET.get("status", "")).strip()
    per_page = request.POST.get("per_page", request.GET.get("per_page", "")).strip().lower()
    address_query = request.POST.get("address_q", request.GET.get("address_q", "")).strip()

    if edit_id:
        params["edit"] = str(edit_id)
    if query:
        params["q"] = query
    if status:
        params["status"] = status
    if per_page in {"10", "30", "50", "100", "all"}:
        params["per_page"] = per_page
    if address_query:
        params["address_q"] = address_query

    base_url = reverse("service_centers")
    if params:
        base_url = f"{base_url}?{urlencode(params)}"
    if fragment:
        base_url = f"{base_url}#{fragment}"
    return base_url


def home(request):
    return render(request, "home.html")


def home_logo(request):
    logo_path = settings.BASE_DIR / "templates" / "logo.jpg"
    if not logo_path.exists():
        return HttpResponse(status=404)
    return FileResponse(logo_path.open("rb"), content_type="image/jpeg")


@login_required(login_url="login")
def profile(request):
    return render(request, "profile.html")


@login_required(login_url="login")
def contacts(request):
    query = request.GET.get("q", "").strip()
    normalized_query = _normalize_search_term(query)
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    sort = request.GET.get("sort", "name")
    direction = request.GET.get("dir", "asc")

    sort_map = {
        "name": "name",
        "position": "position",
        "phone": "phone",
        "organization": "organization__name",
    }

    if sort not in sort_map:
        sort = "name"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"

    contacts_qs = OrganizationContact.objects.select_related("organization")
    if normalized_query:
        contacts_qs = contacts_qs.annotate(
            name_lc=Lower("name"),
            position_lc=Lower("position"),
            phone_lc=Lower("phone"),
            organization_name_lc=Lower("organization__name"),
        ).filter(
            Q(name_lc__contains=normalized_query)
            | Q(position_lc__contains=normalized_query)
            | Q(phone_lc__contains=normalized_query)
            | Q(organization_name_lc__contains=normalized_query)
        )
    contacts_qs = contacts_qs.order_by(order_field, "id")
    paged_contacts, page_obj, per_page = _paginate_report_queryset(request, contacts_qs, default_per_page=per_page)

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    context = {
        "contacts": paged_contacts,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "query": query,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "contacts.html", context)


@login_required(login_url="login")
def egrul_lookup(request):
    inn_raw = request.GET.get("inn", "")
    inn = _extract_inn(inn_raw)
    if len(inn) not in {10, 12}:
        return JsonResponse({"ok": False, "error": "Введите корректный ИНН (10 или 12 цифр)."}, status=400)

    try:
        result = _egrul_lookup_by_inn(inn)
    except Exception:
        return JsonResponse(
            {"ok": False, "error": "Не удалось получить данные ЕГРЮЛ. Попробуйте позже."},
            status=502,
        )

    if not result:
        return JsonResponse({"ok": False, "error": "По указанному ИНН данные не найдены."}, status=404)

    return JsonResponse({"ok": True, "data": result})


@login_required(login_url="login")
def organizations(request):
    organization_id = request.GET.get("edit")
    contact_edit_id = request.GET.get("contact_edit")
    query = request.GET.get("q", "").strip()
    normalized_query = _normalize_search_term(query)
    status_filter = request.GET.get("status", "").strip()
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    address_query = request.GET.get("address_q", "").strip()
    editing_organization = None
    editing_contact = None

    if organization_id:
        editing_organization = get_object_or_404(Organization, id=organization_id)
        if contact_edit_id:
            editing_contact = get_object_or_404(
                OrganizationContact,
                id=contact_edit_id,
                organization=editing_organization,
            )

    contact_form = OrganizationContactForm(instance=editing_contact)

    if request.method == "POST":
        org_id = request.POST.get("organization_id")
        action = request.POST.get("action", "save")

        if action == "delete" and org_id:
            target = get_object_or_404(Organization, id=org_id)
            target.delete()
            messages.success(request, "Организация удалена.")
            return redirect(_organizations_redirect_url(request))

        if action == "add_contact" and org_id:
            editing_organization = get_object_or_404(Organization, id=org_id)
            contact_form = OrganizationContactForm(request.POST)
            form = OrganizationForm(instance=editing_organization)
            if contact_form.is_valid():
                contact = contact_form.save(commit=False)
                contact.organization = editing_organization
                contact.save()
                messages.success(request, "Контакт добавлен.")
                return redirect(_organizations_redirect_url(request, edit_id=editing_organization.id))

        if action == "update_contact":
            contact_id = request.POST.get("contact_id")
            if contact_id:
                contact = get_object_or_404(OrganizationContact, id=contact_id)
                editing_organization = contact.organization
                editing_contact = contact
                contact_form = OrganizationContactForm(request.POST, instance=contact)
                form = OrganizationForm(instance=editing_organization)
                if contact_form.is_valid():
                    contact_form.save()
                    messages.success(request, "Контакт обновлен.")
                    return redirect(_organizations_redirect_url(request, edit_id=editing_organization.id))

        if action == "delete_contact":
            contact_id = request.POST.get("contact_id")
            if contact_id:
                contact = get_object_or_404(OrganizationContact, id=contact_id)
                organization_for_redirect = contact.organization_id
                contact.delete()
                messages.success(request, "Контакт удален.")
                return redirect(_organizations_redirect_url(request, edit_id=organization_for_redirect))

        if action == "add_address" and org_id:
            editing_organization = get_object_or_404(Organization, id=org_id)
            address_id = request.POST.get("address_id")
            if address_id:
                address = get_object_or_404(Address, id=address_id)
                OrganizationAddress.objects.get_or_create(
                    organization=editing_organization,
                    address=address,
                )
                messages.success(request, "Адрес привязан.")
            return redirect(_organizations_redirect_url(request, edit_id=editing_organization.id, fragment="addresses-section"))

        if action == "remove_address" and org_id:
            editing_organization = get_object_or_404(Organization, id=org_id)
            address_id = request.POST.get("address_id")
            if address_id:
                OrganizationAddress.objects.filter(
                    organization=editing_organization,
                    address_id=address_id,
                ).delete()
                messages.success(request, "Адрес откреплён.")
            return redirect(_organizations_redirect_url(request, edit_id=editing_organization.id, fragment="addresses-section"))

        if action == "set_main_office" and org_id:
            editing_organization = get_object_or_404(Organization, id=org_id)
            address_id = request.POST.get("address_id")
            make_main_office = request.POST.get("main_office") == "on"
            address_link = get_object_or_404(
                OrganizationAddress,
                organization=editing_organization,
                address_id=address_id,
            )
            if make_main_office:
                OrganizationAddress.objects.filter(
                    organization=editing_organization,
                    main_office=True,
                ).exclude(pk=address_link.pk).update(main_office=False)
            address_link.main_office = make_main_office
            address_link.save(update_fields=["main_office"])
            messages.success(request, "Признак главного офиса обновлён.")
            return redirect(_organizations_redirect_url(request, edit_id=editing_organization.id, fragment="addresses-section"))

        if org_id:
            editing_organization = get_object_or_404(Organization, id=org_id)
            form = OrganizationForm(request.POST, instance=editing_organization)
            is_new_organization = False
        else:
            form = OrganizationForm(request.POST)
            is_new_organization = True

        if form.is_valid():
            saved = form.save()
            if is_new_organization:
                messages.success(request, f"Организация {saved.name} создана. Теперь можно добавить контакт ниже.")
            else:
                messages.success(request, f"Организация {saved.name} сохранена.")
            return redirect(_organizations_redirect_url(request, edit_id=saved.id, fragment="contacts-section"))
    else:
        form = OrganizationForm(instance=editing_organization)

    organizations_qs = Organization.objects.all()
    if normalized_query:
        organizations_qs = organizations_qs.annotate(
            name_lc=Lower("name"),
            inn_kpp_lc=Lower("inn_kpp"),
        ).filter(Q(name_lc__contains=normalized_query) | Q(inn_kpp_lc__contains=normalized_query))
    if status_filter in {Organization.Status.ACTIVE, Organization.Status.REGISTERED}:
        organizations_qs = organizations_qs.filter(status=status_filter)
    organizations_qs = organizations_qs.order_by("name", "id")
    paged_organizations, page_obj, per_page = _paginate_report_queryset(request, organizations_qs, default_per_page=per_page)

    linked_addresses = (
        editing_organization.address_links.select_related("address")
        .order_by("-main_office", "address__locality", "address__street", "address__house", "id")
        if editing_organization
        else []
    )
    if editing_organization:
        linked_ids = editing_organization.address_links.values_list("address_id", flat=True)
        address_candidates_qs = Address.objects.exclude(id__in=linked_ids)
        if address_query:
            normalized_aq = _normalize_search_term(address_query)
            address_candidates_qs = address_candidates_qs.annotate(
                locality_lc=Lower("locality"),
                street_lc=Lower("street"),
                house_lc=Lower("house"),
            ).filter(
                Q(locality_lc__contains=normalized_aq)
                | Q(street_lc__contains=normalized_aq)
                | Q(house_lc__contains=normalized_aq)
            )
        address_candidates = address_candidates_qs.order_by("locality", "street", "house")
    else:
        address_candidates = Address.objects.none()

    context = {
        "form": form,
        "contact_form": contact_form,
        "editing_organization": editing_organization,
        "editing_contact": editing_contact,
        "contacts": editing_organization.contacts.all() if editing_organization else [],
        "linked_addresses": linked_addresses,
        "address_candidates": address_candidates,
        "address_query": address_query,
        "organizations": paged_organizations,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "query": query,
        "status_filter": status_filter,
        "status_choices": Organization.Status.choices,
    }
    return render(request, "organizations.html", context)


@login_required(login_url="login")
def service_centers(request):
    center_id = request.GET.get("edit")
    contact_edit_id = request.GET.get("contact_edit")
    query = request.GET.get("q", "").strip()
    normalized_query = _normalize_search_term(query)
    status_filter = request.GET.get("status", "").strip()
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    address_query = request.GET.get("address_q", "").strip()
    editing_service_center = None
    editing_contact = None

    if center_id:
        editing_service_center = get_object_or_404(ServiceCenter, id=center_id)
        if contact_edit_id:
            editing_contact = get_object_or_404(
                ServiceCenterContact,
                id=contact_edit_id,
                service_center=editing_service_center,
            )

    contact_form = ServiceCenterContactForm(instance=editing_contact)

    if request.method == "POST":
        service_center_id = request.POST.get("service_center_id")
        action = request.POST.get("action", "save")

        if action == "delete" and service_center_id:
            target = get_object_or_404(ServiceCenter, id=service_center_id)
            target.delete()
            messages.success(request, "Сервисный центр удален.")
            return redirect(_service_centers_redirect_url(request))

        if action == "add_contact" and service_center_id:
            editing_service_center = get_object_or_404(ServiceCenter, id=service_center_id)
            contact_form = ServiceCenterContactForm(request.POST)
            form = ServiceCenterForm(instance=editing_service_center)
            if contact_form.is_valid():
                contact = contact_form.save(commit=False)
                contact.service_center = editing_service_center
                contact.save()
                messages.success(request, "Контакт добавлен.")
                return redirect(_service_centers_redirect_url(request, edit_id=editing_service_center.id))

        if action == "update_contact":
            contact_id = request.POST.get("contact_id")
            if contact_id:
                contact = get_object_or_404(ServiceCenterContact, id=contact_id)
                editing_service_center = contact.service_center
                editing_contact = contact
                contact_form = ServiceCenterContactForm(request.POST, instance=contact)
                form = ServiceCenterForm(instance=editing_service_center)
                if contact_form.is_valid():
                    contact_form.save()
                    messages.success(request, "Контакт обновлен.")
                    return redirect(_service_centers_redirect_url(request, edit_id=editing_service_center.id))

        if action == "delete_contact":
            contact_id = request.POST.get("contact_id")
            if contact_id:
                contact = get_object_or_404(ServiceCenterContact, id=contact_id)
                service_center_for_redirect = contact.service_center_id
                contact.delete()
                messages.success(request, "Контакт удален.")
                return redirect(_service_centers_redirect_url(request, edit_id=service_center_for_redirect))

        if action == "add_address" and service_center_id:
            editing_service_center = get_object_or_404(ServiceCenter, id=service_center_id)
            address_id = request.POST.get("address_id")
            if address_id:
                address = get_object_or_404(Address, id=address_id)
                ServiceCenterAddress.objects.get_or_create(
                    service_center=editing_service_center,
                    address=address,
                )
                messages.success(request, "Адрес привязан.")
            return redirect(_service_centers_redirect_url(request, edit_id=editing_service_center.id, fragment="addresses-section"))

        if action == "remove_address" and service_center_id:
            editing_service_center = get_object_or_404(ServiceCenter, id=service_center_id)
            address_id = request.POST.get("address_id")
            if address_id:
                ServiceCenterAddress.objects.filter(
                    service_center=editing_service_center,
                    address_id=address_id,
                ).delete()
                messages.success(request, "Адрес откреплён.")
            return redirect(_service_centers_redirect_url(request, edit_id=editing_service_center.id, fragment="addresses-section"))

        if action == "set_main_office" and service_center_id:
            editing_service_center = get_object_or_404(ServiceCenter, id=service_center_id)
            address_id = request.POST.get("address_id")
            make_main_office = request.POST.get("main_office") == "on"
            address_link = get_object_or_404(
                ServiceCenterAddress,
                service_center=editing_service_center,
                address_id=address_id,
            )
            if make_main_office:
                ServiceCenterAddress.objects.filter(
                    service_center=editing_service_center,
                    main_office=True,
                ).exclude(pk=address_link.pk).update(main_office=False)
            address_link.main_office = make_main_office
            address_link.save(update_fields=["main_office"])
            messages.success(request, "Признак главного офиса обновлён.")
            return redirect(_service_centers_redirect_url(request, edit_id=editing_service_center.id, fragment="addresses-section"))

        if service_center_id:
            editing_service_center = get_object_or_404(ServiceCenter, id=service_center_id)
            form = ServiceCenterForm(request.POST, instance=editing_service_center)
            is_new_service_center = False
        else:
            form = ServiceCenterForm(request.POST)
            is_new_service_center = True

        if form.is_valid():
            saved = form.save()
            if is_new_service_center:
                messages.success(request, f"Сервисный центр {saved.name} создан. Теперь можно добавить контакт ниже.")
            else:
                messages.success(request, f"Сервисный центр {saved.name} сохранен.")
            return redirect(_service_centers_redirect_url(request, edit_id=saved.id, fragment="contacts-section"))
    else:
        form = ServiceCenterForm(instance=editing_service_center)

    service_centers_qs = ServiceCenter.objects.all()
    if normalized_query:
        service_centers_qs = service_centers_qs.annotate(
            name_lc=Lower("name"),
            inn_kpp_lc=Lower("inn_kpp"),
        ).filter(Q(name_lc__contains=normalized_query) | Q(inn_kpp_lc__contains=normalized_query))
    if status_filter in {ServiceCenter.Status.ACTIVE, ServiceCenter.Status.REGISTERED}:
        service_centers_qs = service_centers_qs.filter(status=status_filter)
    service_centers_qs = service_centers_qs.order_by("name", "id")
    paged_service_centers, page_obj, per_page = _paginate_report_queryset(request, service_centers_qs, default_per_page=per_page)

    linked_addresses = (
        editing_service_center.address_links.select_related("address")
        .order_by("-main_office", "address__locality", "address__street", "address__house", "id")
        if editing_service_center
        else []
    )
    if editing_service_center:
        linked_ids = editing_service_center.address_links.values_list("address_id", flat=True)
        address_candidates_qs = Address.objects.exclude(id__in=linked_ids)
        if address_query:
            normalized_aq = _normalize_search_term(address_query)
            address_candidates_qs = address_candidates_qs.annotate(
                locality_lc=Lower("locality"),
                street_lc=Lower("street"),
                house_lc=Lower("house"),
            ).filter(
                Q(locality_lc__contains=normalized_aq)
                | Q(street_lc__contains=normalized_aq)
                | Q(house_lc__contains=normalized_aq)
            )
        address_candidates = address_candidates_qs.order_by("locality", "street", "house")
    else:
        address_candidates = Address.objects.none()

    context = {
        "form": form,
        "contact_form": contact_form,
        "editing_service_center": editing_service_center,
        "editing_contact": editing_contact,
        "contacts": editing_service_center.contacts.all() if editing_service_center else [],
        "linked_addresses": linked_addresses,
        "address_candidates": address_candidates,
        "address_query": address_query,
        "service_centers": paged_service_centers,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "query": query,
        "status_filter": status_filter,
        "status_choices": ServiceCenter.Status.choices,
    }
    return render(request, "service_centers.html", context)


@user_passes_test(lambda u: u.is_staff, login_url="home")
def users(request):
    users_qs = User.objects.all().order_by("username")
    groups = Group.objects.all().order_by("name")
    create_form = AdminUserCreateForm()
    group_form = GroupCreateForm()

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "add_user":
            create_form = AdminUserCreateForm(request.POST)
            if create_form.is_valid():
                new_user = create_form.save()
                new_user.groups.set(create_form.cleaned_data["groups"])
                messages.success(request, "Пользователь успешно создан.")
                return redirect("users")

        elif action == "add_group":
            group_form = GroupCreateForm(request.POST)
            if group_form.is_valid():
                group_form.save()
                messages.success(request, "Группа успешно создана.")
                return redirect("users")

        elif action == "delete_user":
            user_id = request.POST.get("user_id")
            target_user = get_object_or_404(User, id=user_id)
            if target_user == request.user:
                messages.error(request, "Нельзя удалить текущего пользователя.")
            else:
                target_user.delete()
                messages.success(request, "Пользователь удален.")
            return redirect("users")

        elif action == "update_groups":
            user_id = request.POST.get("user_id")
            target_user = get_object_or_404(User, id=user_id)
            selected_group_ids = request.POST.getlist("groups")
            target_user.groups.set(groups.filter(id__in=selected_group_ids))
            messages.success(request, f"Группы пользователя {target_user.username} обновлены.")
            return redirect("users")

    context = {
        "create_form": create_form,
        "group_form": group_form,
        "groups": groups,
        "users": users_qs,
    }
    return render(request, "users.html", context)


@login_required(login_url="login")
def serviceman(request):
    serviceman_id = request.GET.get("edit")
    query = request.GET.get("q", "").strip()
    normalized_query = _normalize_search_term(query)
    status_filter = request.GET.get("status", "").strip()
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    sort = request.GET.get("sort", "full_name")
    direction = request.GET.get("dir", "asc")
    editing_serviceman = None

    sort_map = {
        "full_name": "full_name",
        "phone": "phone",
        "status": "status",
    }

    if sort not in sort_map:
        sort = "full_name"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"

    if serviceman_id:
        editing_serviceman = get_object_or_404(ServiceMan, id=serviceman_id)

    if request.method == "POST":
        action = request.POST.get("action", "save")
        target_id = request.POST.get("serviceman_id")
        post_query = request.POST.get("q", query).strip()
        post_status = request.POST.get("status", status_filter).strip()
        post_per_page = request.POST.get("per_page", per_page).strip().lower()
        post_sort = request.POST.get("sort", sort).strip() or "full_name"
        post_direction = request.POST.get("dir", direction).strip() or "asc"

        params = {}
        if post_query:
            params["q"] = post_query
        if post_status in {ServiceMan.Status.ACTIVE, ServiceMan.Status.DISABLED, ServiceMan.Status.DELETED}:
            params["status"] = post_status
        if post_per_page in {"10", "30", "50", "100", "all"}:
            params["per_page"] = post_per_page
        if post_sort in sort_map:
            params["sort"] = post_sort
        if post_direction in {"asc", "desc"}:
            params["dir"] = post_direction

        redirect_url = reverse("serviceman")
        if params:
            redirect_url = f"{redirect_url}?{urlencode(params)}"

        if action == "delete" and target_id:
            target = get_object_or_404(ServiceMan, id=target_id)
            target.delete()
            messages.success(request, "Сервисный инженер удален.")
            return redirect(redirect_url)

        if target_id:
            editing_serviceman = get_object_or_404(ServiceMan, id=target_id)
            form = ServiceManForm(request.POST, instance=editing_serviceman)
            is_new = False
        else:
            form = ServiceManForm(request.POST)
            is_new = True

        if form.is_valid():
            saved = form.save()
            if is_new:
                messages.success(request, f"Сервисный инженер {saved.full_name} создан.")
            else:
                messages.success(request, f"Сервисный инженер {saved.full_name} сохранен.")
            params["edit"] = saved.id
            return redirect(f"{reverse('serviceman')}?{urlencode(params)}")
    else:
        form = ServiceManForm(instance=editing_serviceman)

    servicemen_qs = ServiceMan.objects.all()
    if normalized_query:
        servicemen_qs = servicemen_qs.annotate(
            full_name_lc=Lower("full_name"),
            phone_lc=Lower("phone"),
            status_lc=Lower("status"),
        ).filter(
            Q(full_name_lc__contains=normalized_query)
            | Q(phone_lc__contains=normalized_query)
            | Q(status_lc__contains=normalized_query)
        )
    if status_filter in {ServiceMan.Status.ACTIVE, ServiceMan.Status.DISABLED, ServiceMan.Status.DELETED}:
        servicemen_qs = servicemen_qs.filter(status=status_filter)
    servicemen_qs = servicemen_qs.order_by(order_field, "id")
    paged_servicemen, page_obj, per_page = _paginate_report_queryset(request, servicemen_qs, default_per_page=per_page)

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if status_filter:
            params["status"] = status_filter
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    context = {
        "form": form,
        "editing_serviceman": editing_serviceman,
        "servicemen": paged_servicemen,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "query": query,
        "status_filter": status_filter,
        "status_choices": ServiceMan.Status.choices,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "serviceman.html", context)


@login_required(login_url="login")
def product_category(request):
    category_id = request.GET.get("edit")
    query = request.GET.get("q", "").strip()
    normalized_query = _normalize_search_term(query)
    group_filter = request.GET.get("group", "").strip()
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    sort = request.GET.get("sort", "name")
    direction = request.GET.get("dir", "asc")
    editing_category = None

    sort_map = {
        "name": "name",
        "group": "group",
    }

    if sort not in sort_map:
        sort = "name"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"

    if category_id:
        editing_category = get_object_or_404(ProductCategory, id=category_id)

    if request.method == "POST":
        action = request.POST.get("action", "save")
        target_id = request.POST.get("category_id")
        post_query = request.POST.get("q", query).strip()
        post_group = request.POST.get("group", group_filter).strip()
        post_per_page = request.POST.get("per_page", per_page).strip().lower()
        post_sort = request.POST.get("sort", sort).strip() or "name"
        post_direction = request.POST.get("dir", direction).strip() or "asc"

        params = {}
        if post_query:
            params["q"] = post_query
        if post_group:
            params["group"] = post_group
        if post_per_page in {"10", "30", "50", "100", "all"}:
            params["per_page"] = post_per_page
        if post_sort in sort_map:
            params["sort"] = post_sort
        if post_direction in {"asc", "desc"}:
            params["dir"] = post_direction

        redirect_url = reverse("product_category")
        if params:
            redirect_url = f"{redirect_url}?{urlencode(params)}"

        if action == "delete" and target_id:
            target = get_object_or_404(ProductCategory, id=target_id)
            target.delete()
            messages.success(request, "Категория товара удалена.")
            return redirect(redirect_url)

        if target_id:
            editing_category = get_object_or_404(ProductCategory, id=target_id)
            form = ProductCategoryForm(request.POST, instance=editing_category)
            is_new = False
        else:
            form = ProductCategoryForm(request.POST)
            is_new = True

        if form.is_valid():
            saved = form.save()
            if is_new:
                messages.success(request, f"Категория \"{saved.name}\" создана.")
            else:
                messages.success(request, f"Категория \"{saved.name}\" сохранена.")
            params["edit"] = saved.id
            return redirect(f"{reverse('product_category')}?{urlencode(params)}")
    else:
        form = ProductCategoryForm(instance=editing_category)

    categories_qs = ProductCategory.objects.all()
    if normalized_query:
        categories_qs = categories_qs.annotate(
            name_lc=Lower("name"),
            group_lc=Lower("group"),
        ).filter(
            Q(name_lc__contains=normalized_query)
            | Q(group_lc__contains=normalized_query)
        )
    if group_filter:
        group_filter_lc = group_filter.lower()
        categories_qs = categories_qs.annotate(
            group_lc_f=Lower("group"),
        ).filter(group_lc_f__contains=group_filter_lc)
    categories_qs = categories_qs.order_by(order_field, "id")
    paged_categories, page_obj, per_page = _paginate_report_queryset(request, categories_qs, default_per_page=per_page)

    all_groups = (
        ProductCategory.objects.exclude(group="")
        .values_list("group", flat=True)
        .distinct()
        .order_by("group")
    )

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if group_filter:
            params["group"] = group_filter
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    context = {
        "form": form,
        "editing_category": editing_category,
        "categories": paged_categories,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "all_groups": all_groups,
        "query": query,
        "group_filter": group_filter,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "product_category.html", context)


@login_required(login_url="login")
def brand(request):
    brand_id = request.GET.get("edit")
    query = request.GET.get("q", "").strip()
    normalized_query = _normalize_search_term(query)
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    sort = request.GET.get("sort", "name")
    direction = request.GET.get("dir", "asc")
    editing_brand = None

    sort_map = {"name": "name", "site": "site"}

    if sort not in sort_map:
        sort = "name"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"

    if brand_id:
        editing_brand = get_object_or_404(Brand, id=brand_id)

    if request.method == "POST":
        action = request.POST.get("action", "save")
        target_id = request.POST.get("brand_id")
        post_query = request.POST.get("q", query).strip()
        post_per_page = request.POST.get("per_page", per_page).strip().lower()
        post_sort = request.POST.get("sort", sort).strip() or "name"
        post_direction = request.POST.get("dir", direction).strip() or "asc"

        params = {}
        if post_query:
            params["q"] = post_query
        if post_per_page in {"10", "30", "50", "100", "all"}:
            params["per_page"] = post_per_page
        if post_sort in sort_map:
            params["sort"] = post_sort
        if post_direction in {"asc", "desc"}:
            params["dir"] = post_direction

        redirect_url = reverse("brand")
        if params:
            redirect_url = f"{redirect_url}?{urlencode(params)}"

        if action == "delete" and target_id:
            target = get_object_or_404(Brand, id=target_id)
            target.delete()
            messages.success(request, "Бренд удален.")
            return redirect(redirect_url)

        if target_id:
            editing_brand = get_object_or_404(Brand, id=target_id)
            form = BrandForm(request.POST, instance=editing_brand)
            is_new = False
        else:
            form = BrandForm(request.POST)
            is_new = True

        if form.is_valid():
            saved = form.save()
            if is_new:
                messages.success(request, f"Бренд \"{saved.name}\" создан.")
            else:
                messages.success(request, f"Бренд \"{saved.name}\" сохранен.")
            params["edit"] = saved.id
            return redirect(f"{reverse('brand')}?{urlencode(params)}")
    else:
        form = BrandForm(instance=editing_brand)

    brands_qs = Brand.objects.all()
    if normalized_query:
        brands_qs = brands_qs.annotate(
            name_lc=Lower("name"),
            site_lc=Lower("site"),
        ).filter(
            Q(name_lc__contains=normalized_query) | Q(site_lc__contains=normalized_query)
        )
    brands_qs = brands_qs.order_by(order_field, "id")
    paged_brands, page_obj, per_page = _paginate_report_queryset(request, brands_qs, default_per_page=per_page)

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        p = {"sort": key, "dir": next_dir}
        if query:
            p["q"] = query
        if per_page:
            p["per_page"] = per_page
        sort_links[key] = urlencode(p)

    context = {
        "form": form,
        "editing_brand": editing_brand,
        "brands": paged_brands,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "query": query,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "brand.html", context)


@login_required(login_url="login")
def equipment_characteristic_type(request):
    type_id = request.GET.get("edit")
    query = request.GET.get("q", "").strip()
    normalized_query = _normalize_search_term(query)
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    sort = request.GET.get("sort", "sort_order")
    direction = request.GET.get("dir", "asc")
    editing_type = None

    sort_map = {
        "sort_order": "sort_order",
        "code": "code",
        "name": "name",
        "value_kind": "value_kind",
    }

    if sort not in sort_map:
        sort = "sort_order"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"

    if type_id:
        editing_type = get_object_or_404(EquipmentCharacteristicType, id=type_id)

    if request.method == "POST":
        action = request.POST.get("action", "save")
        target_id = request.POST.get("type_id")
        post_query = request.POST.get("q", query).strip()
        post_per_page = request.POST.get("per_page", per_page).strip().lower()
        post_sort = request.POST.get("sort", sort).strip() or "sort_order"
        post_direction = request.POST.get("dir", direction).strip() or "asc"

        params = {}
        if post_query:
            params["q"] = post_query
        if post_per_page in {"10", "30", "50", "100", "all"}:
            params["per_page"] = post_per_page
        if post_sort in sort_map:
            params["sort"] = post_sort
        if post_direction in {"asc", "desc"}:
            params["dir"] = post_direction

        redirect_url = reverse("equipment_characteristic_type")
        if params:
            redirect_url = f"{redirect_url}?{urlencode(params)}"

        if action == "delete" and target_id:
            target = get_object_or_404(EquipmentCharacteristicType, id=target_id)
            target.delete()
            messages.success(request, "Тип характеристики удален.")
            return redirect(redirect_url)

        if target_id:
            editing_type = get_object_or_404(EquipmentCharacteristicType, id=target_id)
            form = EquipmentCharacteristicTypeForm(request.POST, instance=editing_type)
            is_new = False
        else:
            form = EquipmentCharacteristicTypeForm(request.POST)
            is_new = True

        if form.is_valid():
            saved = form.save()
            if is_new:
                messages.success(request, f"Тип характеристики \"{saved.name}\" создан.")
            else:
                messages.success(request, f"Тип характеристики \"{saved.name}\" сохранен.")
            params["edit"] = saved.id
            return redirect(f"{reverse('equipment_characteristic_type')}?{urlencode(params)}")
    else:
        form = EquipmentCharacteristicTypeForm(instance=editing_type)

    types_qs = EquipmentCharacteristicType.objects.all()
    if normalized_query:
        types_qs = types_qs.annotate(code_lc=Lower("code"), name_lc=Lower("name"), value_kind_lc=Lower("value_kind")).filter(
            Q(code_lc__contains=normalized_query)
            | Q(name_lc__contains=normalized_query)
            | Q(value_kind_lc__contains=normalized_query)
        )
    types_qs = types_qs.order_by(order_field, "id")
    paged_types, page_obj, per_page = _paginate_report_queryset(request, types_qs, default_per_page=per_page)

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        p = {"sort": key, "dir": next_dir}
        if query:
            p["q"] = query
        if per_page:
            p["per_page"] = per_page
        sort_links[key] = urlencode(p)

    context = {
        "form": form,
        "editing_type": editing_type,
        "types": paged_types,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "query": query,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "equipment_characteristic_type.html", context)


@login_required(login_url="login")
def product_model(request):
    query = request.GET.get("q", "").strip()
    normalized_query = _normalize_search_term(query)
    brand_filter = request.GET.get("brand", "").strip()
    category_filter = request.GET.get("category", "").strip()
    device_type_filter = request.GET.get("device_type", "").strip()
    attachment_filter = request.GET.get("attachments", "").strip()
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    sort = request.GET.get("sort", "name")
    direction = request.GET.get("dir", "asc")

    sort_map = {
        "name": "name",
        "site": "site",
        "category": "category__name",
        "brand": "brand__name",
        "sku": "sku",
        "device_type": "char_device_type",
        "speed_print": "char_speed_print_number",
    }

    if sort not in sort_map:
        sort = "name"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"

    models_qs = _annotate_characteristic_values(
        ProductModel.objects.select_related("brand", "category").annotate(
            attachment_count=Count("attachments", distinct=True),
        ),
        ProductModelCharacteristic,
        "product_model_id",
    )
    if normalized_query:
        models_qs = models_qs.annotate(
            name_lc=Lower("name"),
            site_lc=Lower("site"),
            sku_lc=Lower("sku"),
            brand_name_lc=Lower("brand__name"),
            category_name_lc=Lower("category__name"),
        ).filter(
            Q(name_lc__contains=normalized_query)
            | Q(site_lc__contains=normalized_query)
            | Q(sku_lc__contains=normalized_query)
            | Q(brand_name_lc__contains=normalized_query)
            | Q(category_name_lc__contains=normalized_query)
        )
    if brand_filter:
        try:
            models_qs = models_qs.filter(brand_id=int(brand_filter))
        except ValueError:
            pass
    if category_filter:
        try:
            models_qs = models_qs.filter(category_id=int(category_filter))
        except ValueError:
            pass
    if device_type_filter:
        models_qs = models_qs.filter(char_device_type=device_type_filter)
    if attachment_filter == "with":
        models_qs = models_qs.filter(attachment_count__gt=0)
    elif attachment_filter == "without":
        models_qs = models_qs.filter(attachment_count=0)
    models_qs = models_qs.order_by(order_field, "id")
    paged_models, page_obj, per_page = _paginate_report_queryset(request, models_qs, default_per_page=per_page)

    all_brands = Brand.objects.all().order_by("name")
    all_categories = ProductCategory.objects.all().order_by("group", "name")
    all_device_types = (
        ProductModelCharacteristic.objects.filter(characteristic_type__code="device_type")
        .exclude(value="")
        .values_list("value", flat=True)
        .distinct()
        .order_by("value")
    )

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        p = {"sort": key, "dir": next_dir}
        if query:
            p["q"] = query
        if brand_filter:
            p["brand"] = brand_filter
        if category_filter:
            p["category"] = category_filter
        if device_type_filter:
            p["device_type"] = device_type_filter
        if attachment_filter:
            p["attachments"] = attachment_filter
        if per_page:
            p["per_page"] = per_page
        sort_links[key] = urlencode(p)

    context = {
        "product_models": paged_models,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "all_brands": all_brands,
        "all_categories": all_categories,
        "all_device_types": all_device_types,
        "query": query,
        "brand_filter": brand_filter,
        "category_filter": category_filter,
        "device_type_filter": device_type_filter,
        "attachment_filter": attachment_filter,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "product_model.html", context)


@login_required(login_url="login")
def product_model_edit(request, model_id=None):
    editing_model = get_object_or_404(ProductModel, id=model_id) if model_id else None
    attachment_form = CatalogAttachmentForm()

    back_query = request.GET.get("q", "").strip()
    back_brand = request.GET.get("brand", "").strip()
    back_category = request.GET.get("category", "").strip()
    back_device_type = request.GET.get("device_type", "").strip()
    back_attachment = request.GET.get("attachments", "").strip()
    back_sort = request.GET.get("sort", "name").strip() or "name"
    back_dir = request.GET.get("dir", "asc").strip() or "asc"

    cons_query = request.GET.get("cons_q", "").strip()
    cons_query_lc = _normalize_search_term(cons_query)
    cons_brand_filter = request.GET.get("cons_brand", "").strip()
    cons_category_filter = request.GET.get("cons_category", "").strip()
    cons_sort = request.GET.get("cons_sort", "name").strip() or "name"
    cons_dir = request.GET.get("cons_dir", "asc").strip() or "asc"

    part_query = request.GET.get("part_q", "").strip()
    part_query_lc = _normalize_search_term(part_query)
    part_brand_filter = request.GET.get("part_brand", "").strip()
    part_category_filter = request.GET.get("part_category", "").strip()
    part_sort = request.GET.get("part_sort", "name").strip() or "name"
    part_dir = request.GET.get("part_dir", "asc").strip() or "asc"

    back_params = {}
    if back_query:
        back_params["q"] = back_query
    if back_brand:
        back_params["brand"] = back_brand
    if back_category:
        back_params["category"] = back_category
    if back_device_type:
        back_params["device_type"] = back_device_type
    if back_attachment:
        back_params["attachments"] = back_attachment
    if back_sort:
        back_params["sort"] = back_sort
    if back_dir:
        back_params["dir"] = back_dir

    cons_sort_map = {
        "name": "name",
        "brand": "brand__name",
        "category": "category__name",
        "sku": "sku",
    }
    if cons_sort not in cons_sort_map:
        cons_sort = "name"
    if cons_dir not in {"asc", "desc"}:
        cons_dir = "asc"

    part_sort_map = {
        "name": "name",
        "brand": "brand__name",
        "category": "category__name",
        "sku": "sku",
    }
    if part_sort not in part_sort_map:
        part_sort = "name"
    if part_dir not in {"asc", "desc"}:
        part_dir = "asc"

    model_characteristics = ProductModelCharacteristic.objects.none()
    characteristic_types = EquipmentCharacteristicType.objects.all().order_by("sort_order", "name", "id")

    if request.method == "POST":
        post_back_params = {}
        for key in (
            "q", "brand", "category", "device_type", "attachments", "sort", "dir",
            "cons_q", "cons_brand", "cons_category", "cons_sort", "cons_dir",
            "part_q", "part_brand", "part_category", "part_sort", "part_dir",
        ):
            value = request.POST.get(key, request.GET.get(key, "")).strip()
            if value:
                post_back_params[key] = value

        action = request.POST.get("action", "save").strip()
        if editing_model and action in {"add_characteristic", "remove_characteristic", "update_characteristic"}:
            if action == "add_characteristic":
                characteristic_type_id = request.POST.get("characteristic_type_id", "").strip()
                if not characteristic_type_id.isdigit():
                    messages.error(request, "Выберите характеристику.")
                else:
                    c_type = get_object_or_404(EquipmentCharacteristicType, id=int(characteristic_type_id))
                    value, error_message = _normalize_characteristic_input(c_type, request)
                    if error_message:
                        messages.error(request, error_message)
                    else:
                        _, created = ProductModelCharacteristic.objects.update_or_create(
                            product_model=editing_model,
                            characteristic_type=c_type,
                            defaults={"value": value},
                        )
                        if created:
                            messages.success(request, "Характеристика добавлена.")
                        else:
                            messages.success(request, "Характеристика обновлена.")

            if action == "update_characteristic":
                characteristic_id = request.POST.get("characteristic_id", "").strip()
                if not characteristic_id.isdigit():
                    messages.error(request, "Характеристика для редактирования не выбрана.")
                else:
                    characteristic = get_object_or_404(
                        ProductModelCharacteristic,
                        id=int(characteristic_id),
                        product_model=editing_model,
                    )
                    c_type = characteristic.characteristic_type
                    value, error_message = _normalize_characteristic_input(c_type, request)
                    if error_message:
                        messages.error(request, error_message)
                    else:
                        characteristic.value = value
                        characteristic.save(update_fields=["value", "updated_at"])
                        messages.success(request, "Характеристика обновлена.")

            if action == "remove_characteristic":
                characteristic_id = request.POST.get("characteristic_id", "").strip()
                if characteristic_id.isdigit():
                    characteristic = get_object_or_404(ProductModelCharacteristic, id=int(characteristic_id), product_model=editing_model)
                    characteristic.delete()
                    messages.success(request, "Характеристика удалена.")

            qs = urlencode(post_back_params)
            edit_url = reverse("product_model_edit", kwargs={"model_id": editing_model.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        if editing_model and action == "add_attachment":
            attachment_form = CatalogAttachmentForm(request.POST, request.FILES)
            if attachment_form.is_valid():
                _save_catalog_attachment("product_model", editing_model, ProductModelAttachment, attachment_form)
                messages.success(request, "Вложение добавлено.")
            else:
                messages.error(request, _first_form_error(attachment_form))

            qs = urlencode(post_back_params)
            edit_url = reverse("product_model_edit", kwargs={"model_id": editing_model.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        if editing_model and action == "remove_attachment":
            attachment_id = request.POST.get("attachment_id", "").strip()
            if attachment_id.isdigit():
                attachment = get_object_or_404(
                    ProductModelAttachment,
                    id=int(attachment_id),
                    product_model=editing_model,
                )
                attachment.delete()
                messages.success(request, "Вложение удалено.")

            qs = urlencode(post_back_params)
            edit_url = reverse("product_model_edit", kwargs={"model_id": editing_model.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        form = ProductModelForm(request.POST, instance=editing_model)
        if form.is_valid():
            saved = form.save()
            if editing_model:
                messages.success(request, f"Техника \"{saved.name}\" сохранена.")
            else:
                messages.success(request, f"Техника \"{saved.name}\" создана.")

            qs = urlencode(post_back_params)
            edit_url = reverse("product_model_edit", kwargs={"model_id": saved.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)
    else:
        form = ProductModelForm(instance=editing_model)

    back_qs = urlencode(back_params)
    back_url = reverse("product_model")
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    linked_consumables = Consumable.objects.none()
    linked_parts = Part.objects.none()
    attachments = ProductModelAttachment.objects.none()
    cons_sort_links = {}
    part_sort_links = {}
    consumable_links_url = ""
    part_links_url = ""

    if editing_model:
        attachments = editing_model.attachments.all()
        model_characteristics = list(
            ProductModelCharacteristic.objects.select_related("characteristic_type").filter(product_model=editing_model)
        )
        for item in model_characteristics:
            if item.characteristic_type.value_kind == EquipmentCharacteristicType.ValueKind.TAGS:
                item.display_tags = _parse_tags_value(item.value)
            else:
                item.display_tags = []

        linked_consumables = Consumable.objects.select_related("brand", "category").filter(
            compatibilities__product_model=editing_model
        )
        if cons_query_lc:
            linked_consumables = linked_consumables.annotate(
                name_lc=Lower("name"),
                sku_lc=Lower("sku"),
                brand_name_lc=Lower("brand__name"),
                category_name_lc=Lower("category__name"),
            ).filter(
                Q(name_lc__contains=cons_query_lc)
                | Q(sku_lc__contains=cons_query_lc)
                | Q(brand_name_lc__contains=cons_query_lc)
                | Q(category_name_lc__contains=cons_query_lc)
            )
        if cons_brand_filter:
            try:
                linked_consumables = linked_consumables.filter(brand_id=int(cons_brand_filter))
            except ValueError:
                pass
        if cons_category_filter:
            try:
                linked_consumables = linked_consumables.filter(category_id=int(cons_category_filter))
            except ValueError:
                pass

        cons_order = cons_sort_map[cons_sort]
        if cons_dir == "desc":
            cons_order = f"-{cons_order}"
        linked_consumables = linked_consumables.order_by(cons_order, "id")

        linked_parts = Part.objects.select_related("brand", "category").filter(
            compatibilities__product_model=editing_model
        )
        if part_query_lc:
            linked_parts = linked_parts.annotate(
                name_lc=Lower("name"),
                sku_lc=Lower("sku"),
                brand_name_lc=Lower("brand__name"),
                category_name_lc=Lower("category__name"),
            ).filter(
                Q(name_lc__contains=part_query_lc)
                | Q(sku_lc__contains=part_query_lc)
                | Q(brand_name_lc__contains=part_query_lc)
                | Q(category_name_lc__contains=part_query_lc)
            )
        if part_brand_filter:
            try:
                linked_parts = linked_parts.filter(brand_id=int(part_brand_filter))
            except ValueError:
                pass
        if part_category_filter:
            try:
                linked_parts = linked_parts.filter(category_id=int(part_category_filter))
            except ValueError:
                pass

        part_order = part_sort_map[part_sort]
        if part_dir == "desc":
            part_order = f"-{part_order}"
        linked_parts = linked_parts.order_by(part_order, "id")

        for key in cons_sort_map:
            next_dir = "desc" if cons_sort == key and cons_dir == "asc" else "asc"
            params = {
                "sort": back_sort,
                "dir": back_dir,
                "cons_sort": key,
                "cons_dir": next_dir,
                "part_sort": part_sort,
                "part_dir": part_dir,
            }
            if back_query:
                params["q"] = back_query
            if back_brand:
                params["brand"] = back_brand
            if back_category:
                params["category"] = back_category
            if back_device_type:
                params["device_type"] = back_device_type
            if cons_query:
                params["cons_q"] = cons_query
            if cons_brand_filter:
                params["cons_brand"] = cons_brand_filter
            if cons_category_filter:
                params["cons_category"] = cons_category_filter
            if part_query:
                params["part_q"] = part_query
            if part_brand_filter:
                params["part_brand"] = part_brand_filter
            if part_category_filter:
                params["part_category"] = part_category_filter
            cons_sort_links[key] = urlencode(params)

        for key in part_sort_map:
            next_dir = "desc" if part_sort == key and part_dir == "asc" else "asc"
            params = {
                "sort": back_sort,
                "dir": back_dir,
                "cons_sort": cons_sort,
                "cons_dir": cons_dir,
                "part_sort": key,
                "part_dir": next_dir,
            }
            if back_query:
                params["q"] = back_query
            if back_brand:
                params["brand"] = back_brand
            if back_category:
                params["category"] = back_category
            if back_device_type:
                params["device_type"] = back_device_type
            if cons_query:
                params["cons_q"] = cons_query
            if cons_brand_filter:
                params["cons_brand"] = cons_brand_filter
            if cons_category_filter:
                params["cons_category"] = cons_category_filter
            if part_query:
                params["part_q"] = part_query
            if part_brand_filter:
                params["part_brand"] = part_brand_filter
            if part_category_filter:
                params["part_category"] = part_category_filter
            part_sort_links[key] = urlencode(params)

        relation_back_params = dict(back_params)
        if cons_query:
            relation_back_params["cons_q"] = cons_query
        if cons_brand_filter:
            relation_back_params["cons_brand"] = cons_brand_filter
        if cons_category_filter:
            relation_back_params["cons_category"] = cons_category_filter
        if cons_sort:
            relation_back_params["cons_sort"] = cons_sort
        if cons_dir:
            relation_back_params["cons_dir"] = cons_dir
        if part_query:
            relation_back_params["part_q"] = part_query
        if part_brand_filter:
            relation_back_params["part_brand"] = part_brand_filter
        if part_category_filter:
            relation_back_params["part_category"] = part_category_filter
        if part_sort:
            relation_back_params["part_sort"] = part_sort
        if part_dir:
            relation_back_params["part_dir"] = part_dir

        relation_back_qs = urlencode(relation_back_params)
        consumable_links_url = reverse("product_model_consumables", kwargs={"model_id": editing_model.id})
        part_links_url = reverse("product_model_parts", kwargs={"model_id": editing_model.id})
        if relation_back_qs:
            consumable_links_url = f"{consumable_links_url}?{relation_back_qs}"
            part_links_url = f"{part_links_url}?{relation_back_qs}"

    all_brands = Brand.objects.all().order_by("name")
    all_categories = ProductCategory.objects.all().order_by("group", "name")

    context = {
        "form": form,
        "editing_model": editing_model,
        "back_url": back_url,
        "q": back_query,
        "brand": back_brand,
        "category": back_category,
        "device_type": back_device_type,
        "attachment": back_attachment,
        "sort": back_sort,
        "dir": back_dir,
        "linked_consumables": linked_consumables,
        "linked_parts": linked_parts,
        "all_brands": all_brands,
        "all_categories": all_categories,
        "cons_query": cons_query,
        "cons_brand_filter": cons_brand_filter,
        "cons_category_filter": cons_category_filter,
        "cons_sort": cons_sort,
        "cons_dir": cons_dir,
        "cons_sort_links": cons_sort_links,
        "part_query": part_query,
        "part_brand_filter": part_brand_filter,
        "part_category_filter": part_category_filter,
        "part_sort": part_sort,
        "part_dir": part_dir,
        "part_sort_links": part_sort_links,
        "consumable_links_url": consumable_links_url,
        "part_links_url": part_links_url,
        "model_characteristics": model_characteristics,
        "characteristic_types": characteristic_types,
        "attachments": attachments,
        "attachment_form": attachment_form,
    }
    return render(request, "product_model_edit.html", context)


@login_required(login_url="login")
def product_model_consumables(request, model_id):
    model = get_object_or_404(ProductModel, id=model_id)

    back_params = {}
    for key in (
        "q", "brand", "category", "device_type", "sort", "dir",
        "cons_q", "cons_brand", "cons_category", "cons_sort", "cons_dir",
        "part_q", "part_brand", "part_category", "part_sort", "part_dir",
    ):
        value = request.GET.get(key, "").strip() if request.method == "GET" else request.POST.get(key, "").strip()
        if value:
            back_params[key] = value

    relation_query = request.GET.get("relation_q", "").strip() if request.method == "GET" else request.POST.get("relation_q", "").strip()
    relation_query_lc = _normalize_search_term(relation_query)
    relation_brand = request.GET.get("relation_brand", "").strip() if request.method == "GET" else request.POST.get("relation_brand", "").strip()
    relation_category = request.GET.get("relation_category", "").strip() if request.method == "GET" else request.POST.get("relation_category", "").strip()

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add_relation":
            consumable_id = request.POST.get("consumable_id")
            if consumable_id:
                consumable = get_object_or_404(Consumable, id=consumable_id)
                _, created = ConsumableCompatibility.objects.get_or_create(consumable=consumable, product_model=model)
                if created:
                    messages.success(request, f"Связь добавлена: {consumable.name}")
                else:
                    messages.info(request, "Такая связь уже существует.")
        elif action == "remove_relation":
            relation_id = request.POST.get("relation_id")
            if relation_id:
                relation = get_object_or_404(ConsumableCompatibility, id=relation_id, product_model=model)
                relation.delete()
                messages.success(request, "Связь удалена.")

        redirect_params = dict(back_params)
        if relation_query:
            redirect_params["relation_q"] = relation_query
        if relation_brand:
            redirect_params["relation_brand"] = relation_brand
        if relation_category:
            redirect_params["relation_category"] = relation_category
        qs = urlencode(redirect_params)
        url = reverse("product_model_consumables", kwargs={"model_id": model.id})
        return redirect(f"{url}?{qs}" if qs else url)

    linked_relations = ConsumableCompatibility.objects.select_related("consumable__brand", "consumable__category").filter(product_model=model)
    linked_ids = linked_relations.values_list("consumable_id", flat=True)

    candidates = Consumable.objects.select_related("brand", "category").exclude(id__in=linked_ids)
    if relation_query_lc:
        candidates = candidates.annotate(
            name_lc=Lower("name"),
            sku_lc=Lower("sku"),
            brand_name_lc=Lower("brand__name"),
            category_name_lc=Lower("category__name"),
        ).filter(
            Q(name_lc__contains=relation_query_lc)
            | Q(sku_lc__contains=relation_query_lc)
            | Q(brand_name_lc__contains=relation_query_lc)
            | Q(category_name_lc__contains=relation_query_lc)
        )
    if relation_brand:
        try:
            candidates = candidates.filter(brand_id=int(relation_brand))
        except ValueError:
            pass
    if relation_category:
        try:
            candidates = candidates.filter(category_id=int(relation_category))
        except ValueError:
            pass
    candidates = candidates.order_by("name", "id")[:200]

    edit_back_qs = urlencode(back_params)
    back_url = reverse("product_model_edit", kwargs={"model_id": model.id})
    if edit_back_qs:
        back_url = f"{back_url}?{edit_back_qs}"

    context = {
        "editing_model": model,
        "linked_relations": linked_relations,
        "relation_candidates": candidates,
        "relation_query": relation_query,
        "relation_brand": relation_brand,
        "relation_category": relation_category,
        "all_brands": Brand.objects.all().order_by("name"),
        "all_categories": ProductCategory.objects.all().order_by("group", "name"),
        "back_url": back_url,
        "back_params": back_params,
    }
    return render(request, "product_model_consumables.html", context)


@login_required(login_url="login")
def product_model_parts(request, model_id):
    model = get_object_or_404(ProductModel, id=model_id)

    back_params = {}
    for key in (
        "q", "brand", "category", "device_type", "sort", "dir",
        "cons_q", "cons_brand", "cons_category", "cons_sort", "cons_dir",
        "part_q", "part_brand", "part_category", "part_sort", "part_dir",
    ):
        value = request.GET.get(key, "").strip() if request.method == "GET" else request.POST.get(key, "").strip()
        if value:
            back_params[key] = value

    relation_query = request.GET.get("relation_q", "").strip() if request.method == "GET" else request.POST.get("relation_q", "").strip()
    relation_query_lc = _normalize_search_term(relation_query)
    relation_brand = request.GET.get("relation_brand", "").strip() if request.method == "GET" else request.POST.get("relation_brand", "").strip()
    relation_category = request.GET.get("relation_category", "").strip() if request.method == "GET" else request.POST.get("relation_category", "").strip()

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add_relation":
            part_id = request.POST.get("part_id")
            if part_id:
                part = get_object_or_404(Part, id=part_id)
                _, created = PartCompatibility.objects.get_or_create(part=part, product_model=model)
                if created:
                    messages.success(request, f"Связь добавлена: {part.name}")
                else:
                    messages.info(request, "Такая связь уже существует.")
        elif action == "remove_relation":
            relation_id = request.POST.get("relation_id")
            if relation_id:
                relation = get_object_or_404(PartCompatibility, id=relation_id, product_model=model)
                relation.delete()
                messages.success(request, "Связь удалена.")

        redirect_params = dict(back_params)
        if relation_query:
            redirect_params["relation_q"] = relation_query
        if relation_brand:
            redirect_params["relation_brand"] = relation_brand
        if relation_category:
            redirect_params["relation_category"] = relation_category
        qs = urlencode(redirect_params)
        url = reverse("product_model_parts", kwargs={"model_id": model.id})
        return redirect(f"{url}?{qs}" if qs else url)

    linked_relations = PartCompatibility.objects.select_related("part__brand", "part__category").filter(product_model=model)
    linked_ids = linked_relations.values_list("part_id", flat=True)

    candidates = Part.objects.select_related("brand", "category").exclude(id__in=linked_ids)
    if relation_query_lc:
        candidates = candidates.annotate(
            name_lc=Lower("name"),
            sku_lc=Lower("sku"),
            brand_name_lc=Lower("brand__name"),
            category_name_lc=Lower("category__name"),
        ).filter(
            Q(name_lc__contains=relation_query_lc)
            | Q(sku_lc__contains=relation_query_lc)
            | Q(brand_name_lc__contains=relation_query_lc)
            | Q(category_name_lc__contains=relation_query_lc)
        )
    if relation_brand:
        try:
            candidates = candidates.filter(brand_id=int(relation_brand))
        except ValueError:
            pass
    if relation_category:
        try:
            candidates = candidates.filter(category_id=int(relation_category))
        except ValueError:
            pass
    candidates = candidates.order_by("name", "id")[:200]

    edit_back_qs = urlencode(back_params)
    back_url = reverse("product_model_edit", kwargs={"model_id": model.id})
    if edit_back_qs:
        back_url = f"{back_url}?{edit_back_qs}"

    context = {
        "editing_model": model,
        "linked_relations": linked_relations,
        "relation_candidates": candidates,
        "relation_query": relation_query,
        "relation_brand": relation_brand,
        "relation_category": relation_category,
        "all_brands": Brand.objects.all().order_by("name"),
        "all_categories": ProductCategory.objects.all().order_by("group", "name"),
        "back_url": back_url,
        "back_params": back_params,
    }
    return render(request, "product_model_parts.html", context)


@login_required(login_url="login")
def brand_delete(request, brand_id):
    item = get_object_or_404(Brand, id=brand_id)

    back_params = {}
    for key in ("q", "per_page", "sort", "dir"):
        value = request.GET.get(key, "").strip() if request.method == "GET" else request.POST.get(key, "").strip()
        if value:
            back_params[key] = value

    back_qs = urlencode(back_params)
    back_url = reverse("brand")
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    if request.method == "POST":
        item_name = item.name
        item.delete()
        messages.success(request, f"Бренд \"{item_name}\" удален.")
        return redirect(back_url)

    context = {
        "item": item,
        "back_url": back_url,
        "title": "Удаление бренда",
        "object_label": "бренд",
        "q": back_params.get("q", ""),
        "sort": back_params.get("sort", "name"),
        "dir": back_params.get("dir", "asc"),
    }
    return render(request, "dictionary_delete.html", context)


@login_required(login_url="login")
def product_category_delete(request, category_id):
    item = get_object_or_404(ProductCategory, id=category_id)

    back_params = {}
    for key in ("q", "group", "per_page", "sort", "dir"):
        value = request.GET.get(key, "").strip() if request.method == "GET" else request.POST.get(key, "").strip()
        if value:
            back_params[key] = value

    back_qs = urlencode(back_params)
    back_url = reverse("product_category")
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    if request.method == "POST":
        item_name = item.name
        item.delete()
        messages.success(request, f"Категория \"{item_name}\" удалена.")
        return redirect(back_url)

    context = {
        "item": item,
        "back_url": back_url,
        "title": "Удаление категории товара",
        "object_label": "категорию",
        "q": back_params.get("q", ""),
        "group": back_params.get("group", ""),
        "sort": back_params.get("sort", "name"),
        "dir": back_params.get("dir", "asc"),
    }
    return render(request, "dictionary_delete.html", context)


@login_required(login_url="login")
def product_model_delete(request, model_id):
    item = get_object_or_404(ProductModel, id=model_id)

    back_params = {}
    for key in ("q", "brand", "category", "device_type", "attachments", "per_page", "sort", "dir"):
        value = request.GET.get(key, "").strip() if request.method == "GET" else request.POST.get(key, "").strip()
        if value:
            back_params[key] = value

    back_qs = urlencode(back_params)
    back_url = reverse("product_model")
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    if request.method == "POST":
        item_name = item.name
        item.delete()
        messages.success(request, f"Техника \"{item_name}\" удалена.")
        return redirect(back_url)

    context = {
        "item": item,
        "back_url": back_url,
        "title": "Удаление техники",
        "object_label": "технику",
        "q": back_params.get("q", ""),
        "brand": back_params.get("brand", ""),
        "category": back_params.get("category", ""),
        "device_type": back_params.get("device_type", ""),
        "attachment_filter": back_params.get("attachments", ""),
        "sort": back_params.get("sort", "name"),
        "dir": back_params.get("dir", "asc"),
    }
    return render(request, "dictionary_delete.html", context)


@login_required(login_url="login")
def consumable(request):
    consumable_id = request.GET.get("edit")
    query = request.GET.get("q", "").strip()
    normalized_query = _normalize_search_term(query)
    brand_filter = request.GET.get("brand", "").strip()
    category_filter = request.GET.get("category", "").strip()
    device_type_filter = request.GET.get("device_type", "").strip()
    attachment_filter = request.GET.get("attachments", "").strip()
    relation_query = request.GET.get("relation_q", "").strip()
    relation_query_lc = _normalize_search_term(relation_query)
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    sort = request.GET.get("sort", "name")
    direction = request.GET.get("dir", "asc")
    editing_consumable = None

    sort_map = {
        "name": "name",
        "site": "site",
        "category": "category__name",
        "brand": "brand__name",
        "sku": "sku",
        "device_type": "char_device_type",
        "speed_print": "char_speed_print_number",
    }

    if sort not in sort_map:
        sort = "name"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"

    if consumable_id:
        editing_consumable = get_object_or_404(Consumable, id=consumable_id)

    if request.method == "POST":
        action = request.POST.get("action", "save")
        target_id = request.POST.get("consumable_id")
        post_query = request.POST.get("q", query).strip()
        post_brand = request.POST.get("brand", brand_filter).strip()
        post_category = request.POST.get("category", category_filter).strip()
        post_device_type = request.POST.get("device_type", device_type_filter).strip()
        post_attachment_filter = request.POST.get("attachments", attachment_filter).strip()
        post_relation_q = request.POST.get("relation_q", relation_query).strip()
        post_per_page = request.POST.get("per_page", per_page).strip().lower()
        post_sort = request.POST.get("sort", sort).strip() or "name"
        post_direction = request.POST.get("dir", direction).strip() or "asc"

        params = {}
        if post_query:
            params["q"] = post_query
        if post_brand:
            params["brand"] = post_brand
        if post_category:
            params["category"] = post_category
        if post_device_type:
            params["device_type"] = post_device_type
        if post_attachment_filter in {"with", "without"}:
            params["attachments"] = post_attachment_filter
        if post_relation_q:
            params["relation_q"] = post_relation_q
        if post_per_page in {"10", "30", "50", "100", "all"}:
            params["per_page"] = post_per_page
        if post_sort in sort_map:
            params["sort"] = post_sort
        if post_direction in {"asc", "desc"}:
            params["dir"] = post_direction

        redirect_url = reverse("consumable")
        if params:
            redirect_url = f"{redirect_url}?{urlencode(params)}"

        if action == "delete" and target_id:
            target = get_object_or_404(Consumable, id=target_id)
            if target.compatibilities.exists():
                messages.error(request, "Нельзя удалить расходный материал, пока есть связи с техникой.")
                return redirect(redirect_url)
            target.delete()
            messages.success(request, "Расходный материал удален.")
            return redirect(redirect_url)

        if action == "add_relation" and target_id:
            editing_consumable = get_object_or_404(Consumable, id=target_id)
            product_model_id = request.POST.get("product_model_id")
            if product_model_id:
                product_model = get_object_or_404(ProductModel, id=product_model_id)
                _, created = ConsumableCompatibility.objects.get_or_create(
                    consumable=editing_consumable,
                    product_model=product_model,
                )
                if created:
                    messages.success(request, f"Связь добавлена: {product_model.name}")
                else:
                    messages.info(request, "Такая связь уже существует.")
            params["edit"] = editing_consumable.id
            return redirect(f"{reverse('consumable')}?{urlencode(params)}")

        if action == "remove_relation" and target_id:
            editing_consumable = get_object_or_404(Consumable, id=target_id)
            relation_id = request.POST.get("relation_id")
            if relation_id:
                relation = get_object_or_404(ConsumableCompatibility, id=relation_id, consumable=editing_consumable)
                relation.delete()
                messages.success(request, "Связь удалена.")
            params["edit"] = editing_consumable.id
            return redirect(f"{reverse('consumable')}?{urlencode(params)}")

        if target_id:
            editing_consumable = get_object_or_404(Consumable, id=target_id)
            form = ConsumableForm(request.POST, instance=editing_consumable)
            is_new = False
        else:
            form = ConsumableForm(request.POST)
            is_new = True

        if form.is_valid():
            saved = form.save()
            if is_new:
                messages.success(request, f"Расходный материал «{saved.name}» создан.")
            else:
                messages.success(request, f"Расходный материал «{saved.name}» сохранен.")
            params["edit"] = saved.id
            return redirect(f"{reverse('consumable')}?{urlencode(params)}")
    else:
        form = ConsumableForm(instance=editing_consumable)

    consumables_qs = _annotate_characteristic_values(
        Consumable.objects.select_related("brand", "category").annotate(
            attachment_count=Count("attachments", distinct=True),
        ),
        ConsumableCharacteristic,
        "consumable_id",
    )
    if normalized_query:
        consumables_qs = consumables_qs.annotate(
            name_lc=Lower("name"),
            site_lc=Lower("site"),
            sku_lc=Lower("sku"),
            brand_name_lc=Lower("brand__name"),
            category_name_lc=Lower("category__name"),
        ).filter(
            Q(name_lc__contains=normalized_query)
            | Q(site_lc__contains=normalized_query)
            | Q(sku_lc__contains=normalized_query)
            | Q(brand_name_lc__contains=normalized_query)
            | Q(category_name_lc__contains=normalized_query)
        )
    if brand_filter:
        try:
            consumables_qs = consumables_qs.filter(brand_id=int(brand_filter))
        except ValueError:
            pass
    if category_filter:
        try:
            consumables_qs = consumables_qs.filter(category_id=int(category_filter))
        except ValueError:
            pass
    if device_type_filter:
        consumables_qs = consumables_qs.filter(char_device_type=device_type_filter)
    if attachment_filter == "with":
        consumables_qs = consumables_qs.filter(attachment_count__gt=0)
    elif attachment_filter == "without":
        consumables_qs = consumables_qs.filter(attachment_count=0)
    consumables_qs = consumables_qs.order_by(order_field, "id")
    paged_consumables, page_obj, per_page = _paginate_report_queryset(request, consumables_qs, default_per_page=per_page)

    linked_relations = ConsumableCompatibility.objects.none()
    relation_candidates = ProductModel.objects.none()
    if editing_consumable:
        linked_relations = ConsumableCompatibility.objects.select_related(
            "product_model__brand",
            "product_model__category",
        ).filter(consumable=editing_consumable)

        linked_ids = linked_relations.values_list("product_model_id", flat=True)
        relation_candidates = ProductModel.objects.select_related("brand", "category").exclude(id__in=linked_ids)
        if relation_query_lc:
            relation_candidates = relation_candidates.annotate(
                name_lc=Lower("name"),
                sku_lc=Lower("sku"),
                brand_name_lc=Lower("brand__name"),
                category_name_lc=Lower("category__name"),
            ).filter(
                Q(name_lc__contains=relation_query_lc)
                | Q(sku_lc__contains=relation_query_lc)
                | Q(brand_name_lc__contains=relation_query_lc)
                | Q(category_name_lc__contains=relation_query_lc)
            )
        relation_candidates = relation_candidates.order_by("name", "id")[:100]

    all_brands = Brand.objects.all().order_by("name")
    all_categories = ProductCategory.objects.all().order_by("group", "name")
    all_device_types = (
        ConsumableCharacteristic.objects.filter(characteristic_type__code="device_type")
        .exclude(value="")
        .values_list("value", flat=True)
        .distinct()
        .order_by("value")
    )

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        p = {"sort": key, "dir": next_dir}
        if query:
            p["q"] = query
        if brand_filter:
            p["brand"] = brand_filter
        if category_filter:
            p["category"] = category_filter
        if device_type_filter:
            p["device_type"] = device_type_filter
        if attachment_filter:
            p["attachments"] = attachment_filter
        if relation_query:
            p["relation_q"] = relation_query
        if editing_consumable:
            p["edit"] = editing_consumable.id
        if per_page:
            p["per_page"] = per_page
        sort_links[key] = urlencode(p)

    context = {
        "form": form,
        "editing_consumable": editing_consumable,
        "consumables": paged_consumables,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "linked_relations": linked_relations,
        "relation_candidates": relation_candidates,
        "all_brands": all_brands,
        "all_categories": all_categories,
        "all_device_types": all_device_types,
        "query": query,
        "relation_query": relation_query,
        "brand_filter": brand_filter,
        "category_filter": category_filter,
        "device_type_filter": device_type_filter,
        "attachment_filter": attachment_filter,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "consumable.html", context)


@login_required(login_url="login")
def part(request):
    part_id = request.GET.get("edit")
    query = request.GET.get("q", "").strip()
    normalized_query = _normalize_search_term(query)
    brand_filter = request.GET.get("brand", "").strip()
    category_filter = request.GET.get("category", "").strip()
    device_type_filter = request.GET.get("device_type", "").strip()
    attachment_filter = request.GET.get("attachments", "").strip()
    relation_query = request.GET.get("relation_q", "").strip()
    relation_query_lc = _normalize_search_term(relation_query)
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    sort = request.GET.get("sort", "name")
    direction = request.GET.get("dir", "asc")
    editing_part = None

    sort_map = {
        "name": "name",
        "site": "site",
        "category": "category__name",
        "brand": "brand__name",
        "sku": "sku",
        "device_type": "char_device_type",
        "speed_print": "char_speed_print_number",
    }

    if sort not in sort_map:
        sort = "name"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"

    if part_id:
        editing_part = get_object_or_404(Part, id=part_id)

    if request.method == "POST":
        action = request.POST.get("action", "save")
        target_id = request.POST.get("part_id")
        post_query = request.POST.get("q", query).strip()
        post_brand = request.POST.get("brand", brand_filter).strip()
        post_category = request.POST.get("category", category_filter).strip()
        post_device_type = request.POST.get("device_type", device_type_filter).strip()
        post_attachment_filter = request.POST.get("attachments", attachment_filter).strip()
        post_relation_q = request.POST.get("relation_q", relation_query).strip()
        post_per_page = request.POST.get("per_page", per_page).strip().lower()
        post_sort = request.POST.get("sort", sort).strip() or "name"
        post_direction = request.POST.get("dir", direction).strip() or "asc"

        params = {}
        if post_query:
            params["q"] = post_query
        if post_brand:
            params["brand"] = post_brand
        if post_category:
            params["category"] = post_category
        if post_device_type:
            params["device_type"] = post_device_type
        if post_attachment_filter in {"with", "without"}:
            params["attachments"] = post_attachment_filter
        if post_relation_q:
            params["relation_q"] = post_relation_q
        if post_per_page in {"10", "30", "50", "100", "all"}:
            params["per_page"] = post_per_page
        if post_sort in sort_map:
            params["sort"] = post_sort
        if post_direction in {"asc", "desc"}:
            params["dir"] = post_direction

        redirect_url = reverse("part")
        if params:
            redirect_url = f"{redirect_url}?{urlencode(params)}"

        if action == "delete" and target_id:
            target = get_object_or_404(Part, id=target_id)
            if target.compatibilities.exists():
                messages.error(request, "Нельзя удалить запчасть, пока есть связи с техникой.")
                return redirect(redirect_url)
            target.delete()
            messages.success(request, "Запчасть удалена.")
            return redirect(redirect_url)

        if action == "add_relation" and target_id:
            editing_part = get_object_or_404(Part, id=target_id)
            product_model_id = request.POST.get("product_model_id")
            if product_model_id:
                product_model = get_object_or_404(ProductModel, id=product_model_id)
                _, created = PartCompatibility.objects.get_or_create(
                    part=editing_part,
                    product_model=product_model,
                )
                if created:
                    messages.success(request, f"Связь добавлена: {product_model.name}")
                else:
                    messages.info(request, "Такая связь уже существует.")
            params["edit"] = editing_part.id
            return redirect(f"{reverse('part')}?{urlencode(params)}")

        if action == "remove_relation" and target_id:
            editing_part = get_object_or_404(Part, id=target_id)
            relation_id = request.POST.get("relation_id")
            if relation_id:
                relation = get_object_or_404(PartCompatibility, id=relation_id, part=editing_part)
                relation.delete()
                messages.success(request, "Связь удалена.")
            params["edit"] = editing_part.id
            return redirect(f"{reverse('part')}?{urlencode(params)}")

        if target_id:
            editing_part = get_object_or_404(Part, id=target_id)
            form = PartForm(request.POST, instance=editing_part)
            is_new = False
        else:
            form = PartForm(request.POST)
            is_new = True

        if form.is_valid():
            saved = form.save()
            if is_new:
                messages.success(request, f"Запчасть \"{saved.name}\" создана.")
            else:
                messages.success(request, f"Запчасть \"{saved.name}\" сохранена.")
            params["edit"] = saved.id
            return redirect(f"{reverse('part')}?{urlencode(params)}")
    else:
        form = PartForm(instance=editing_part)

    parts_qs = _annotate_characteristic_values(
        Part.objects.select_related("brand", "category").annotate(
            attachment_count=Count("attachments", distinct=True),
        ),
        PartCharacteristic,
        "part_id",
    )
    if normalized_query:
        parts_qs = parts_qs.annotate(
            name_lc=Lower("name"),
            site_lc=Lower("site"),
            sku_lc=Lower("sku"),
            brand_name_lc=Lower("brand__name"),
            category_name_lc=Lower("category__name"),
        ).filter(
            Q(name_lc__contains=normalized_query)
            | Q(site_lc__contains=normalized_query)
            | Q(sku_lc__contains=normalized_query)
            | Q(brand_name_lc__contains=normalized_query)
            | Q(category_name_lc__contains=normalized_query)
        )
    if brand_filter:
        try:
            parts_qs = parts_qs.filter(brand_id=int(brand_filter))
        except ValueError:
            pass
    if category_filter:
        try:
            parts_qs = parts_qs.filter(category_id=int(category_filter))
        except ValueError:
            pass
    if device_type_filter:
        parts_qs = parts_qs.filter(char_device_type=device_type_filter)
    if attachment_filter == "with":
        parts_qs = parts_qs.filter(attachment_count__gt=0)
    elif attachment_filter == "without":
        parts_qs = parts_qs.filter(attachment_count=0)
    parts_qs = parts_qs.order_by(order_field, "id")
    paged_parts, page_obj, per_page = _paginate_report_queryset(request, parts_qs, default_per_page=per_page)

    linked_relations = PartCompatibility.objects.none()
    relation_candidates = ProductModel.objects.none()
    if editing_part:
        linked_relations = PartCompatibility.objects.select_related(
            "product_model__brand",
            "product_model__category",
        ).filter(part=editing_part)

        linked_ids = linked_relations.values_list("product_model_id", flat=True)
        relation_candidates = ProductModel.objects.select_related("brand", "category").exclude(id__in=linked_ids)
        if relation_query_lc:
            relation_candidates = relation_candidates.annotate(
                name_lc=Lower("name"),
                sku_lc=Lower("sku"),
                brand_name_lc=Lower("brand__name"),
                category_name_lc=Lower("category__name"),
            ).filter(
                Q(name_lc__contains=relation_query_lc)
                | Q(sku_lc__contains=relation_query_lc)
                | Q(brand_name_lc__contains=relation_query_lc)
                | Q(category_name_lc__contains=relation_query_lc)
            )
        relation_candidates = relation_candidates.order_by("name", "id")[:100]

    all_brands = Brand.objects.all().order_by("name")
    all_categories = ProductCategory.objects.all().order_by("group", "name")
    all_device_types = (
        PartCharacteristic.objects.filter(characteristic_type__code="device_type")
        .exclude(value="")
        .values_list("value", flat=True)
        .distinct()
        .order_by("value")
    )

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        p = {"sort": key, "dir": next_dir}
        if query:
            p["q"] = query
        if brand_filter:
            p["brand"] = brand_filter
        if category_filter:
            p["category"] = category_filter
        if device_type_filter:
            p["device_type"] = device_type_filter
        if attachment_filter:
            p["attachments"] = attachment_filter
        if relation_query:
            p["relation_q"] = relation_query
        if editing_part:
            p["edit"] = editing_part.id
        if per_page:
            p["per_page"] = per_page
        sort_links[key] = urlencode(p)

    context = {
        "form": form,
        "editing_part": editing_part,
        "parts": paged_parts,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "linked_relations": linked_relations,
        "relation_candidates": relation_candidates,
        "all_brands": all_brands,
        "all_categories": all_categories,
        "all_device_types": all_device_types,
        "query": query,
        "relation_query": relation_query,
        "brand_filter": brand_filter,
        "category_filter": category_filter,
        "device_type_filter": device_type_filter,
        "attachment_filter": attachment_filter,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "part.html", context)


@login_required(login_url="login")
def consumable_edit(request, consumable_id=None):
    editing_consumable = get_object_or_404(Consumable, id=consumable_id) if consumable_id else None
    attachment_form = CatalogAttachmentForm()

    relation_query = request.GET.get("relation_q", "").strip()
    relation_query_lc = _normalize_search_term(relation_query)

    back_query = request.GET.get("q", "").strip()
    back_brand = request.GET.get("brand", "").strip()
    back_category = request.GET.get("category", "").strip()
    back_device_type = request.GET.get("device_type", "").strip()
    back_attachment = request.GET.get("attachments", "").strip()
    back_sort = request.GET.get("sort", "name").strip() or "name"
    back_dir = request.GET.get("dir", "asc").strip() or "asc"

    back_params = {}
    if back_query:
        back_params["q"] = back_query
    if back_brand:
        back_params["brand"] = back_brand
    if back_category:
        back_params["category"] = back_category
    if back_device_type:
        back_params["device_type"] = back_device_type
    if back_attachment:
        back_params["attachments"] = back_attachment
    if back_sort:
        back_params["sort"] = back_sort
    if back_dir:
        back_params["dir"] = back_dir

    model_characteristics = ConsumableCharacteristic.objects.none()
    characteristic_types = EquipmentCharacteristicType.objects.all().order_by("sort_order", "name", "id")

    if request.method == "POST":
        action = request.POST.get("action", "save")
        target_id = request.POST.get("consumable_id")
        relation_query = request.POST.get("relation_q", relation_query).strip()

        post_back_params = {}
        for key in ("q", "brand", "category", "device_type", "attachments", "per_page", "sort", "dir"):
            value = request.POST.get(key, request.GET.get(key, "")).strip()
            if value:
                post_back_params[key] = value

        if target_id:
            editing_consumable = get_object_or_404(Consumable, id=target_id)

        if editing_consumable and action in {"add_characteristic", "remove_characteristic", "update_characteristic"}:
            if action == "add_characteristic":
                characteristic_type_id = request.POST.get("characteristic_type_id", "").strip()
                if not characteristic_type_id.isdigit():
                    messages.error(request, "Выберите характеристику.")
                else:
                    c_type = get_object_or_404(EquipmentCharacteristicType, id=int(characteristic_type_id))
                    value, error_message = _normalize_characteristic_input(c_type, request)
                    if error_message:
                        messages.error(request, error_message)
                    else:
                        _, created = ConsumableCharacteristic.objects.update_or_create(
                            consumable=editing_consumable,
                            characteristic_type=c_type,
                            defaults={"value": value},
                        )
                        if created:
                            messages.success(request, "Характеристика добавлена.")
                        else:
                            messages.success(request, "Характеристика обновлена.")

            if action == "update_characteristic":
                characteristic_id = request.POST.get("characteristic_id", "").strip()
                if not characteristic_id.isdigit():
                    messages.error(request, "Характеристика для редактирования не выбрана.")
                else:
                    characteristic = get_object_or_404(
                        ConsumableCharacteristic,
                        id=int(characteristic_id),
                        consumable=editing_consumable,
                    )
                    c_type = characteristic.characteristic_type
                    value, error_message = _normalize_characteristic_input(c_type, request)
                    if error_message:
                        messages.error(request, error_message)
                    else:
                        characteristic.value = value
                        characteristic.save(update_fields=["value", "updated_at"])
                        messages.success(request, "Характеристика обновлена.")

            if action == "remove_characteristic":
                characteristic_id = request.POST.get("characteristic_id", "").strip()
                if characteristic_id.isdigit():
                    characteristic = get_object_or_404(
                        ConsumableCharacteristic,
                        id=int(characteristic_id),
                        consumable=editing_consumable,
                    )
                    characteristic.delete()
                    messages.success(request, "Характеристика удалена.")

            edit_params = dict(post_back_params)
            if relation_query:
                edit_params["relation_q"] = relation_query
            qs = urlencode(edit_params)
            edit_url = reverse("consumable_edit", kwargs={"consumable_id": editing_consumable.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        if action == "add_relation" and editing_consumable:
            product_model_id = request.POST.get("product_model_id")
            if product_model_id:
                product_model = get_object_or_404(ProductModel, id=product_model_id)
                _, created = ConsumableCompatibility.objects.get_or_create(
                    consumable=editing_consumable,
                    product_model=product_model,
                )
                if created:
                    messages.success(request, f"Связь добавлена: {product_model.name}")
                else:
                    messages.info(request, "Такая связь уже существует.")

            edit_params = dict(post_back_params)
            if relation_query:
                edit_params["relation_q"] = relation_query
            qs = urlencode(edit_params)
            edit_url = reverse("consumable_edit", kwargs={"consumable_id": editing_consumable.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        if action == "remove_relation" and editing_consumable:
            relation_id = request.POST.get("relation_id")
            if relation_id:
                relation = get_object_or_404(ConsumableCompatibility, id=relation_id, consumable=editing_consumable)
                relation.delete()
                messages.success(request, "Связь удалена.")

            edit_params = dict(post_back_params)
            if relation_query:
                edit_params["relation_q"] = relation_query
            qs = urlencode(edit_params)
            edit_url = reverse("consumable_edit", kwargs={"consumable_id": editing_consumable.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        if action == "add_attachment" and editing_consumable:
            attachment_form = CatalogAttachmentForm(request.POST, request.FILES)
            if attachment_form.is_valid():
                _save_catalog_attachment("consumable", editing_consumable, ConsumableAttachment, attachment_form)
                messages.success(request, "Вложение добавлено.")
            else:
                messages.error(request, _first_form_error(attachment_form))

            edit_params = dict(post_back_params)
            if relation_query:
                edit_params["relation_q"] = relation_query
            qs = urlencode(edit_params)
            edit_url = reverse("consumable_edit", kwargs={"consumable_id": editing_consumable.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        if action == "remove_attachment" and editing_consumable:
            attachment_id = request.POST.get("attachment_id", "").strip()
            if attachment_id.isdigit():
                attachment = get_object_or_404(
                    ConsumableAttachment,
                    id=int(attachment_id),
                    consumable=editing_consumable,
                )
                attachment.delete()
                messages.success(request, "Вложение удалено.")

            edit_params = dict(post_back_params)
            if relation_query:
                edit_params["relation_q"] = relation_query
            qs = urlencode(edit_params)
            edit_url = reverse("consumable_edit", kwargs={"consumable_id": editing_consumable.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        form = ConsumableForm(request.POST, instance=editing_consumable)
        if form.is_valid():
            saved = form.save()
            if editing_consumable:
                messages.success(request, f"Расходный материал \"{saved.name}\" сохранен.")
            else:
                messages.success(request, f"Расходный материал \"{saved.name}\" создан.")

            edit_params = dict(post_back_params)
            if relation_query:
                edit_params["relation_q"] = relation_query
            qs = urlencode(edit_params)
            edit_url = reverse("consumable_edit", kwargs={"consumable_id": saved.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)
    else:
        form = ConsumableForm(instance=editing_consumable)

    linked_relations = ConsumableCompatibility.objects.none()
    relation_candidates = ProductModel.objects.none()
    attachments = ConsumableAttachment.objects.none()
    if editing_consumable:
        attachments = editing_consumable.attachments.all()
        model_characteristics = list(
            ConsumableCharacteristic.objects.select_related("characteristic_type").filter(consumable=editing_consumable)
        )
        for item in model_characteristics:
            if item.characteristic_type.value_kind == EquipmentCharacteristicType.ValueKind.TAGS:
                item.display_tags = _parse_tags_value(item.value)
            else:
                item.display_tags = []

        linked_relations = ConsumableCompatibility.objects.select_related(
            "product_model__brand",
            "product_model__category",
        ).filter(consumable=editing_consumable)

        linked_ids = linked_relations.values_list("product_model_id", flat=True)
        relation_candidates = ProductModel.objects.select_related("brand", "category").exclude(id__in=linked_ids)
        if relation_query_lc:
            relation_candidates = relation_candidates.annotate(
                name_lc=Lower("name"),
                sku_lc=Lower("sku"),
                brand_name_lc=Lower("brand__name"),
                category_name_lc=Lower("category__name"),
            ).filter(
                Q(name_lc__contains=relation_query_lc)
                | Q(sku_lc__contains=relation_query_lc)
                | Q(brand_name_lc__contains=relation_query_lc)
                | Q(category_name_lc__contains=relation_query_lc)
            )
        relation_candidates = relation_candidates.order_by("name", "id")[:100]

    back_qs = urlencode(back_params)
    back_url = reverse("consumable")
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    context = {
        "form": form,
        "editing_consumable": editing_consumable,
        "model_characteristics": model_characteristics,
        "characteristic_types": characteristic_types,
        "linked_relations": linked_relations,
        "relation_candidates": relation_candidates,
        "relation_query": relation_query,
        "back_url": back_url,
        "attachments": attachments,
        "attachment_form": attachment_form,
        "q": back_query,
        "brand": back_brand,
        "category": back_category,
        "device_type": back_device_type,
        "attachment": back_attachment,
        "sort": back_sort,
        "dir": back_dir,
    }
    return render(request, "consumable_edit.html", context)


@login_required(login_url="login")
def part_edit(request, part_id=None):
    editing_part = get_object_or_404(Part, id=part_id) if part_id else None
    attachment_form = CatalogAttachmentForm()

    relation_query = request.GET.get("relation_q", "").strip()
    relation_query_lc = _normalize_search_term(relation_query)

    back_query = request.GET.get("q", "").strip()
    back_brand = request.GET.get("brand", "").strip()
    back_category = request.GET.get("category", "").strip()
    back_device_type = request.GET.get("device_type", "").strip()
    back_attachment = request.GET.get("attachments", "").strip()
    back_sort = request.GET.get("sort", "name").strip() or "name"
    back_dir = request.GET.get("dir", "asc").strip() or "asc"

    back_params = {}
    if back_query:
        back_params["q"] = back_query
    if back_brand:
        back_params["brand"] = back_brand
    if back_category:
        back_params["category"] = back_category
    if back_device_type:
        back_params["device_type"] = back_device_type
    if back_attachment:
        back_params["attachments"] = back_attachment
    if back_sort:
        back_params["sort"] = back_sort
    if back_dir:
        back_params["dir"] = back_dir

    model_characteristics = PartCharacteristic.objects.none()
    characteristic_types = EquipmentCharacteristicType.objects.all().order_by("sort_order", "name", "id")

    if request.method == "POST":
        action = request.POST.get("action", "save")
        target_id = request.POST.get("part_id")
        relation_query = request.POST.get("relation_q", relation_query).strip()

        post_back_params = {}
        for key in ("q", "brand", "category", "device_type", "attachments", "per_page", "sort", "dir"):
            value = request.POST.get(key, request.GET.get(key, "")).strip()
            if value:
                post_back_params[key] = value

        if target_id:
            editing_part = get_object_or_404(Part, id=target_id)

        if editing_part and action in {"add_characteristic", "remove_characteristic", "update_characteristic"}:
            if action == "add_characteristic":
                characteristic_type_id = request.POST.get("characteristic_type_id", "").strip()
                if not characteristic_type_id.isdigit():
                    messages.error(request, "Выберите характеристику.")
                else:
                    c_type = get_object_or_404(EquipmentCharacteristicType, id=int(characteristic_type_id))
                    value, error_message = _normalize_characteristic_input(c_type, request)
                    if error_message:
                        messages.error(request, error_message)
                    else:
                        _, created = PartCharacteristic.objects.update_or_create(
                            part=editing_part,
                            characteristic_type=c_type,
                            defaults={"value": value},
                        )
                        if created:
                            messages.success(request, "Характеристика добавлена.")
                        else:
                            messages.success(request, "Характеристика обновлена.")

            if action == "update_characteristic":
                characteristic_id = request.POST.get("characteristic_id", "").strip()
                if not characteristic_id.isdigit():
                    messages.error(request, "Характеристика для редактирования не выбрана.")
                else:
                    characteristic = get_object_or_404(
                        PartCharacteristic,
                        id=int(characteristic_id),
                        part=editing_part,
                    )
                    c_type = characteristic.characteristic_type
                    value, error_message = _normalize_characteristic_input(c_type, request)
                    if error_message:
                        messages.error(request, error_message)
                    else:
                        characteristic.value = value
                        characteristic.save(update_fields=["value", "updated_at"])
                        messages.success(request, "Характеристика обновлена.")

            if action == "remove_characteristic":
                characteristic_id = request.POST.get("characteristic_id", "").strip()
                if characteristic_id.isdigit():
                    characteristic = get_object_or_404(
                        PartCharacteristic,
                        id=int(characteristic_id),
                        part=editing_part,
                    )
                    characteristic.delete()
                    messages.success(request, "Характеристика удалена.")

            edit_params = dict(post_back_params)
            if relation_query:
                edit_params["relation_q"] = relation_query
            qs = urlencode(edit_params)
            edit_url = reverse("part_edit", kwargs={"part_id": editing_part.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        if action == "add_relation" and editing_part:
            product_model_id = request.POST.get("product_model_id")
            if product_model_id:
                product_model = get_object_or_404(ProductModel, id=product_model_id)
                _, created = PartCompatibility.objects.get_or_create(
                    part=editing_part,
                    product_model=product_model,
                )
                if created:
                    messages.success(request, f"Связь добавлена: {product_model.name}")
                else:
                    messages.info(request, "Такая связь уже существует.")

            edit_params = dict(post_back_params)
            if relation_query:
                edit_params["relation_q"] = relation_query
            qs = urlencode(edit_params)
            edit_url = reverse("part_edit", kwargs={"part_id": editing_part.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        if action == "remove_relation" and editing_part:
            relation_id = request.POST.get("relation_id")
            if relation_id:
                relation = get_object_or_404(PartCompatibility, id=relation_id, part=editing_part)
                relation.delete()
                messages.success(request, "Связь удалена.")

            edit_params = dict(post_back_params)
            if relation_query:
                edit_params["relation_q"] = relation_query
            qs = urlencode(edit_params)
            edit_url = reverse("part_edit", kwargs={"part_id": editing_part.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        if action == "add_attachment" and editing_part:
            attachment_form = CatalogAttachmentForm(request.POST, request.FILES)
            if attachment_form.is_valid():
                _save_catalog_attachment("part", editing_part, PartAttachment, attachment_form)
                messages.success(request, "Вложение добавлено.")
            else:
                messages.error(request, _first_form_error(attachment_form))

            edit_params = dict(post_back_params)
            if relation_query:
                edit_params["relation_q"] = relation_query
            qs = urlencode(edit_params)
            edit_url = reverse("part_edit", kwargs={"part_id": editing_part.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        if action == "remove_attachment" and editing_part:
            attachment_id = request.POST.get("attachment_id", "").strip()
            if attachment_id.isdigit():
                attachment = get_object_or_404(
                    PartAttachment,
                    id=int(attachment_id),
                    part=editing_part,
                )
                attachment.delete()
                messages.success(request, "Вложение удалено.")

            edit_params = dict(post_back_params)
            if relation_query:
                edit_params["relation_q"] = relation_query
            qs = urlencode(edit_params)
            edit_url = reverse("part_edit", kwargs={"part_id": editing_part.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        form = PartForm(request.POST, instance=editing_part)
        if form.is_valid():
            saved = form.save()
            if editing_part:
                messages.success(request, f"Запчасть \"{saved.name}\" сохранена.")
            else:
                messages.success(request, f"Запчасть \"{saved.name}\" создана.")

            edit_params = dict(post_back_params)
            if relation_query:
                edit_params["relation_q"] = relation_query
            qs = urlencode(edit_params)
            edit_url = reverse("part_edit", kwargs={"part_id": saved.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)
    else:
        form = PartForm(instance=editing_part)

    linked_relations = PartCompatibility.objects.none()
    relation_candidates = ProductModel.objects.none()
    attachments = PartAttachment.objects.none()
    if editing_part:
        attachments = editing_part.attachments.all()
        model_characteristics = list(
            PartCharacteristic.objects.select_related("characteristic_type").filter(part=editing_part)
        )
        for item in model_characteristics:
            if item.characteristic_type.value_kind == EquipmentCharacteristicType.ValueKind.TAGS:
                item.display_tags = _parse_tags_value(item.value)
            else:
                item.display_tags = []

        linked_relations = PartCompatibility.objects.select_related(
            "product_model__brand",
            "product_model__category",
        ).filter(part=editing_part)

        linked_ids = linked_relations.values_list("product_model_id", flat=True)
        relation_candidates = ProductModel.objects.select_related("brand", "category").exclude(id__in=linked_ids)
        if relation_query_lc:
            relation_candidates = relation_candidates.annotate(
                name_lc=Lower("name"),
                sku_lc=Lower("sku"),
                brand_name_lc=Lower("brand__name"),
                category_name_lc=Lower("category__name"),
            ).filter(
                Q(name_lc__contains=relation_query_lc)
                | Q(sku_lc__contains=relation_query_lc)
                | Q(brand_name_lc__contains=relation_query_lc)
                | Q(category_name_lc__contains=relation_query_lc)
            )
        relation_candidates = relation_candidates.order_by("name", "id")[:100]

    back_qs = urlencode(back_params)
    back_url = reverse("part")
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    context = {
        "form": form,
        "editing_part": editing_part,
        "model_characteristics": model_characteristics,
        "characteristic_types": characteristic_types,
        "linked_relations": linked_relations,
        "relation_candidates": relation_candidates,
        "relation_query": relation_query,
        "back_url": back_url,
        "attachments": attachments,
        "attachment_form": attachment_form,
        "q": back_query,
        "brand": back_brand,
        "category": back_category,
        "device_type": back_device_type,
        "attachment": back_attachment,
        "sort": back_sort,
        "dir": back_dir,
    }
    return render(request, "part_edit.html", context)


@login_required(login_url="login")
def consumable_product_models(request, consumable_id):
    consumable = get_object_or_404(Consumable, id=consumable_id)

    back_params = {}
    for key in ("q", "brand", "category", "device_type", "per_page", "sort", "dir"):
        value = request.GET.get(key, "").strip() if request.method == "GET" else request.POST.get(key, "").strip()
        if value:
            back_params[key] = value

    relation_query = request.GET.get("relation_q", "").strip() if request.method == "GET" else request.POST.get("relation_q", "").strip()
    relation_query_lc = _normalize_search_term(relation_query)
    relation_brand = request.GET.get("relation_brand", "").strip() if request.method == "GET" else request.POST.get("relation_brand", "").strip()
    relation_category = request.GET.get("relation_category", "").strip() if request.method == "GET" else request.POST.get("relation_category", "").strip()

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add_relation":
            product_model_id = request.POST.get("product_model_id")
            if product_model_id:
                model = get_object_or_404(ProductModel, id=product_model_id)
                _, created = ConsumableCompatibility.objects.get_or_create(consumable=consumable, product_model=model)
                if created:
                    messages.success(request, f"Связь добавлена: {model.name}")
                else:
                    messages.info(request, "Такая связь уже существует.")
        elif action == "remove_relation":
            relation_id = request.POST.get("relation_id")
            if relation_id:
                relation = get_object_or_404(ConsumableCompatibility, id=relation_id, consumable=consumable)
                relation.delete()
                messages.success(request, "Связь удалена.")

        redirect_params = dict(back_params)
        if relation_query:
            redirect_params["relation_q"] = relation_query
        if relation_brand:
            redirect_params["relation_brand"] = relation_brand
        if relation_category:
            redirect_params["relation_category"] = relation_category
        qs = urlencode(redirect_params)
        url = reverse("consumable_product_models", kwargs={"consumable_id": consumable.id})
        return redirect(f"{url}?{qs}" if qs else url)

    linked_relations = ConsumableCompatibility.objects.select_related(
        "product_model__brand",
        "product_model__category",
    ).filter(consumable=consumable)
    linked_ids = linked_relations.values_list("product_model_id", flat=True)

    candidates = ProductModel.objects.select_related("brand", "category").exclude(id__in=linked_ids)
    if relation_query_lc:
        candidates = candidates.annotate(
            name_lc=Lower("name"),
            sku_lc=Lower("sku"),
            brand_name_lc=Lower("brand__name"),
            category_name_lc=Lower("category__name"),
        ).filter(
            Q(name_lc__contains=relation_query_lc)
            | Q(sku_lc__contains=relation_query_lc)
            | Q(brand_name_lc__contains=relation_query_lc)
            | Q(category_name_lc__contains=relation_query_lc)
        )
    if relation_brand:
        try:
            candidates = candidates.filter(brand_id=int(relation_brand))
        except ValueError:
            pass
    if relation_category:
        try:
            candidates = candidates.filter(category_id=int(relation_category))
        except ValueError:
            pass
    candidates = candidates.order_by("name", "id")[:200]

    back_qs = urlencode(back_params)
    back_url = reverse("consumable_edit", kwargs={"consumable_id": consumable.id})
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    context = {
        "item": consumable,
        "linked_relations": linked_relations,
        "relation_candidates": candidates,
        "relation_query": relation_query,
        "relation_brand": relation_brand,
        "relation_category": relation_category,
        "all_brands": Brand.objects.all().order_by("name"),
        "all_categories": ProductCategory.objects.all().order_by("group", "name"),
        "back_url": back_url,
        "back_params": back_params,
    }
    return render(request, "consumable_product_models.html", context)


@login_required(login_url="login")
def part_product_models(request, part_id):
    part = get_object_or_404(Part, id=part_id)

    back_params = {}
    for key in ("q", "brand", "category", "device_type", "per_page", "sort", "dir"):
        value = request.GET.get(key, "").strip() if request.method == "GET" else request.POST.get(key, "").strip()
        if value:
            back_params[key] = value

    relation_query = request.GET.get("relation_q", "").strip() if request.method == "GET" else request.POST.get("relation_q", "").strip()
    relation_query_lc = _normalize_search_term(relation_query)
    relation_brand = request.GET.get("relation_brand", "").strip() if request.method == "GET" else request.POST.get("relation_brand", "").strip()
    relation_category = request.GET.get("relation_category", "").strip() if request.method == "GET" else request.POST.get("relation_category", "").strip()

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add_relation":
            product_model_id = request.POST.get("product_model_id")
            if product_model_id:
                model = get_object_or_404(ProductModel, id=product_model_id)
                _, created = PartCompatibility.objects.get_or_create(part=part, product_model=model)
                if created:
                    messages.success(request, f"Связь добавлена: {model.name}")
                else:
                    messages.info(request, "Такая связь уже существует.")
        elif action == "remove_relation":
            relation_id = request.POST.get("relation_id")
            if relation_id:
                relation = get_object_or_404(PartCompatibility, id=relation_id, part=part)
                relation.delete()
                messages.success(request, "Связь удалена.")

        redirect_params = dict(back_params)
        if relation_query:
            redirect_params["relation_q"] = relation_query
        if relation_brand:
            redirect_params["relation_brand"] = relation_brand
        if relation_category:
            redirect_params["relation_category"] = relation_category
        qs = urlencode(redirect_params)
        url = reverse("part_product_models", kwargs={"part_id": part.id})
        return redirect(f"{url}?{qs}" if qs else url)

    linked_relations = PartCompatibility.objects.select_related(
        "product_model__brand",
        "product_model__category",
    ).filter(part=part)
    linked_ids = linked_relations.values_list("product_model_id", flat=True)

    candidates = ProductModel.objects.select_related("brand", "category").exclude(id__in=linked_ids)
    if relation_query_lc:
        candidates = candidates.annotate(
            name_lc=Lower("name"),
            sku_lc=Lower("sku"),
            brand_name_lc=Lower("brand__name"),
            category_name_lc=Lower("category__name"),
        ).filter(
            Q(name_lc__contains=relation_query_lc)
            | Q(sku_lc__contains=relation_query_lc)
            | Q(brand_name_lc__contains=relation_query_lc)
            | Q(category_name_lc__contains=relation_query_lc)
        )
    if relation_brand:
        try:
            candidates = candidates.filter(brand_id=int(relation_brand))
        except ValueError:
            pass
    if relation_category:
        try:
            candidates = candidates.filter(category_id=int(relation_category))
        except ValueError:
            pass
    candidates = candidates.order_by("name", "id")[:200]

    back_qs = urlencode(back_params)
    back_url = reverse("part_edit", kwargs={"part_id": part.id})
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    context = {
        "item": part,
        "linked_relations": linked_relations,
        "relation_candidates": candidates,
        "relation_query": relation_query,
        "relation_brand": relation_brand,
        "relation_category": relation_category,
        "all_brands": Brand.objects.all().order_by("name"),
        "all_categories": ProductCategory.objects.all().order_by("group", "name"),
        "back_url": back_url,
        "back_params": back_params,
    }
    return render(request, "part_product_models.html", context)


@login_required(login_url="login")
def consumable_delete(request, consumable_id):
    item = get_object_or_404(Consumable, id=consumable_id)
    relation_count = item.compatibilities.count()
    can_delete = relation_count == 0

    back_params = {}
    for key in ("q", "brand", "category", "device_type", "attachments", "per_page", "sort", "dir"):
        value = request.GET.get(key, "").strip() if request.method == "GET" else request.POST.get(key, "").strip()
        if value:
            back_params[key] = value

    back_qs = urlencode(back_params)
    back_url = reverse("consumable")
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    if request.method == "POST":
        if not can_delete:
            messages.error(request, "Нельзя удалить расходный материал, пока есть связи с техникой.")
            return redirect(reverse("consumable_edit", kwargs={"consumable_id": item.id}))
        item_name = item.name
        item.delete()
        messages.success(request, f"Расходный материал «{item_name}» удален.")
        return redirect(back_url)

    context = {
        "item": item,
        "back_url": back_url,
        "q": back_params.get("q", ""),
        "brand": back_params.get("brand", ""),
        "category": back_params.get("category", ""),
        "device_type": back_params.get("device_type", ""),
        "attachment_filter": back_params.get("attachments", ""),
        "sort": back_params.get("sort", "name"),
        "dir": back_params.get("dir", "asc"),
        "relation_count": relation_count,
        "can_delete": can_delete,
    }
    return render(request, "consumable_delete.html", context)


@login_required(login_url="login")
def part_delete(request, part_id):
    item = get_object_or_404(Part, id=part_id)
    relation_count = item.compatibilities.count()
    can_delete = relation_count == 0

    back_params = {}
    for key in ("q", "brand", "category", "device_type", "attachments", "per_page", "sort", "dir"):
        value = request.GET.get(key, "").strip() if request.method == "GET" else request.POST.get(key, "").strip()
        if value:
            back_params[key] = value

    back_qs = urlencode(back_params)
    back_url = reverse("part")
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    if request.method == "POST":
        if not can_delete:
            messages.error(request, "Нельзя удалить запчасть, пока есть связи с техникой.")
            return redirect(reverse("part_edit", kwargs={"part_id": item.id}))
        item_name = item.name
        item.delete()
        messages.success(request, f"Запчасть \"{item_name}\" удалена.")
        return redirect(back_url)

    context = {
        "item": item,
        "back_url": back_url,
        "q": back_params.get("q", ""),
        "brand": back_params.get("brand", ""),
        "category": back_params.get("category", ""),
        "device_type": back_params.get("device_type", ""),
        "attachment_filter": back_params.get("attachments", ""),
        "sort": back_params.get("sort", "name"),
        "dir": back_params.get("dir", "asc"),
        "relation_count": relation_count,
        "can_delete": can_delete,
    }
    return render(request, "part_delete.html", context)


@login_required(login_url="login")
def work_directory(request):
    query = request.GET.get("q", "").strip()
    query_lc = _normalize_search_term(query)
    component_filter = request.GET.get("component", "").strip()
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    sort = request.GET.get("sort", "code")
    direction = request.GET.get("dir", "asc")

    sort_map = {
        "code": "code",
        "name": "name",
        "unit_price": "unit_price",
        "consumables_count": "consumables_count",
        "parts_count": "parts_count",
    }

    if sort not in sort_map:
        sort = "code"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    works_qs = WorkDirectory.objects.annotate(
        consumables_count=Count("consumable_links", distinct=True),
        parts_count=Count("part_links", distinct=True),
    )

    if query_lc:
        works_qs = works_qs.annotate(
            code_lc=Lower("code"),
            name_lc=Lower("name"),
        ).filter(
            Q(code_lc__contains=query_lc) | Q(name_lc__contains=query_lc)
        )

    if component_filter == "with_consumables":
        works_qs = works_qs.filter(consumables_count__gt=0)
    elif component_filter == "with_parts":
        works_qs = works_qs.filter(parts_count__gt=0)
    elif component_filter == "with_any":
        works_qs = works_qs.filter(Q(consumables_count__gt=0) | Q(parts_count__gt=0))

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"
    works_qs = works_qs.order_by(order_field, "id")
    paged_works, page_obj, per_page = _paginate_report_queryset(request, works_qs, default_per_page=per_page)

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if component_filter:
            params["component"] = component_filter
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    context = {
        "works": paged_works,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "query": query,
        "component_filter": component_filter,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "work_directory.html", context)


@login_required(login_url="login")
def work_directory_edit(request, work_id=None):
    editing_work = get_object_or_404(WorkDirectory, id=work_id) if work_id else None

    back_query = request.GET.get("q", "").strip()
    back_component = request.GET.get("component", "").strip()
    back_sort = request.GET.get("sort", "code").strip() or "code"
    back_dir = request.GET.get("dir", "asc").strip() or "asc"
    consumable_query = request.GET.get("consumable_q", "").strip()
    part_query = request.GET.get("part_q", "").strip()

    back_params = {}
    if back_query:
        back_params["q"] = back_query
    if back_component:
        back_params["component"] = back_component
    if back_sort:
        back_params["sort"] = back_sort
    if back_dir:
        back_params["dir"] = back_dir

    if request.method == "POST":
        action = request.POST.get("action", "save")
        target_id = request.POST.get("work_id")
        consumable_query = request.POST.get("consumable_q", consumable_query).strip()
        part_query = request.POST.get("part_q", part_query).strip()

        post_back_params = {}
        for key in ("q", "component", "per_page", "sort", "dir"):
            value = request.POST.get(key, "").strip()
            if value:
                post_back_params[key] = value

        if target_id:
            editing_work = get_object_or_404(WorkDirectory, id=target_id)

        if action == "add_consumable" and editing_work:
            consumable_id = request.POST.get("consumable_id")
            quantity_raw = request.POST.get("consumable_quantity", "1").strip()
            try:
                quantity = max(1, int(quantity_raw or "1"))
            except ValueError:
                quantity = 1

            if consumable_id:
                consumable = get_object_or_404(Consumable, id=consumable_id)
                link, created = WorkDirectoryConsumable.objects.get_or_create(
                    work=editing_work,
                    consumable=consumable,
                    defaults={"quantity": quantity},
                )
                if not created:
                    link.quantity = quantity
                    link.save(update_fields=["quantity"])
                    messages.success(request, f"Количество расходного материала «{consumable.name}» обновлено.")
                else:
                    messages.success(request, f"Расходный материал «{consumable.name}» добавлен.")

            edit_params = dict(post_back_params)
            if consumable_query:
                edit_params["consumable_q"] = consumable_query
            if part_query:
                edit_params["part_q"] = part_query
            qs = urlencode(edit_params)
            edit_url = reverse("work_directory_edit", kwargs={"work_id": editing_work.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        if action == "remove_consumable" and editing_work:
            link_id = request.POST.get("link_id")
            if link_id:
                link = get_object_or_404(WorkDirectoryConsumable, id=link_id, work=editing_work)
                link.delete()
                messages.success(request, "Строка расходного материала удалена.")

            edit_params = dict(post_back_params)
            if consumable_query:
                edit_params["consumable_q"] = consumable_query
            if part_query:
                edit_params["part_q"] = part_query
            qs = urlencode(edit_params)
            edit_url = reverse("work_directory_edit", kwargs={"work_id": editing_work.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        if action == "add_part" and editing_work:
            part_id_value = request.POST.get("part_id")
            quantity_raw = request.POST.get("part_quantity", "1").strip()
            try:
                quantity = max(1, int(quantity_raw or "1"))
            except ValueError:
                quantity = 1

            if part_id_value:
                part = get_object_or_404(Part, id=part_id_value)
                link, created = WorkDirectoryPart.objects.get_or_create(
                    work=editing_work,
                    part=part,
                    defaults={"quantity": quantity},
                )
                if not created:
                    link.quantity = quantity
                    link.save(update_fields=["quantity"])
                    messages.success(request, f"Количество запчасти «{part.name}» обновлено.")
                else:
                    messages.success(request, f"Запчасть «{part.name}» добавлена.")

            edit_params = dict(post_back_params)
            if consumable_query:
                edit_params["consumable_q"] = consumable_query
            if part_query:
                edit_params["part_q"] = part_query
            qs = urlencode(edit_params)
            edit_url = reverse("work_directory_edit", kwargs={"work_id": editing_work.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        if action == "remove_part" and editing_work:
            link_id = request.POST.get("link_id")
            if link_id:
                link = get_object_or_404(WorkDirectoryPart, id=link_id, work=editing_work)
                link.delete()
                messages.success(request, "Строка запчасти удалена.")

            edit_params = dict(post_back_params)
            if consumable_query:
                edit_params["consumable_q"] = consumable_query
            if part_query:
                edit_params["part_q"] = part_query
            qs = urlencode(edit_params)
            edit_url = reverse("work_directory_edit", kwargs={"work_id": editing_work.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)

        form = WorkDirectoryForm(request.POST, instance=editing_work)
        if form.is_valid():
            saved = form.save()
            if editing_work:
                messages.success(request, f"Работа \"{saved.name}\" сохранена.")
            else:
                messages.success(request, f"Работа \"{saved.name}\" создана.")

            edit_params = dict(post_back_params)
            if consumable_query:
                edit_params["consumable_q"] = consumable_query
            if part_query:
                edit_params["part_q"] = part_query
            qs = urlencode(edit_params)
            edit_url = reverse("work_directory_edit", kwargs={"work_id": saved.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)
    else:
        form = WorkDirectoryForm(instance=editing_work)

    linked_consumables = WorkDirectoryConsumable.objects.none()
    linked_parts = WorkDirectoryPart.objects.none()
    consumable_candidates = Consumable.objects.none()
    part_candidates = Part.objects.none()

    consumable_query_lc = _normalize_search_term(consumable_query)
    part_query_lc = _normalize_search_term(part_query)

    if editing_work:
        linked_consumables = WorkDirectoryConsumable.objects.select_related(
            "consumable__brand",
            "consumable__category",
        ).filter(work=editing_work)
        linked_parts = WorkDirectoryPart.objects.select_related(
            "part__brand",
            "part__category",
        ).filter(work=editing_work)

        linked_consumable_ids = linked_consumables.values_list("consumable_id", flat=True)
        linked_part_ids = linked_parts.values_list("part_id", flat=True)

        consumable_candidates = Consumable.objects.select_related("brand", "category").exclude(id__in=linked_consumable_ids)
        if consumable_query_lc:
            consumable_candidates = consumable_candidates.annotate(
                name_lc=Lower("name"),
                sku_lc=Lower("sku"),
                brand_name_lc=Lower("brand__name"),
                category_name_lc=Lower("category__name"),
            ).filter(
                Q(name_lc__contains=consumable_query_lc)
                | Q(sku_lc__contains=consumable_query_lc)
                | Q(brand_name_lc__contains=consumable_query_lc)
                | Q(category_name_lc__contains=consumable_query_lc)
            )
        consumable_candidates = consumable_candidates.order_by("name", "id")[:100]

        part_candidates = Part.objects.select_related("brand", "category").exclude(id__in=linked_part_ids)
        if part_query_lc:
            part_candidates = part_candidates.annotate(
                name_lc=Lower("name"),
                sku_lc=Lower("sku"),
                brand_name_lc=Lower("brand__name"),
                category_name_lc=Lower("category__name"),
            ).filter(
                Q(name_lc__contains=part_query_lc)
                | Q(sku_lc__contains=part_query_lc)
                | Q(brand_name_lc__contains=part_query_lc)
                | Q(category_name_lc__contains=part_query_lc)
            )
        part_candidates = part_candidates.order_by("name", "id")[:100]

    back_qs = urlencode(back_params)
    back_url = reverse("work_directory")
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    context = {
        "form": form,
        "editing_work": editing_work,
        "linked_consumables": linked_consumables,
        "linked_parts": linked_parts,
        "consumable_candidates": consumable_candidates,
        "part_candidates": part_candidates,
        "consumable_query": consumable_query,
        "part_query": part_query,
        "back_url": back_url,
        "q": back_query,
        "component": back_component,
        "sort": back_sort,
        "dir": back_dir,
    }
    return render(request, "work_directory_edit.html", context)


@login_required(login_url="login")
def work_directory_delete(request, work_id):
    item = get_object_or_404(WorkDirectory, id=work_id)

    back_params = {}
    for key in ("q", "component", "per_page", "sort", "dir"):
        value = request.GET.get(key, "").strip() if request.method == "GET" else request.POST.get(key, "").strip()
        if value:
            back_params[key] = value

    back_qs = urlencode(back_params)
    back_url = reverse("work_directory")
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    if request.method == "POST":
        item_name = item.name
        item.delete()
        messages.success(request, f"Работа \"{item_name}\" удалена.")
        return redirect(back_url)

    context = {
        "item": item,
        "back_url": back_url,
        "title": "Удаление работы",
        "object_label": "работу",
        "q": back_params.get("q", ""),
        "component": back_params.get("component", ""),
        "sort": back_params.get("sort", "code"),
        "dir": back_params.get("dir", "asc"),
    }
    return render(request, "dictionary_delete.html", context)


@login_required(login_url="login")
def serviceman_delete(request, serviceman_id):
    item = get_object_or_404(ServiceMan, id=serviceman_id)

    back_params = {}
    for key in ("q", "status", "per_page", "sort", "dir"):
        value = request.GET.get(key, "").strip() if request.method == "GET" else request.POST.get(key, "").strip()
        if value:
            back_params[key] = value

    back_qs = urlencode(back_params)
    back_url = reverse("serviceman")
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    if request.method == "POST":
        item_name = item.full_name
        item.delete()
        messages.success(request, f"Сервисный инженер \"{item_name}\" удален.")
        return redirect(back_url)

    context = {
        "item": item,
        "back_url": back_url,
        "title": "Удаление сервисного инженера",
        "object_label": "сервисного инженера",
        "q": back_params.get("q", ""),
        "status": back_params.get("status", ""),
        "sort": back_params.get("sort", "full_name"),
        "dir": back_params.get("dir", "asc"),
    }
    return render(request, "dictionary_delete.html", context)


@login_required(login_url="login")
def status_directory_delete(request, status_id):
    item = get_object_or_404(StatusDirectory, id=status_id)

    back_params = {}
    for key in ("q", "code", "per_page", "sort", "dir"):
        value = request.GET.get(key, "").strip() if request.method == "GET" else request.POST.get(key, "").strip()
        if value:
            back_params[key] = value

    back_qs = urlencode(back_params)
    back_url = reverse("status_directory")
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    if request.method == "POST":
        item_name = item.name
        item.delete()
        messages.success(request, f"Статус \"{item_name}\" удален.")
        return redirect(back_url)

    context = {
        "item": item,
        "back_url": back_url,
        "title": "Удаление статуса",
        "object_label": "статус",
        "q": back_params.get("q", ""),
        "code": back_params.get("code", ""),
        "sort": back_params.get("sort", "code"),
        "dir": back_params.get("dir", "asc"),
    }
    return render(request, "dictionary_delete.html", context)


@login_required(login_url="login")
def status_directory(request):
    status_id = request.GET.get("edit")
    query = request.GET.get("q", "").strip()
    normalized_query = _normalize_search_term(query)
    code_filter = request.GET.get("code", "").strip()
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    sort = request.GET.get("sort", "code")
    direction = request.GET.get("dir", "asc")
    editing_status = None

    sort_map = {
        "code": "code",
        "name": "name",
        "description": "description",
    }

    if sort not in sort_map:
        sort = "code"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"

    if status_id:
        editing_status = get_object_or_404(StatusDirectory, id=status_id)

    if request.method == "POST":
        action = request.POST.get("action", "save")
        target_id = request.POST.get("status_id")
        post_query = request.POST.get("q", query).strip()
        post_code = request.POST.get("code", code_filter).strip()
        post_per_page = request.POST.get("per_page", per_page).strip().lower()
        post_sort = request.POST.get("sort", sort).strip() or "code"
        post_direction = request.POST.get("dir", direction).strip() or "asc"

        params = {}
        if post_query:
            params["q"] = post_query
        if post_code:
            params["code"] = post_code
        if post_per_page in {"10", "30", "50", "100", "all"}:
            params["per_page"] = post_per_page
        if post_sort in sort_map:
            params["sort"] = post_sort
        if post_direction in {"asc", "desc"}:
            params["dir"] = post_direction

        redirect_url = reverse("status_directory")
        if params:
            redirect_url = f"{redirect_url}?{urlencode(params)}"

        if action == "delete" and target_id:
            target = get_object_or_404(StatusDirectory, id=target_id)
            target.delete()
            messages.success(request, "Статус удален.")
            return redirect(redirect_url)

        if target_id:
            editing_status = get_object_or_404(StatusDirectory, id=target_id)
            form = StatusDirectoryForm(request.POST, instance=editing_status)
            is_new = False
        else:
            form = StatusDirectoryForm(request.POST)
            is_new = True

        if form.is_valid():
            saved = form.save()
            if is_new:
                messages.success(request, f"Статус \"{saved.name}\" создан.")
            else:
                messages.success(request, f"Статус \"{saved.name}\" сохранен.")
            params["edit"] = saved.id
            return redirect(f"{reverse('status_directory')}?{urlencode(params)}")
    else:
        form = StatusDirectoryForm(instance=editing_status)

    statuses_qs = StatusDirectory.objects.all()
    if normalized_query:
        query_filter = Q(name__icontains=normalized_query) | Q(description__icontains=normalized_query)
        if normalized_query.isdigit():
            query_filter = query_filter | Q(code=int(normalized_query))
        statuses_qs = statuses_qs.filter(query_filter)

    if code_filter:
        try:
            statuses_qs = statuses_qs.filter(code=int(code_filter))
        except ValueError:
            pass

    statuses_qs = statuses_qs.order_by(order_field, "id")
    paged_statuses, page_obj, per_page = _paginate_report_queryset(request, statuses_qs, default_per_page=per_page)

    all_codes = StatusDirectory.objects.values_list("code", flat=True).order_by("code")

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        p = {"sort": key, "dir": next_dir}
        if query:
            p["q"] = query
        if code_filter:
            p["code"] = code_filter
        if per_page:
            p["per_page"] = per_page
        sort_links[key] = urlencode(p)

    context = {
        "form": form,
        "editing_status": editing_status,
        "statuses": paged_statuses,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "all_codes": all_codes,
        "query": query,
        "code_filter": code_filter,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "status_directory.html", context)


@login_required(login_url="login")
def address_directory_delete(request, address_id):
    item = get_object_or_404(Address, id=address_id)

    back_params = {}
    for key in ("q", "locality", "per_page", "sort", "dir"):
        value = request.GET.get(key, "").strip() if request.method == "GET" else request.POST.get(key, "").strip()
        if value:
            back_params[key] = value

    back_qs = urlencode(back_params)
    back_url = reverse("address_directory")
    if back_qs:
        back_url = f"{back_url}?{back_qs}"

    if request.method == "POST":
        item_name = str(item)
        item.delete()
        messages.success(request, f"Адрес \"{item_name}\" удален.")
        return redirect(back_url)

    context = {
        "item": item,
        "back_url": back_url,
        "title": "Удаление адреса",
        "object_label": "адрес",
        "q": back_params.get("q", ""),
        "locality": back_params.get("locality", ""),
        "sort": back_params.get("sort", "locality"),
        "dir": back_params.get("dir", "asc"),
    }
    return render(request, "dictionary_delete.html", context)


@login_required(login_url="login")
def address_directory(request):
    address_id = request.GET.get("edit")
    query = request.GET.get("q", "").strip()
    normalized_query = _normalize_search_term(query)
    locality_filter = request.GET.get("locality", "").strip()
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    sort = request.GET.get("sort", "locality")
    direction = request.GET.get("dir", "asc")
    editing_address = None

    sort_map = {
        "postal_code": "postal_code",
        "locality": "locality",
        "street": "street",
        "house": "house",
        "room": "room",
    }

    if sort not in sort_map:
        sort = "locality"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"

    if address_id:
        editing_address = get_object_or_404(Address, id=address_id)

    if request.method == "POST":
        target_id = request.POST.get("address_id")
        post_query = request.POST.get("q", query).strip()
        post_locality = request.POST.get("locality", locality_filter).strip()
        post_per_page = request.POST.get("per_page", per_page).strip().lower()
        post_sort = request.POST.get("sort", sort).strip() or "locality"
        post_direction = request.POST.get("dir", direction).strip() or "asc"

        params = {}
        if post_query:
            params["q"] = post_query
        if post_locality:
            params["locality"] = post_locality
        if post_per_page in {"10", "30", "50", "100", "all"}:
            params["per_page"] = post_per_page
        if post_sort in sort_map:
            params["sort"] = post_sort
        if post_direction in {"asc", "desc"}:
            params["dir"] = post_direction

        if target_id:
            editing_address = get_object_or_404(Address, id=target_id)
            form = AddressForm(request.POST, instance=editing_address)
            is_new = False
        else:
            form = AddressForm(request.POST)
            is_new = True

        if form.is_valid():
            saved = form.save()
            if is_new:
                messages.success(request, f"Адрес \"{saved}\" создан.")
            else:
                messages.success(request, f"Адрес \"{saved}\" сохранен.")
            params["edit"] = saved.id
            return redirect(f"{reverse('address_directory')}?{urlencode(params)}")
    else:
        form = AddressForm(instance=editing_address)

    addresses_qs = Address.objects.all()
    if normalized_query:
        addresses_qs = addresses_qs.annotate(
            postal_code_lc=Lower("postal_code"),
            locality_lc=Lower("locality"),
            street_lc=Lower("street"),
            house_lc=Lower("house"),
            building_lc=Lower("building"),
            structure_lc=Lower("structure"),
            floor_lc=Lower("floor"),
            room_lc=Lower("room"),
            note_lc=Lower("note"),
        ).filter(
            Q(postal_code_lc__contains=normalized_query)
            | Q(locality_lc__contains=normalized_query)
            | Q(street_lc__contains=normalized_query)
            | Q(house_lc__contains=normalized_query)
            | Q(building_lc__contains=normalized_query)
            | Q(structure_lc__contains=normalized_query)
            | Q(floor_lc__contains=normalized_query)
            | Q(room_lc__contains=normalized_query)
            | Q(note_lc__contains=normalized_query)
        )

    if locality_filter:
        addresses_qs = addresses_qs.filter(locality=locality_filter)

    addresses_qs = addresses_qs.order_by(order_field, "id")
    paged_addresses, page_obj, per_page = _paginate_report_queryset(request, addresses_qs, default_per_page=per_page)

    all_localities = Address.objects.exclude(locality="").values_list("locality", flat=True).distinct().order_by("locality")

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        p = {"sort": key, "dir": next_dir}
        if query:
            p["q"] = query
        if locality_filter:
            p["locality"] = locality_filter
        if per_page:
            p["per_page"] = per_page
        sort_links[key] = urlencode(p)

    context = {
        "form": form,
        "editing_address": editing_address,
        "addresses": paged_addresses,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "all_localities": all_localities,
        "query": query,
        "locality_filter": locality_filter,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "address_directory.html", context)


def _acceptance_document_redirect_url(request: HttpRequest, edit_id: int | None = None) -> str:
    params = {}
    query = request.POST.get("q", request.GET.get("q", "")).strip()

    if edit_id:
        params["edit"] = str(edit_id)
    if query:
        params["q"] = query

    base_url = reverse("acceptance_document")
    if params:
        base_url = f"{base_url}?{urlencode(params)}"
    return base_url


def _ensure_repair_document_from_acceptance_link(link: AcceptanceDocumentEquipment) -> bool:
    if link.repair_document_id:
        return False

    acceptance_document = link.acceptance_document
    if not acceptance_document.serviceman_id:
        return False

    status_acceptance, _ = StatusDirectory.objects.get_or_create(
        code=10,
        defaults={"name": "Приемка", "description": "Автоматический статус приемки"},
    )

    repair_document = RepairDocument.objects.create(
        date=acceptance_document.date,
        repair_place=RepairDocument.RepairPlace.OFFICE,
        organization=acceptance_document.organization,
        serviceman=acceptance_document.serviceman,
        status=status_acceptance,
        client_equipment=link.client_equipment,
        malfunction="",
        work_performed="",
        note=f"Создано автоматически из документа приемки #{acceptance_document.id}",
    )

    link.repair_document = repair_document
    link.save(update_fields=["repair_document"])
    return True


@login_required(login_url="login")
def acceptance_document(request):
    query = request.GET.get("q", "").strip()
    query_lc = _normalize_search_term(query)
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"

    documents_qs = AcceptanceDocument.objects.select_related("organization", "serviceman").all()
    if query_lc:
        documents_qs = documents_qs.annotate(
            organization_name_lc=Lower("organization__name"),
        ).filter(
            Q(organization_name_lc__contains=query_lc)
        )

    documents_qs = documents_qs.order_by("-date", "-id")
    paged_documents, page_obj, per_page = _paginate_report_queryset(request, documents_qs, default_per_page=per_page)

    context = {
        "query": query,
        "documents": paged_documents,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
    }
    return render(request, "acceptance_document.html", context)


@login_required(login_url="login")
def acceptance_document_edit(request, document_id=None):
    editing_document = get_object_or_404(AcceptanceDocument, id=document_id) if document_id else None
    ask_add_serial = request.GET.get("ask_add_serial", "").strip()
    serial_query = request.GET.get("serial", "").strip()
    attachment_form = CatalogAttachmentForm()

    if request.method == "POST":
        action = request.POST.get("action", "save")
        form = AcceptanceDocumentForm(request.POST, instance=editing_document)

        if action == "save":
            if form.is_valid():
                saved = form.save()
                generated_count = 0
                for link in saved.equipment_links.select_related("acceptance_document", "client_equipment"):
                    if _ensure_repair_document_from_acceptance_link(link):
                        generated_count += 1

                if editing_document:
                    messages.success(request, f"Документ приемки #{saved.id} сохранен.")
                else:
                    messages.success(request, f"Документ приемки #{saved.id} создан.")

                if generated_count:
                    messages.success(request, f"Автоматически создано документов ремонта: {generated_count}.")

                return redirect(reverse("acceptance_document_edit", kwargs={"document_id": saved.id}))

        elif action == "search_serial":
            if not editing_document:
                messages.error(request, "Сначала сохраните документ, затем выполняйте поиск по серийному номеру.")
            else:
                serial_query = request.POST.get("serial_query", "").strip()
                if not serial_query:
                    messages.error(request, "Введите серийный номер для поиска.")
                else:
                    found_equipment = ClientEquipment.objects.filter(
                        organization_id=editing_document.organization_id,
                        serial_number__iexact=serial_query,
                    ).order_by("id").first()

                    if found_equipment:
                        duplicate_serial = AcceptanceDocumentEquipment.objects.filter(
                            acceptance_document=editing_document,
                            client_equipment__serial_number__iexact=serial_query,
                        ).exists()
                        if duplicate_serial:
                            messages.info(request, "Техника с этим серийным номером уже добавлена в текущий документ.")
                        else:
                            link, _ = AcceptanceDocumentEquipment.objects.get_or_create(
                                acceptance_document=editing_document,
                                client_equipment=found_equipment,
                            )
                            created_repair = _ensure_repair_document_from_acceptance_link(link)
                            messages.success(request, f"Техника с серийным номером {serial_query} добавлена в документ.")
                            if created_repair:
                                messages.success(request, "Для добавленной техники автоматически создан документ ремонта.")
                        redirect_params = {"serial": serial_query}
                        return redirect(
                            f"{reverse('acceptance_document_edit', kwargs={'document_id': editing_document.id})}?{urlencode(redirect_params)}"
                        )

                    redirect_params = {"ask_add_serial": serial_query, "serial": serial_query}
                    return redirect(
                        f"{reverse('acceptance_document_edit', kwargs={'document_id': editing_document.id})}"
                        f"?{urlencode(redirect_params)}"
                    )

        elif action == "remove_equipment" and editing_document:
            link_id = request.POST.get("link_id")
            if link_id:
                link = get_object_or_404(
                    AcceptanceDocumentEquipment,
                    id=link_id,
                    acceptance_document=editing_document,
                )
                link.delete()
                messages.success(request, "Строка техники удалена из документа приемки.")
            return redirect(reverse("acceptance_document_edit", kwargs={"document_id": editing_document.id}))

        elif action == "add_attachment" and editing_document:
            attachment_form = CatalogAttachmentForm(request.POST, request.FILES)
            if attachment_form.is_valid():
                _save_catalog_attachment("acceptance_document", editing_document, AcceptanceDocumentAttachment, attachment_form)
                messages.success(request, "Вложение добавлено.")
            else:
                messages.error(request, _first_form_error(attachment_form))
            return redirect(reverse("acceptance_document_edit", kwargs={"document_id": editing_document.id}))

        elif action == "remove_attachment" and editing_document:
            attachment_id = request.POST.get("attachment_id", "").strip()
            if attachment_id.isdigit():
                attachment = get_object_or_404(
                    AcceptanceDocumentAttachment,
                    id=int(attachment_id),
                    acceptance_document=editing_document,
                )
                attachment.delete()
                messages.success(request, "Вложение удалено.")
            return redirect(reverse("acceptance_document_edit", kwargs={"document_id": editing_document.id}))
    else:
        initial = {"date": timezone.localdate()} if not editing_document else {}
        selected_org = request.GET.get("organization", "").strip()
        if selected_org.isdigit() and not editing_document:
            initial["organization"] = int(selected_org)
        form = AcceptanceDocumentForm(instance=editing_document, initial=initial)

    linked_equipment = AcceptanceDocumentEquipment.objects.none()
    attachments = AcceptanceDocumentAttachment.objects.none()
    if editing_document:
        linked_equipment = AcceptanceDocumentEquipment.objects.select_related(
            "client_equipment__product_model",
            "client_equipment__organization",
            "repair_document",
        ).filter(
            acceptance_document=editing_document
        )
        attachments = editing_document.attachments.all()

    add_equipment_url = ""
    if editing_document and ask_add_serial:
        add_equipment_url = f"{reverse('acceptance_document_add_equipment', kwargs={'document_id': editing_document.id})}?{urlencode({'serial': ask_add_serial})}"

    context = {
        "form": form,
        "editing_document": editing_document,
        "linked_equipment": linked_equipment,
        "serial_query": serial_query,
        "ask_add_serial": ask_add_serial,
        "add_equipment_url": add_equipment_url,
        "attachments": attachments,
        "attachment_form": attachment_form,
        "back_url": reverse("acceptance_document"),
    }
    return render(request, "acceptance_document_edit.html", context)


@login_required(login_url="login")
def acceptance_document_add_equipment(request, document_id):
    acceptance_document = get_object_or_404(AcceptanceDocument, id=document_id)
    serial = request.GET.get("serial", "").strip() if request.method == "GET" else request.POST.get("serial_number", "").strip()

    if not serial:
        messages.error(request, "Не передан серийный номер для добавления техники.")
        return redirect(reverse("acceptance_document_edit", kwargs={"document_id": acceptance_document.id}))

    if request.method == "POST":
        form = AcceptanceEquipmentCreateForm(request.POST)
        form._organization_id = acceptance_document.organization_id
        form.instance.organization_id = acceptance_document.organization_id
        if form.is_valid():
            equipment = form.save(commit=False)
            equipment.organization_id = acceptance_document.organization_id
            equipment.serial_number = serial
            equipment.save()

            link, _ = AcceptanceDocumentEquipment.objects.get_or_create(
                acceptance_document=acceptance_document,
                client_equipment=equipment,
            )
            _ensure_repair_document_from_acceptance_link(link)
            messages.success(request, f"Техника с серийным номером {serial} создана и добавлена в документ.")
            return redirect(reverse("acceptance_document_edit", kwargs={"document_id": acceptance_document.id}))
    else:
        form = AcceptanceEquipmentCreateForm(initial={"serial_number": serial})

    context = {
        "form": form,
        "acceptance_document": acceptance_document,
        "serial": serial,
        "back_url": reverse("acceptance_document_edit", kwargs={"document_id": acceptance_document.id}),
    }
    return render(request, "acceptance_document_add_equipment.html", context)


def _ensure_repair_document_from_shipment_link(link: ShipmentDocumentEquipment) -> bool:
    if link.repair_document_id:
        return False

    shipment_document = link.shipment_document
    if not shipment_document.serviceman_id:
        return False

    status_shipped, _ = StatusDirectory.objects.get_or_create(
        code=120,
        defaults={"name": "Отгружено клиенту", "description": "Автоматический статус отгрузки"},
    )

    repair_document = RepairDocument.objects.create(
        date=shipment_document.date,
        repair_place=RepairDocument.RepairPlace.OFFICE,
        organization=shipment_document.organization,
        serviceman=shipment_document.serviceman,
        status=status_shipped,
        client_equipment=link.client_equipment,
        malfunction="",
        work_performed="",
        note=f"Создано автоматически из документа отгрузки #{shipment_document.id}",
    )

    link.repair_document = repair_document
    link.save(update_fields=["repair_document"])
    return True


@login_required(login_url="login")
def shipment_document(request):
    query = request.GET.get("q", "").strip()
    query_lc = _normalize_search_term(query)
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"

    documents_qs = ShipmentDocument.objects.select_related("organization", "serviceman").all()
    if query_lc:
        documents_qs = documents_qs.annotate(
            organization_name_lc=Lower("organization__name"),
        ).filter(
            Q(organization_name_lc__contains=query_lc)
        )

    documents_qs = documents_qs.order_by("-date", "-id")
    paged_documents, page_obj, per_page = _paginate_report_queryset(request, documents_qs, default_per_page=per_page)

    context = {
        "query": query,
        "documents": paged_documents,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
    }
    return render(request, "shipment_document.html", context)


@login_required(login_url="login")
def shipment_document_edit(request, document_id=None):
    editing_document = get_object_or_404(ShipmentDocument, id=document_id) if document_id else None
    serial_query = request.GET.get("serial", "").strip()
    attachment_form = CatalogAttachmentForm()

    if request.method == "POST":
        action = request.POST.get("action", "save")
        form = ShipmentDocumentForm(request.POST, instance=editing_document)

        if action == "save":
            if form.is_valid():
                saved = form.save()
                generated_count = 0
                for link in saved.equipment_links.select_related("shipment_document", "client_equipment"):
                    if _ensure_repair_document_from_shipment_link(link):
                        generated_count += 1

                if editing_document:
                    messages.success(request, f"Документ отгрузки #{saved.id} сохранен.")
                else:
                    messages.success(request, f"Документ отгрузки #{saved.id} создан.")

                if generated_count:
                    messages.success(request, f"Автоматически создано документов ремонта: {generated_count}.")

                return redirect(reverse("shipment_document_edit", kwargs={"document_id": saved.id}))

        elif action == "search_serial":
            if not editing_document:
                messages.error(request, "Сначала сохраните документ, затем выполняйте поиск по серийному номеру.")
            else:
                serial_query = request.POST.get("serial_query", "").strip()
                if not serial_query:
                    messages.error(request, "Введите серийный номер для поиска.")
                else:
                    found_equipment = ClientEquipment.objects.filter(
                        organization_id=editing_document.organization_id,
                        serial_number__iexact=serial_query,
                    ).order_by("id").first()

                    if found_equipment:
                        duplicate_serial = ShipmentDocumentEquipment.objects.filter(
                            shipment_document=editing_document,
                            client_equipment__serial_number__iexact=serial_query,
                        ).exists()
                        if duplicate_serial:
                            messages.info(request, "Техника с этим серийным номером уже добавлена в текущий документ.")
                        else:
                            link, _ = ShipmentDocumentEquipment.objects.get_or_create(
                                shipment_document=editing_document,
                                client_equipment=found_equipment,
                            )
                            created_repair = _ensure_repair_document_from_shipment_link(link)
                            messages.success(request, f"Техника с серийным номером {serial_query} добавлена в документ.")
                            if created_repair:
                                messages.success(request, "Для добавленной техники автоматически создан документ ремонта.")
                    else:
                        messages.info(request, "Серийный номер не найден. Строка техники не добавлена.")

                    redirect_params = {"serial": serial_query}
                    return redirect(
                        f"{reverse('shipment_document_edit', kwargs={'document_id': editing_document.id})}?{urlencode(redirect_params)}"
                    )

        elif action == "remove_equipment" and editing_document:
            link_id = request.POST.get("link_id")
            if link_id:
                link = get_object_or_404(
                    ShipmentDocumentEquipment,
                    id=link_id,
                    shipment_document=editing_document,
                )
                link.delete()
                messages.success(request, "Строка техники удалена из документа отгрузки.")
            return redirect(reverse("shipment_document_edit", kwargs={"document_id": editing_document.id}))

        elif action == "add_attachment" and editing_document:
            attachment_form = CatalogAttachmentForm(request.POST, request.FILES)
            if attachment_form.is_valid():
                _save_catalog_attachment("shipment_document", editing_document, ShipmentDocumentAttachment, attachment_form)
                messages.success(request, "Вложение добавлено.")
            else:
                messages.error(request, _first_form_error(attachment_form))
            return redirect(reverse("shipment_document_edit", kwargs={"document_id": editing_document.id}))

        elif action == "remove_attachment" and editing_document:
            attachment_id = request.POST.get("attachment_id", "").strip()
            if attachment_id.isdigit():
                attachment = get_object_or_404(
                    ShipmentDocumentAttachment,
                    id=int(attachment_id),
                    shipment_document=editing_document,
                )
                attachment.delete()
                messages.success(request, "Вложение удалено.")
            return redirect(reverse("shipment_document_edit", kwargs={"document_id": editing_document.id}))
    else:
        initial = {"date": timezone.localdate()} if not editing_document else {}
        selected_org = request.GET.get("organization", "").strip()
        if selected_org.isdigit() and not editing_document:
            initial["organization"] = int(selected_org)
        form = ShipmentDocumentForm(instance=editing_document, initial=initial)

    linked_equipment = ShipmentDocumentEquipment.objects.none()
    attachments = ShipmentDocumentAttachment.objects.none()
    if editing_document:
        linked_equipment = ShipmentDocumentEquipment.objects.select_related(
            "client_equipment__product_model",
            "client_equipment__organization",
            "repair_document",
        ).filter(
            shipment_document=editing_document
        )
        attachments = editing_document.attachments.all()

    context = {
        "form": form,
        "editing_document": editing_document,
        "linked_equipment": linked_equipment,
        "serial_query": serial_query,
        "attachments": attachments,
        "attachment_form": attachment_form,
        "back_url": reverse("shipment_document"),
    }
    return render(request, "shipment_document_edit.html", context)


@login_required(login_url="login")
def report_shipment_document(request):
    organization_id = request.GET.get("organization", "").strip()
    date_value = request.GET.get("date", "").strip()
    document_id = request.GET.get("document", "").strip()
    export_format = request.GET.get("export", "").strip().lower()

    document = None
    available_documents = ShipmentDocument.objects.none()
    linked_equipment = ShipmentDocumentEquipment.objects.none()

    if organization_id.isdigit() and date_value:
        documents_qs = ShipmentDocument.objects.select_related(
            "organization",
            "serviceman",
        ).filter(
            organization_id=int(organization_id),
            date=date_value,
        ).order_by("-id")
        available_documents = documents_qs

        if documents_qs.exists():
            if document_id.isdigit():
                document = documents_qs.filter(id=int(document_id)).first()
                if not document:
                    messages.warning(request, "Выбранный документ не найден для указанной организации и даты.")
            elif documents_qs.count() == 1:
                document = documents_qs.first()
            else:
                messages.info(request, "Найдено несколько документов. Выберите нужный документ в поле ниже.")

            if document:
                linked_equipment = ShipmentDocumentEquipment.objects.select_related(
                    "client_equipment__product_model",
                    "client_equipment__organization",
                ).filter(
                    shipment_document=document
                )
        else:
            messages.warning(request, "Документ отгрузки по выбранной организации и дате не найден.")

    if export_format == "excel":
        if not document:
            messages.warning(request, "Для экспорта выберите конкретный документ отгрузки.")
            return redirect(reverse("report_shipment_document"))

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_shipment_document"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Отгрузка"

        sheet["A1"] = "Документ отгрузки техники"
        sheet["B1"] = f"#{document.id}"
        sheet["A2"] = "Дата"
        sheet["B2"] = document.date.strftime("%d.%m.%Y")
        sheet["A3"] = "Организация"
        sheet["B3"] = document.organization.name
        sheet["A4"] = "Сервисный инженер"
        sheet["B4"] = document.serviceman.full_name if document.serviceman_id else "-"

        header_row = 6
        headers = ["Наименование", "Серийный номер", "Инвентарный номер", "Счетчик"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)

        for index, link in enumerate(linked_equipment, start=header_row + 1):
            sheet.cell(row=index, column=1, value=str(link.client_equipment.product_model or "-"))
            sheet.cell(row=index, column=2, value=link.client_equipment.serial_number or "-")
            sheet.cell(row=index, column=3, value=link.client_equipment.inventory_number or "-")
            sheet.cell(
                row=index,
                column=4,
                value=link.client_equipment.print_counter if link.client_equipment.print_counter is not None else "-",
            )

        sheet.column_dimensions["A"].width = 45
        sheet.column_dimensions["B"].width = 25
        sheet.column_dimensions["C"].width = 25
        sheet.column_dimensions["D"].width = 12

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"shipment_report_{document.id}_{document.date}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    context = {
        "organizations": Organization.objects.all().order_by("name"),
        "organization_filter": organization_id,
        "date_filter": date_value,
        "document_filter": document_id,
        "available_documents": available_documents,
        "document": document,
        "linked_equipment": linked_equipment,
    }
    return render(request, "report_shipment_document.html", context)


@login_required(login_url="login")
def report_acceptance_document(request):
    organization_id = request.GET.get("organization", "").strip()
    date_value = request.GET.get("date", "").strip()
    document_id = request.GET.get("document", "").strip()
    export_format = request.GET.get("export", "").strip().lower()

    document = None
    available_documents = AcceptanceDocument.objects.none()
    linked_equipment = AcceptanceDocumentEquipment.objects.none()

    if organization_id.isdigit() and date_value:
        documents_qs = AcceptanceDocument.objects.select_related(
            "organization",
            "serviceman",
        ).filter(
            organization_id=int(organization_id),
            date=date_value,
        ).order_by("-id")
        available_documents = documents_qs

        if documents_qs.exists():
            if document_id.isdigit():
                document = documents_qs.filter(id=int(document_id)).first()
                if not document:
                    messages.warning(request, "Выбранный документ не найден для указанной организации и даты.")
            elif documents_qs.count() == 1:
                document = documents_qs.first()
            else:
                messages.info(request, "Найдено несколько документов. Выберите нужный документ в поле ниже.")

            if document:
                linked_equipment = AcceptanceDocumentEquipment.objects.select_related(
                    "client_equipment__product_model",
                    "client_equipment__organization",
                ).filter(
                    acceptance_document=document
                )
        else:
            messages.warning(request, "Документ приемки по выбранной организации и дате не найден.")

    if export_format == "excel":
        if not document:
            messages.warning(request, "Для экспорта выберите конкретный документ приемки.")
            return redirect(reverse("report_acceptance_document"))

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_acceptance_document"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Приемка"

        sheet["A1"] = "Документ приемки техники"
        sheet["B1"] = f"#{document.id}"
        sheet["A2"] = "Дата"
        sheet["B2"] = document.date.strftime("%d.%m.%Y")
        sheet["A3"] = "Организация"
        sheet["B3"] = document.organization.name
        sheet["A4"] = "Сервисный инженер"
        sheet["B4"] = document.serviceman.full_name if document.serviceman_id else "-"

        header_row = 6
        headers = ["Наименование", "Серийный номер", "Инвентарный номер", "Счетчик"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)

        for index, link in enumerate(linked_equipment, start=header_row + 1):
            sheet.cell(row=index, column=1, value=str(link.client_equipment.product_model or "-"))
            sheet.cell(row=index, column=2, value=link.client_equipment.serial_number or "-")
            sheet.cell(row=index, column=3, value=link.client_equipment.inventory_number or "-")
            sheet.cell(
                row=index,
                column=4,
                value=link.client_equipment.print_counter if link.client_equipment.print_counter is not None else "-",
            )

        sheet.column_dimensions["A"].width = 45
        sheet.column_dimensions["B"].width = 25
        sheet.column_dimensions["C"].width = 25
        sheet.column_dimensions["D"].width = 12

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"acceptance_report_{document.id}_{document.date}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    context = {
        "organizations": Organization.objects.all().order_by("name"),
        "organization_filter": organization_id,
        "date_filter": date_value,
        "document_filter": document_id,
        "available_documents": available_documents,
        "document": document,
        "linked_equipment": linked_equipment,
    }
    return render(request, "report_acceptance_document.html", context)


@login_required(login_url="login")
def report_repair_document(request):
    query = request.GET.get("q", "").strip()
    query_lc = _normalize_search_term(query)
    organization_filter = request.GET.get("organization", "").strip()
    status_filter = request.GET.get("status", "").strip()
    document_id = request.GET.get("document", "").strip()
    export_format = request.GET.get("export", "").strip().lower()
    sort = request.GET.get("sort", "date")
    direction = request.GET.get("dir", "desc")
    detail_view = request.GET.get("view", "list").strip().lower() == "detail"

    sort_map = {
        "id": "id",
        "date": "date",
        "organization": "organization__name",
        "equipment": "client_equipment__product_model__name",
        "serviceman": "serviceman__full_name",
        "status": "status__name",
        "status_edited_at": "status_edited_at",
    }

    if sort not in sort_map:
        sort = "date"
    if direction not in {"asc", "desc"}:
        direction = "desc"

    documents_qs = RepairDocument.objects.select_related(
        "organization",
        "serviceman",
        "status",
        "client_equipment",
        "client_equipment__product_model",
        "source_document",
    ).all()

    if query_lc:
        documents_qs = documents_qs.annotate(
            organization_name_lc=Lower("organization__name"),
            serviceman_name_lc=Lower("serviceman__full_name"),
            status_name_lc=Lower("status__name"),
            malfunction_lc=Lower("malfunction"),
            equipment_model_lc=Lower("client_equipment__product_model__name"),
            equipment_serial_lc=Lower("client_equipment__serial_number"),
            equipment_inventory_lc=Lower("client_equipment__inventory_number"),
        ).filter(
            Q(organization_name_lc__contains=query_lc)
            | Q(serviceman_name_lc__contains=query_lc)
            | Q(status_name_lc__contains=query_lc)
            | Q(malfunction_lc__contains=query_lc)
            | Q(equipment_model_lc__contains=query_lc)
            | Q(equipment_serial_lc__contains=query_lc)
            | Q(equipment_inventory_lc__contains=query_lc)
        )

    if organization_filter.isdigit():
        documents_qs = documents_qs.filter(organization_id=int(organization_filter))

    if status_filter.isdigit():
        documents_qs = documents_qs.filter(status_id=int(status_filter))

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"
    documents_qs = documents_qs.order_by(order_field, "-id")

    selected_document = None
    linked_works = RepairDocumentWork.objects.none()
    linked_parts = RepairDocumentPart.objects.none()
    linked_consumables = RepairDocumentConsumable.objects.none()

    if document_id.isdigit():
        selected_document = documents_qs.filter(id=int(document_id)).first()
        if not selected_document:
            messages.warning(request, "Выбранный документ ремонта не найден с текущими параметрами поиска.")

    if selected_document:
        linked_works = RepairDocumentWork.objects.select_related("work").filter(repair_document=selected_document)
        linked_parts = RepairDocumentPart.objects.select_related("part__brand").filter(repair_document=selected_document)
        linked_consumables = RepairDocumentConsumable.objects.select_related("consumable__brand").filter(
            repair_document=selected_document
        )

    paged_documents, page_obj, per_page = _paginate_report_queryset(request, documents_qs)

    if export_format == "excel":
        if not selected_document:
            messages.warning(request, "Для экспорта выберите конкретный документ ремонта.")
            return redirect(reverse("report_repair_document"))

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_repair_document"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Р РµРјРѕРЅС‚"

        sheet["A1"] = "Документ ремонта"
        sheet["B1"] = f"#{selected_document.id}"
        sheet["A2"] = "Р”Р°С‚Р°"
        sheet["B2"] = selected_document.date.strftime("%d.%m.%Y")
        sheet["A3"] = "Организация"
        sheet["B3"] = selected_document.organization.name
        sheet["A4"] = "Сервисный инженер"
        sheet["B4"] = selected_document.serviceman.full_name if selected_document.serviceman_id else "-"
        sheet["A5"] = "Статус"
        sheet["B5"] = selected_document.status.name if selected_document.status_id else "-"

        row = 7
        sheet.cell(row=row, column=1, value="Р Р°Р±РѕС‚С‹").font = Font(bold=True)
        row += 1
        for col, header in enumerate(["РљРѕРґ", "Наименование", "Р¦РµРЅР°", "Количество"], start=1):
            sheet.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for link in linked_works:
            sheet.cell(row=row, column=1, value=link.work.code)
            sheet.cell(row=row, column=2, value=link.work.name)
            sheet.cell(row=row, column=3, value=float(link.work.unit_price))
            sheet.cell(row=row, column=4, value=link.quantity)
            row += 1
        if not linked_works.exists():
            sheet.cell(row=row, column=1, value="Работы не добавлены")
            row += 1

        row += 1
        sheet.cell(row=row, column=1, value="Запчасти").font = Font(bold=True)
        row += 1
        for col, header in enumerate(["Наименование", "Артикул", "Бренд", "Всего", "Ручное", "Из работ"], start=1):
            sheet.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for link in linked_parts:
            sheet.cell(row=row, column=1, value=link.part.name)
            sheet.cell(row=row, column=2, value=link.part.sku or "-")
            sheet.cell(row=row, column=3, value=link.part.brand.name if link.part.brand_id else "-")
            sheet.cell(row=row, column=4, value=link.quantity)
            sheet.cell(row=row, column=5, value=link.manual_quantity)
            sheet.cell(row=row, column=6, value=link.work_quantity)
            row += 1
        if not linked_parts.exists():
            sheet.cell(row=row, column=1, value="Запчасти не добавлены")
            row += 1

        row += 1
        sheet.cell(row=row, column=1, value="Расходные материалы").font = Font(bold=True)
        row += 1
        for col, header in enumerate(["Наименование", "Артикул", "Бренд", "Всего", "Ручное", "Из работ"], start=1):
            sheet.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for link in linked_consumables:
            sheet.cell(row=row, column=1, value=link.consumable.name)
            sheet.cell(row=row, column=2, value=link.consumable.sku or "-")
            sheet.cell(row=row, column=3, value=link.consumable.brand.name if link.consumable.brand_id else "-")
            sheet.cell(row=row, column=4, value=link.quantity)
            sheet.cell(row=row, column=5, value=link.manual_quantity)
            sheet.cell(row=row, column=6, value=link.work_quantity)
            row += 1
        if not linked_consumables.exists():
            sheet.cell(row=row, column=1, value="Расходные материалы не добавлены")

        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 35
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 12
        sheet.column_dimensions["E"].width = 12
        sheet.column_dimensions["F"].width = 12

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"repair_report_{selected_document.id}_{selected_document.date}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if organization_filter:
            params["organization"] = organization_filter
        if status_filter:
            params["status"] = status_filter
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    list_params = {
        "sort": sort,
        "dir": direction,
        "per_page": per_page,
    }
    if query:
        list_params["q"] = query
    if organization_filter:
        list_params["organization"] = organization_filter
    if status_filter:
        list_params["status"] = status_filter
    current_page = request.GET.get("page", "").strip()
    if current_page and current_page.isdigit():
        list_params["page"] = current_page
    list_url = reverse("report_repair_document")
    list_qs = urlencode(list_params)
    if list_qs:
        list_url = f"{list_url}?{list_qs}"

    context = {
        "documents": paged_documents,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "detail_view": detail_view,
        "list_url": list_url,
        "selected_document": selected_document,
        "linked_works": linked_works,
        "linked_parts": linked_parts,
        "linked_consumables": linked_consumables,
        "query": query,
        "organizations": Organization.objects.all().order_by("name"),
        "statuses": StatusDirectory.objects.all().order_by("code", "name"),
        "organization_filter": organization_filter,
        "status_filter": status_filter,
        "document_filter": document_id,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "report_repair_document.html", context)


@login_required(login_url="login")
def report_part(request):
    query = request.GET.get("q", "").strip()
    query_lc = _normalize_search_term(query)
    model_name_filter = request.GET.get("model_name", "").strip()
    brand_filter = request.GET.get("brand", "").strip()
    category_filter = request.GET.get("category", "").strip()
    device_type_filter = request.GET.get("device_type", "").strip()
    selected_id = request.GET.get("item", "").strip()
    export_format = request.GET.get("export", "").strip().lower()
    sort = request.GET.get("sort", "name").strip() or "name"
    direction = request.GET.get("dir", "asc").strip() or "asc"
    detail_view = request.GET.get("view", "list").strip().lower() == "detail"

    sort_map = {
        "name": "name",
        "site": "site",
        "category": "category__name",
        "brand": "brand__name",
        "sku": "sku",
        "device_type": "char_device_type",
        "speed_print": "char_speed_print_number",
    }
    if sort not in sort_map:
        sort = "name"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    items_qs = _annotate_characteristic_values(
        Part.objects.select_related("brand", "category").all(),
        PartCharacteristic,
        "part_id",
    )
    if query_lc:
        items_qs = items_qs.annotate(
            name_lc=Lower("name"),
            site_lc=Lower("site"),
            sku_lc=Lower("sku"),
            brand_name_lc=Lower("brand__name"),
            category_name_lc=Lower("category__name"),
        ).filter(
            Q(name_lc__contains=query_lc)
            | Q(site_lc__contains=query_lc)
            | Q(sku_lc__contains=query_lc)
            | Q(brand_name_lc__contains=query_lc)
            | Q(category_name_lc__contains=query_lc)
        )
    if brand_filter.isdigit():
        items_qs = items_qs.filter(brand_id=int(brand_filter))
    if category_filter.isdigit():
        items_qs = items_qs.filter(category_id=int(category_filter))
    if device_type_filter:
        items_qs = items_qs.filter(char_device_type=device_type_filter)
    if model_name_filter:
        items_qs = items_qs.filter(compatibilities__product_model__name__icontains=model_name_filter).distinct()

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"
    items_qs = items_qs.order_by(order_field, "id")

    selected_item = None
    linked_models = ProductModel.objects.none()
    if selected_id.isdigit():
        selected_item = items_qs.filter(id=int(selected_id)).first()
        if not selected_item:
            messages.warning(request, "Выбранная запись не найдена с текущими параметрами фильтрации.")
        else:
            linked_models = ProductModel.objects.select_related("brand", "category").filter(
                part_links__part=selected_item
            ).order_by("name", "id")

    paged_items, page_obj, per_page = _paginate_report_queryset(request, items_qs)

    if export_format == "excel_list":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_part"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Запчасти"

        sheet["A1"] = "Отчет по списку запчастей"
        sheet["A1"].font = Font(bold=True)

        headers = ["Наименование", "Артикул", "Р‘СЂРµРЅРґ", "Категория", "Тип устройства", "РЎР°Р№С‚"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=3, column=col, value=header).font = Font(bold=True)

        row = 4
        for item in items_qs:
            sheet.cell(row=row, column=1, value=item.name)
            sheet.cell(row=row, column=2, value=item.sku or "-")
            sheet.cell(row=row, column=3, value=item.brand.name if item.brand_id else "-")
            sheet.cell(row=row, column=4, value=str(item.category) if item.category_id else "-")
            sheet.cell(row=row, column=5, value=item.char_device_type or "-")
            sheet.cell(row=row, column=6, value=item.site or "-")
            row += 1

        if row == 4:
            sheet.cell(row=row, column=1, value="Список пуст")

        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 22
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 24
        sheet.column_dimensions["E"].width = 18
        sheet.column_dimensions["F"].width = 34

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        brand_name = ""
        if brand_filter.isdigit():
            brand_name = Brand.objects.filter(id=int(brand_filter)).values_list("name", flat=True).first() or ""
        category_name = ""
        if category_filter.isdigit():
            category_name = ProductCategory.objects.filter(id=int(category_filter)).values_list("name", flat=True).first() or ""

        filename = _build_filtered_list_filename(
            "directory_part_report",
            [
                ("поиск", query),
                ("техника", model_name_filter),
                ("Р±СЂРµРЅРґ", brand_name),
                ("категория", category_name),
                ("тип", device_type_filter),
                ("СЃРѕСЂС‚", sort),
                ("направление", direction),
            ],
        )
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    if export_format == "excel":
        if not selected_item:
            messages.warning(request, "Для экспорта выберите конкретную запись запчасти.")
            return redirect(reverse("report_part"))

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_part"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Запчасть"

        sheet["A1"] = "Отчет по запчасти"
        sheet["B1"] = selected_item.name
        sheet["A2"] = "Артикул"
        sheet["B2"] = selected_item.sku or "-"
        sheet["A3"] = "Р‘СЂРµРЅРґ"
        sheet["B3"] = selected_item.brand.name if selected_item.brand_id else "-"
        sheet["A4"] = "Категория"
        sheet["B4"] = str(selected_item.category) if selected_item.category_id else "-"
        sheet["A5"] = "Тип устройства"
        sheet["B5"] = selected_item.char_device_type or "-"

        row = 7
        sheet.cell(row=row, column=1, value="Привязки к технике").font = Font(bold=True)
        row += 1
        headers = ["Техника", "Р‘СЂРµРЅРґ", "Категория", "Артикул"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for model in linked_models:
            sheet.cell(row=row, column=1, value=model.name)
            sheet.cell(row=row, column=2, value=model.brand.name if model.brand_id else "-")
            sheet.cell(row=row, column=3, value=str(model.category) if model.category_id else "-")
            sheet.cell(row=row, column=4, value=model.sku or "-")
            row += 1
        if not linked_models.exists():
            sheet.cell(row=row, column=1, value="Привязки к технике отсутствуют")

        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 30
        sheet.column_dimensions["C"].width = 25
        sheet.column_dimensions["D"].width = 20

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"directory_part_report_{selected_item.id}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if brand_filter:
            params["brand"] = brand_filter
        if category_filter:
            params["category"] = category_filter
        if device_type_filter:
            params["device_type"] = device_type_filter
        if model_name_filter:
            params["model_name"] = model_name_filter
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    list_params = {
        "sort": sort,
        "dir": direction,
        "per_page": per_page,
    }
    if query:
        list_params["q"] = query
    if model_name_filter:
        list_params["model_name"] = model_name_filter
    if brand_filter:
        list_params["brand"] = brand_filter
    if category_filter:
        list_params["category"] = category_filter
    if device_type_filter:
        list_params["device_type"] = device_type_filter
    current_page = request.GET.get("page", "").strip()
    if current_page and current_page.isdigit():
        list_params["page"] = current_page
    list_url = reverse("report_part")
    list_qs = urlencode(list_params)
    if list_qs:
        list_url = f"{list_url}?{list_qs}"

    context = {
        "items": paged_items,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "detail_view": detail_view,
        "list_url": list_url,
        "selected_item": selected_item,
        "linked_models": linked_models,
        "query": query,
        "model_name_filter": model_name_filter,
        "all_brands": Brand.objects.all().order_by("name"),
        "all_categories": ProductCategory.objects.all().order_by("group", "name"),
        "all_device_types": PartCharacteristic.objects.filter(characteristic_type__code="device_type").exclude(value="").values_list("value", flat=True).distinct().order_by("value"),
        "brand_filter": brand_filter,
        "category_filter": category_filter,
        "device_type_filter": device_type_filter,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "report_part.html", context)


@login_required(login_url="login")
def report_consumable(request):
    query = request.GET.get("q", "").strip()
    query_lc = _normalize_search_term(query)
    model_name_filter = request.GET.get("model_name", "").strip()
    brand_filter = request.GET.get("brand", "").strip()
    category_filter = request.GET.get("category", "").strip()
    device_type_filter = request.GET.get("device_type", "").strip()
    selected_id = request.GET.get("item", "").strip()
    export_format = request.GET.get("export", "").strip().lower()
    sort = request.GET.get("sort", "name").strip() or "name"
    direction = request.GET.get("dir", "asc").strip() or "asc"
    detail_view = request.GET.get("view", "list").strip().lower() == "detail"

    sort_map = {
        "name": "name",
        "site": "site",
        "category": "category__name",
        "brand": "brand__name",
        "sku": "sku",
        "device_type": "char_device_type",
        "speed_print": "char_speed_print_number",
    }
    if sort not in sort_map:
        sort = "name"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    items_qs = _annotate_characteristic_values(
        Consumable.objects.select_related("brand", "category").all(),
        ConsumableCharacteristic,
        "consumable_id",
    )
    if query_lc:
        items_qs = items_qs.annotate(
            name_lc=Lower("name"),
            site_lc=Lower("site"),
            sku_lc=Lower("sku"),
            brand_name_lc=Lower("brand__name"),
            category_name_lc=Lower("category__name"),
        ).filter(
            Q(name_lc__contains=query_lc)
            | Q(site_lc__contains=query_lc)
            | Q(sku_lc__contains=query_lc)
            | Q(brand_name_lc__contains=query_lc)
            | Q(category_name_lc__contains=query_lc)
        )
    if brand_filter.isdigit():
        items_qs = items_qs.filter(brand_id=int(brand_filter))
    if category_filter.isdigit():
        items_qs = items_qs.filter(category_id=int(category_filter))
    if device_type_filter:
        items_qs = items_qs.filter(char_device_type=device_type_filter)
    if model_name_filter:
        items_qs = items_qs.filter(compatibilities__product_model__name__icontains=model_name_filter).distinct()

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"
    items_qs = items_qs.order_by(order_field, "id")

    selected_item = None
    linked_models = ProductModel.objects.none()
    if selected_id.isdigit():
        selected_item = items_qs.filter(id=int(selected_id)).first()
        if not selected_item:
            messages.warning(request, "Выбранная запись не найдена с текущими параметрами фильтрации.")
        else:
            linked_models = ProductModel.objects.select_related("brand", "category").filter(
                consumable_links__consumable=selected_item
            ).order_by("name", "id")

    paged_items, page_obj, per_page = _paginate_report_queryset(request, items_qs)

    if export_format == "excel_list":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_consumable"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Расходные материалы"

        sheet["A1"] = "Отчет по списку расходных материалов"
        sheet["A1"].font = Font(bold=True)

        headers = ["Наименование", "Артикул", "Р‘СЂРµРЅРґ", "Категория", "Тип устройства", "РЎР°Р№С‚"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=3, column=col, value=header).font = Font(bold=True)

        row = 4
        for item in items_qs:
            sheet.cell(row=row, column=1, value=item.name)
            sheet.cell(row=row, column=2, value=item.sku or "-")
            sheet.cell(row=row, column=3, value=item.brand.name if item.brand_id else "-")
            sheet.cell(row=row, column=4, value=str(item.category) if item.category_id else "-")
            sheet.cell(row=row, column=5, value=item.char_device_type or "-")
            sheet.cell(row=row, column=6, value=item.site or "-")
            row += 1

        if row == 4:
            sheet.cell(row=row, column=1, value="Список пуст")

        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 22
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 24
        sheet.column_dimensions["E"].width = 18
        sheet.column_dimensions["F"].width = 34

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        brand_name = ""
        if brand_filter.isdigit():
            brand_name = Brand.objects.filter(id=int(brand_filter)).values_list("name", flat=True).first() or ""
        category_name = ""
        if category_filter.isdigit():
            category_name = ProductCategory.objects.filter(id=int(category_filter)).values_list("name", flat=True).first() or ""

        filename = _build_filtered_list_filename(
            "directory_consumable_report",
            [
                ("поиск", query),
                ("техника", model_name_filter),
                ("Р±СЂРµРЅРґ", brand_name),
                ("категория", category_name),
                ("тип", device_type_filter),
                ("СЃРѕСЂС‚", sort),
                ("направление", direction),
            ],
        )
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    if export_format == "excel":
        if not selected_item:
            messages.warning(request, "Для экспорта выберите конкретную запись расходного материала.")
            return redirect(reverse("report_consumable"))

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_consumable"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Расходный материал"

        sheet["A1"] = "Отчет по расходному материалу"
        sheet["B1"] = selected_item.name
        sheet["A2"] = "Артикул"
        sheet["B2"] = selected_item.sku or "-"
        sheet["A3"] = "Р‘СЂРµРЅРґ"
        sheet["B3"] = selected_item.brand.name if selected_item.brand_id else "-"
        sheet["A4"] = "Категория"
        sheet["B4"] = str(selected_item.category) if selected_item.category_id else "-"
        sheet["A5"] = "Тип устройства"
        sheet["B5"] = selected_item.char_device_type or "-"

        row = 7
        sheet.cell(row=row, column=1, value="Привязки к технике").font = Font(bold=True)
        row += 1
        headers = ["Техника", "Р‘СЂРµРЅРґ", "Категория", "Артикул"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for model in linked_models:
            sheet.cell(row=row, column=1, value=model.name)
            sheet.cell(row=row, column=2, value=model.brand.name if model.brand_id else "-")
            sheet.cell(row=row, column=3, value=str(model.category) if model.category_id else "-")
            sheet.cell(row=row, column=4, value=model.sku or "-")
            row += 1
        if not linked_models.exists():
            sheet.cell(row=row, column=1, value="Привязки к технике отсутствуют")

        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 30
        sheet.column_dimensions["C"].width = 25
        sheet.column_dimensions["D"].width = 20

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"directory_consumable_report_{selected_item.id}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if brand_filter:
            params["brand"] = brand_filter
        if category_filter:
            params["category"] = category_filter
        if device_type_filter:
            params["device_type"] = device_type_filter
        if model_name_filter:
            params["model_name"] = model_name_filter
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    list_params = {
        "sort": sort,
        "dir": direction,
        "per_page": per_page,
    }
    if query:
        list_params["q"] = query
    if model_name_filter:
        list_params["model_name"] = model_name_filter
    if brand_filter:
        list_params["brand"] = brand_filter
    if category_filter:
        list_params["category"] = category_filter
    if device_type_filter:
        list_params["device_type"] = device_type_filter
    current_page = request.GET.get("page", "").strip()
    if current_page and current_page.isdigit():
        list_params["page"] = current_page
    list_url = reverse("report_consumable")
    list_qs = urlencode(list_params)
    if list_qs:
        list_url = f"{list_url}?{list_qs}"

    context = {
        "items": paged_items,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "detail_view": detail_view,
        "list_url": list_url,
        "selected_item": selected_item,
        "linked_models": linked_models,
        "query": query,
        "model_name_filter": model_name_filter,
        "all_brands": Brand.objects.all().order_by("name"),
        "all_categories": ProductCategory.objects.all().order_by("group", "name"),
        "all_device_types": ConsumableCharacteristic.objects.filter(characteristic_type__code="device_type").exclude(value="").values_list("value", flat=True).distinct().order_by("value"),
        "brand_filter": brand_filter,
        "category_filter": category_filter,
        "device_type_filter": device_type_filter,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "report_consumable.html", context)


@login_required(login_url="login")
def report_work_directory(request):
    query = request.GET.get("q", "").strip()
    query_lc = _normalize_search_term(query)
    component_filter = request.GET.get("component", "").strip()
    selected_id = request.GET.get("item", "").strip()
    export_format = request.GET.get("export", "").strip().lower()
    sort = request.GET.get("sort", "code").strip() or "code"
    direction = request.GET.get("dir", "asc").strip() or "asc"
    detail_view = request.GET.get("view", "list").strip().lower() == "detail"

    sort_map = {
        "code": "code",
        "name": "name",
        "unit_price": "unit_price",
        "consumables_count": "consumables_count",
        "parts_count": "parts_count",
    }
    if sort not in sort_map:
        sort = "code"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    works_qs = WorkDirectory.objects.annotate(
        consumables_count=Count("consumable_links", distinct=True),
        parts_count=Count("part_links", distinct=True),
    )

    if query_lc:
        works_qs = works_qs.annotate(code_lc=Lower("code"), name_lc=Lower("name")).filter(
            Q(code_lc__contains=query_lc) | Q(name_lc__contains=query_lc)
        )

    if component_filter == "with_consumables":
        works_qs = works_qs.filter(consumables_count__gt=0)
    elif component_filter == "with_parts":
        works_qs = works_qs.filter(parts_count__gt=0)
    elif component_filter == "with_any":
        works_qs = works_qs.filter(Q(consumables_count__gt=0) | Q(parts_count__gt=0))

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"
    works_qs = works_qs.order_by(order_field, "id")

    selected_item = None
    selected_consumables = WorkDirectoryConsumable.objects.none()
    selected_parts = WorkDirectoryPart.objects.none()
    if selected_id.isdigit():
        selected_item = works_qs.filter(id=int(selected_id)).first()
        if not selected_item:
            messages.warning(request, "Выбранная работа не найдена с текущими параметрами фильтрации.")
        else:
            selected_consumables = WorkDirectoryConsumable.objects.select_related("consumable__brand").filter(work=selected_item).order_by(
                "consumable__name", "id"
            )
            selected_parts = WorkDirectoryPart.objects.select_related("part__brand").filter(work=selected_item).order_by(
                "part__name", "id"
            )

    paged_items, page_obj, per_page = _paginate_report_queryset(request, works_qs)

    if export_format == "excel_list":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_work_directory"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Р Р°Р±РѕС‚С‹"

        sheet["A1"] = "Отчет по списку работ"
        sheet["A1"].font = Font(bold=True)

        headers = ["РљРѕРґ", "Наименование", "Р¦РµРЅР°", "Расходники", "Запчасти"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=3, column=col, value=header).font = Font(bold=True)

        row = 4
        for item in works_qs:
            sheet.cell(row=row, column=1, value=item.code)
            sheet.cell(row=row, column=2, value=item.name)
            sheet.cell(row=row, column=3, value=float(item.unit_price))
            sheet.cell(row=row, column=4, value=item.consumables_count)
            sheet.cell(row=row, column=5, value=item.parts_count)
            row += 1

        if row == 4:
            sheet.cell(row=row, column=1, value="Список пуст")

        sheet.column_dimensions["A"].width = 16
        sheet.column_dimensions["B"].width = 38
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 14
        sheet.column_dimensions["E"].width = 14

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        component_label_map = {
            "with_consumables": "расходники",
            "with_parts": "запчасти",
            "with_any": "Р»СЋР±С‹Рµ",
        }
        filename = _build_filtered_list_filename(
            "directory_work_report",
            [
                ("поиск", query),
                ("компоненты", component_label_map.get(component_filter, "")),
                ("СЃРѕСЂС‚", sort),
                ("направление", direction),
            ],
        )
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    if export_format == "excel":
        if not selected_item:
            messages.warning(request, "Для экспорта выберите конкретную работу.")
            return redirect(reverse("report_work_directory"))

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_work_directory"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Р Р°Р±РѕС‚С‹"

        sheet["A1"] = "РћС‚С‡РµС‚ РїРѕ СЂР°Р±РѕС‚Рµ"
        sheet["B1"] = selected_item.name
        sheet["A2"] = "РљРѕРґ"
        sheet["B2"] = selected_item.code
        sheet["A3"] = "Р¦РµРЅР°"
        sheet["B3"] = float(selected_item.unit_price)

        row = 5
        sheet.cell(row=row, column=1, value="Расходные материалы").font = Font(bold=True)
        row += 1
        for col, header in enumerate(["Наименование", "Артикул", "Р‘СЂРµРЅРґ", "Количество"], start=1):
            sheet.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for link in selected_consumables:
            sheet.cell(row=row, column=1, value=link.consumable.name)
            sheet.cell(row=row, column=2, value=link.consumable.sku or "-")
            sheet.cell(row=row, column=3, value=link.consumable.brand.name if link.consumable.brand_id else "-")
            sheet.cell(row=row, column=4, value=link.quantity)
            row += 1
        if not selected_consumables.exists():
            sheet.cell(row=row, column=1, value="Расходные материалы не назначены")
            row += 1

        row += 1
        sheet.cell(row=row, column=1, value="Запчасти").font = Font(bold=True)
        row += 1
        for col, header in enumerate(["Наименование", "Артикул", "Р‘СЂРµРЅРґ", "Количество"], start=1):
            sheet.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for link in selected_parts:
            sheet.cell(row=row, column=1, value=link.part.name)
            sheet.cell(row=row, column=2, value=link.part.sku or "-")
            sheet.cell(row=row, column=3, value=link.part.brand.name if link.part.brand_id else "-")
            sheet.cell(row=row, column=4, value=link.quantity)
            row += 1
        if not selected_parts.exists():
            sheet.cell(row=row, column=1, value="Запчасти не назначены")

        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 25
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 12

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"directory_work_report_{selected_item.id}_{selected_item.code}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if component_filter:
            params["component"] = component_filter
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    list_params = {
        "sort": sort,
        "dir": direction,
        "per_page": per_page,
    }
    if query:
        list_params["q"] = query
    if component_filter:
        list_params["component"] = component_filter
    current_page = request.GET.get("page", "").strip()
    if current_page and current_page.isdigit():
        list_params["page"] = current_page
    list_url = reverse("report_work_directory")
    list_qs = urlencode(list_params)
    if list_qs:
        list_url = f"{list_url}?{list_qs}"

    context = {
        "items": paged_items,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "detail_view": detail_view,
        "list_url": list_url,
        "selected_item": selected_item,
        "selected_consumables": selected_consumables,
        "selected_parts": selected_parts,
        "query": query,
        "component_filter": component_filter,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "report_work_directory.html", context)


@login_required(login_url="login")
def report_address_directory(request):
    query = request.GET.get("q", "").strip()
    query_lc = _normalize_search_term(query)
    selected_id = request.GET.get("item", "").strip()
    export_format = request.GET.get("export", "").strip().lower()
    sort = request.GET.get("sort", "locality").strip() or "locality"
    direction = request.GET.get("dir", "asc").strip() or "asc"
    detail_view = request.GET.get("view", "list").strip().lower() == "detail"

    sort_map = {
        "locality": "locality",
        "street": "street",
        "house": "house",
        "postal_code": "postal_code",
    }
    if sort not in sort_map:
        sort = "locality"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    items_qs = Address.objects.all()
    if query_lc:
        items_qs = items_qs.annotate(
            locality_lc=Lower("locality"),
            street_lc=Lower("street"),
            house_lc=Lower("house"),
            postal_code_lc=Lower("postal_code"),
            note_lc=Lower("note"),
        ).filter(
            Q(locality_lc__contains=query_lc)
            | Q(street_lc__contains=query_lc)
            | Q(house_lc__contains=query_lc)
            | Q(postal_code_lc__contains=query_lc)
            | Q(note_lc__contains=query_lc)
        )

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"
    items_qs = items_qs.order_by(order_field, "id")

    selected_item = None
    if selected_id.isdigit():
        selected_item = items_qs.filter(id=int(selected_id)).first()
        if not selected_item:
            messages.warning(request, "Выбранный адрес не найден с текущими параметрами фильтрации.")

    paged_items, page_obj, per_page = _paginate_report_queryset(request, items_qs)

    if export_format == "excel_list":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_address_directory"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Адреса"

        sheet["A1"] = "Отчет по справочнику адресов"
        sheet["A1"].font = Font(bold=True)

        headers = ["Населенный пункт", "Улица", "Дом", "Индекс", "Примечание"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=3, column=col, value=header).font = Font(bold=True)

        row = 4
        for item in items_qs:
            sheet.cell(row=row, column=1, value=item.locality)
            sheet.cell(row=row, column=2, value=item.street or "-")
            sheet.cell(row=row, column=3, value=item.house or "-")
            sheet.cell(row=row, column=4, value=item.postal_code or "-")
            sheet.cell(row=row, column=5, value=item.note or "-")
            row += 1

        if row == 4:
            sheet.cell(row=row, column=1, value="Список пуст")

        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 30
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 14
        sheet.column_dimensions["E"].width = 45

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = _build_filtered_list_filename(
            "directory_address_report",
            [
                ("поиск", query),
                ("СЃРѕСЂС‚", sort),
                ("направление", direction),
            ],
        )
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    if export_format == "excel":
        if not selected_item:
            messages.warning(request, "Для экспорта выберите конкретный адрес.")
            return redirect(reverse("report_address_directory"))

        try:
            from openpyxl import Workbook
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_address_directory"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Адрес"

        sheet["A1"] = "Отчет по адресу"
        sheet["B1"] = str(selected_item)
        sheet["A2"] = "Индекс"
        sheet["B2"] = selected_item.postal_code or "-"
        sheet["A3"] = "Населенный пункт"
        sheet["B3"] = selected_item.locality
        sheet["A4"] = "РЈР»РёС†Р°"
        sheet["B4"] = selected_item.street or "-"
        sheet["A5"] = "Р”РѕРј"
        sheet["B5"] = selected_item.house or "-"
        sheet["A6"] = "Корпус"
        sheet["B6"] = selected_item.building or "-"
        sheet["A7"] = "Строение"
        sheet["B7"] = selected_item.structure or "-"
        sheet["A8"] = "Р­С‚Р°Р¶"
        sheet["B8"] = selected_item.floor or "-"
        sheet["A9"] = "РљРѕРјРЅР°С‚Р°"
        sheet["B9"] = selected_item.room or "-"
        sheet["A10"] = "Примечание"
        sheet["B10"] = selected_item.note or "-"

        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 60

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"directory_address_report_{selected_item.id}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    list_params = {"sort": sort, "dir": direction, "per_page": per_page}
    if query:
        list_params["q"] = query
    current_page = request.GET.get("page", "").strip()
    if current_page and current_page.isdigit():
        list_params["page"] = current_page
    list_url = reverse("report_address_directory")
    list_qs = urlencode(list_params)
    if list_qs:
        list_url = f"{list_url}?{list_qs}"

    context = {
        "items": paged_items,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "detail_view": detail_view,
        "list_url": list_url,
        "selected_item": selected_item,
        "query": query,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "report_address_directory.html", context)


@login_required(login_url="login")
def report_brand(request):
    query = request.GET.get("q", "").strip()
    query_lc = _normalize_search_term(query)
    selected_id = request.GET.get("item", "").strip()
    export_format = request.GET.get("export", "").strip().lower()
    sort = request.GET.get("sort", "name").strip() or "name"
    direction = request.GET.get("dir", "asc").strip() or "asc"
    detail_view = request.GET.get("view", "list").strip().lower() == "detail"

    sort_map = {
        "name": "name",
        "site": "site",
    }
    if sort not in sort_map:
        sort = "name"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    items_qs = Brand.objects.all()
    if query_lc:
        items_qs = items_qs.annotate(name_lc=Lower("name"), site_lc=Lower("site")).filter(
            Q(name_lc__contains=query_lc) | Q(site_lc__contains=query_lc)
        )

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"
    items_qs = items_qs.order_by(order_field, "id")

    selected_item = None
    if selected_id.isdigit():
        selected_item = items_qs.filter(id=int(selected_id)).first()
        if not selected_item:
            messages.warning(request, "Выбранный бренд не найден с текущими параметрами фильтрации.")

    paged_items, page_obj, per_page = _paginate_report_queryset(request, items_qs)

    if export_format == "excel_list":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_brand"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Р‘СЂРµРЅРґС‹"

        sheet["A1"] = "Отчет по справочнику брендов"
        sheet["A1"].font = Font(bold=True)

        headers = ["Название", "РЎР°Р№С‚"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=3, column=col, value=header).font = Font(bold=True)

        row = 4
        for item in items_qs:
            sheet.cell(row=row, column=1, value=item.name)
            sheet.cell(row=row, column=2, value=item.site or "-")
            row += 1

        if row == 4:
            sheet.cell(row=row, column=1, value="Список пуст")

        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 45

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = _build_filtered_list_filename(
            "directory_brand_report",
            [
                ("поиск", query),
                ("СЃРѕСЂС‚", sort),
                ("направление", direction),
            ],
        )
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    if export_format == "excel":
        if not selected_item:
            messages.warning(request, "Для экспорта выберите конкретный бренд.")
            return redirect(reverse("report_brand"))

        try:
            from openpyxl import Workbook
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_brand"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Р‘СЂРµРЅРґ"

        sheet["A1"] = "Отчет по бренду"
        sheet["B1"] = selected_item.name
        sheet["A2"] = "РЎР°Р№С‚"
        sheet["B2"] = selected_item.site or "-"

        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 60

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"directory_brand_report_{selected_item.id}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    list_params = {"sort": sort, "dir": direction, "per_page": per_page}
    if query:
        list_params["q"] = query
    current_page = request.GET.get("page", "").strip()
    if current_page and current_page.isdigit():
        list_params["page"] = current_page
    list_url = reverse("report_brand")
    list_qs = urlencode(list_params)
    if list_qs:
        list_url = f"{list_url}?{list_qs}"

    context = {
        "items": paged_items,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "detail_view": detail_view,
        "list_url": list_url,
        "selected_item": selected_item,
        "query": query,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "report_brand.html", context)


@login_required(login_url="login")
def report_status_directory(request):
    query = request.GET.get("q", "").strip()
    query_lc = _normalize_search_term(query)
    selected_id = request.GET.get("item", "").strip()
    export_format = request.GET.get("export", "").strip().lower()
    sort = request.GET.get("sort", "code").strip() or "code"
    direction = request.GET.get("dir", "asc").strip() or "asc"
    detail_view = request.GET.get("view", "list").strip().lower() == "detail"

    sort_map = {
        "code": "code",
        "name": "name",
    }
    if sort not in sort_map:
        sort = "code"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    items_qs = StatusDirectory.objects.all()
    if query_lc:
        search_filter = Q(name_lc__contains=query_lc) | Q(description_lc__contains=query_lc)
        if query_lc.isdigit():
            search_filter |= Q(code=int(query_lc))
        items_qs = items_qs.annotate(name_lc=Lower("name"), description_lc=Lower("description")).filter(search_filter)

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"
    items_qs = items_qs.order_by(order_field, "id")

    selected_item = None
    if selected_id.isdigit():
        selected_item = items_qs.filter(id=int(selected_id)).first()
        if not selected_item:
            messages.warning(request, "Выбранный статус не найден с текущими параметрами фильтрации.")

    paged_items, page_obj, per_page = _paginate_report_queryset(request, items_qs)

    if export_format == "excel_list":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_status_directory"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Статусы"

        sheet["A1"] = "Отчет по справочнику статусов"
        sheet["A1"].font = Font(bold=True)

        headers = ["РљРѕРґ", "Статус", "Описание"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=3, column=col, value=header).font = Font(bold=True)

        row = 4
        for item in items_qs:
            sheet.cell(row=row, column=1, value=item.code)
            sheet.cell(row=row, column=2, value=item.name)
            sheet.cell(row=row, column=3, value=item.description or "-")
            row += 1

        if row == 4:
            sheet.cell(row=row, column=1, value="Список пуст")

        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 30
        sheet.column_dimensions["C"].width = 55

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = _build_filtered_list_filename(
            "directory_status_report",
            [
                ("поиск", query),
                ("СЃРѕСЂС‚", sort),
                ("направление", direction),
            ],
        )
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    if export_format == "excel":
        if not selected_item:
            messages.warning(request, "Для экспорта выберите конкретный статус.")
            return redirect(reverse("report_status_directory"))

        try:
            from openpyxl import Workbook
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_status_directory"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Статус"

        sheet["A1"] = "Отчет по статусу"
        sheet["B1"] = selected_item.name
        sheet["A2"] = "РљРѕРґ"
        sheet["B2"] = selected_item.code
        sheet["A3"] = "Описание"
        sheet["B3"] = selected_item.description or "-"

        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 60

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"directory_status_report_{selected_item.id}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    list_params = {"sort": sort, "dir": direction, "per_page": per_page}
    if query:
        list_params["q"] = query
    current_page = request.GET.get("page", "").strip()
    if current_page and current_page.isdigit():
        list_params["page"] = current_page
    list_url = reverse("report_status_directory")
    list_qs = urlencode(list_params)
    if list_qs:
        list_url = f"{list_url}?{list_qs}"

    context = {
        "items": paged_items,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "detail_view": detail_view,
        "list_url": list_url,
        "selected_item": selected_item,
        "query": query,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "report_status_directory.html", context)


@login_required(login_url="login")
def report_product_category(request):
    query = request.GET.get("q", "").strip()
    query_lc = _normalize_search_term(query)
    group_filter = request.GET.get("group", "").strip()
    selected_id = request.GET.get("item", "").strip()
    export_format = request.GET.get("export", "").strip().lower()
    sort = request.GET.get("sort", "group").strip() or "group"
    direction = request.GET.get("dir", "asc").strip() or "asc"
    detail_view = request.GET.get("view", "list").strip().lower() == "detail"

    sort_map = {
        "group": "group",
        "name": "name",
    }
    if sort not in sort_map:
        sort = "group"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    items_qs = ProductCategory.objects.all()
    if query_lc:
        items_qs = items_qs.annotate(name_lc=Lower("name"), group_lc=Lower("group")).filter(
            Q(name_lc__contains=query_lc) | Q(group_lc__contains=query_lc)
        )

    if group_filter:
        items_qs = items_qs.filter(group=group_filter)

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"
    items_qs = items_qs.order_by(order_field, "id")

    selected_item = None
    if selected_id.isdigit():
        selected_item = items_qs.filter(id=int(selected_id)).first()
        if not selected_item:
            messages.warning(request, "Выбранная категория не найдена с текущими параметрами фильтрации.")

    paged_items, page_obj, per_page = _paginate_report_queryset(request, items_qs)

    if export_format == "excel_list":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_product_category"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Категории"

        sheet["A1"] = "Отчет по справочнику категорий товаров"
        sheet["A1"].font = Font(bold=True)

        headers = ["Группа", "Название"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=3, column=col, value=header).font = Font(bold=True)

        row = 4
        for item in items_qs:
            sheet.cell(row=row, column=1, value=item.group or "-")
            sheet.cell(row=row, column=2, value=item.name)
            row += 1

        if row == 4:
            sheet.cell(row=row, column=1, value="Список пуст")

        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 40

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = _build_filtered_list_filename(
            "directory_product_category_report",
            [
                ("поиск", query),
                ("группа", group_filter),
                ("СЃРѕСЂС‚", sort),
                ("направление", direction),
            ],
        )
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    if export_format == "excel":
        if not selected_item:
            messages.warning(request, "Для экспорта выберите конкретную категорию.")
            return redirect(reverse("report_product_category"))

        try:
            from openpyxl import Workbook
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_product_category"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Категория"

        sheet["A1"] = "Отчет по категории товара"
        sheet["B1"] = selected_item.name
        sheet["A2"] = "Группа"
        sheet["B2"] = selected_item.group or "-"

        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 60

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"directory_product_category_report_{selected_item.id}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if group_filter:
            params["group"] = group_filter
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    list_params = {"sort": sort, "dir": direction, "per_page": per_page}
    if query:
        list_params["q"] = query
    if group_filter:
        list_params["group"] = group_filter
    current_page = request.GET.get("page", "").strip()
    if current_page and current_page.isdigit():
        list_params["page"] = current_page
    list_url = reverse("report_product_category")
    list_qs = urlencode(list_params)
    if list_qs:
        list_url = f"{list_url}?{list_qs}"

    context = {
        "items": paged_items,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "detail_view": detail_view,
        "list_url": list_url,
        "selected_item": selected_item,
        "query": query,
        "group_filter": group_filter,
        "all_groups": ProductCategory.objects.exclude(group="").values_list("group", flat=True).distinct().order_by("group"),
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "report_product_category.html", context)


@login_required(login_url="login")
def report_characteristics(request):
    query = request.GET.get("q", "").strip()
    query_lc = _normalize_search_term(query)
    tags_filter = request.GET.get("tags", "").strip()
    tags_mode = request.GET.get("tags_mode", "and").strip().lower() or "and"
    if tags_mode not in {"and", "or"}:
        tags_mode = "and"
    tag_tokens = _parse_tags_value(tags_filter)
    brand_filter = request.GET.get("brand", "").strip()
    category_filter = request.GET.get("category", "").strip()
    model_filter = request.GET.get("model", "").strip()
    type_filter = request.GET.get("type", "").strip()
    selected_id = request.GET.get("item", "").strip()
    export_format = request.GET.get("export", "").strip().lower()
    sort = request.GET.get("sort", "product_model").strip() or "product_model"
    direction = request.GET.get("dir", "asc").strip() or "asc"
    detail_view = request.GET.get("view", "list").strip().lower() == "detail"

    sort_map = {
        "product_model": "product_model__name",
        "characteristic_type": "characteristic_type__name",
        "value": "value",
    }
    if sort not in sort_map:
        sort = "product_model"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    items_qs = ProductModelCharacteristic.objects.select_related(
        "product_model",
        "product_model__brand",
        "characteristic_type",
    ).all()
    if query_lc:
        items_qs = items_qs.annotate(
            model_lc=Lower("product_model__name"),
            brand_lc=Lower("product_model__brand__name"),
            type_lc=Lower("characteristic_type__name"),
            type_code_lc=Lower("characteristic_type__code"),
            value_lc=Lower("value"),
        ).filter(
            Q(model_lc__contains=query_lc)
            | Q(brand_lc__contains=query_lc)
            | Q(type_lc__contains=query_lc)
            | Q(type_code_lc__contains=query_lc)
            | Q(value_lc__contains=query_lc)
        )

    if model_filter.isdigit():
        items_qs = items_qs.filter(product_model_id=int(model_filter))

    if type_filter.isdigit():
        items_qs = items_qs.filter(characteristic_type_id=int(type_filter))

    if brand_filter.isdigit():
        items_qs = items_qs.filter(product_model__brand_id=int(brand_filter))

    if category_filter.isdigit():
        items_qs = items_qs.filter(product_model__category_id=int(category_filter))

    if tag_tokens:
        items_qs = items_qs.filter(characteristic_type__value_kind=EquipmentCharacteristicType.ValueKind.TAGS)
        if tags_mode == "or":
            tags_query = Q()
            for token in tag_tokens:
                tags_query |= Q(value__icontains=token)
            items_qs = items_qs.filter(tags_query)
        else:
            for token in tag_tokens:
                items_qs = items_qs.filter(value__icontains=token)

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"
    items_qs = items_qs.order_by(order_field, "id")

    selected_item = None
    if selected_id.isdigit():
        selected_item = items_qs.filter(id=int(selected_id)).first()
        if not selected_item:
            messages.warning(request, "Выбранная характеристика не найдена с текущими параметрами фильтрации.")

    paged_items, page_obj, per_page = _paginate_report_queryset(request, items_qs)

    if export_format == "excel_list":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_characteristics"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Характеристики"

        sheet["A1"] = "Отчет по характеристикам техники"
        sheet["A1"].font = Font(bold=True)

        headers = ["Техника", "Р‘СЂРµРЅРґ", "Характеристика", "Р—РЅР°С‡РµРЅРёРµ"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=3, column=col, value=header).font = Font(bold=True)

        row = 4
        for item in items_qs:
            sheet.cell(row=row, column=1, value=item.product_model.name)
            sheet.cell(row=row, column=2, value=item.product_model.brand.name if item.product_model.brand_id else "-")
            sheet.cell(row=row, column=3, value=item.characteristic_type.name)
            sheet.cell(row=row, column=4, value=item.value)
            row += 1

        if row == 4:
            sheet.cell(row=row, column=1, value="Список пуст")

        sheet.column_dimensions["A"].width = 36
        sheet.column_dimensions["B"].width = 24
        sheet.column_dimensions["C"].width = 28
        sheet.column_dimensions["D"].width = 38

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = _build_filtered_list_filename(
            "directory_characteristics_values_report",
            [
                ("поиск", query),
                ("теги", ",".join(tag_tokens)),
                ("режим_тегов", tags_mode),
                ("Р±СЂРµРЅРґ", brand_filter),
                ("категория", category_filter),
                ("техника", model_filter),
                ("тип", type_filter),
                ("СЃРѕСЂС‚", sort),
                ("направление", direction),
            ],
        )
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    if export_format == "excel":
        if not selected_item:
            messages.warning(request, "Для экспорта выберите конкретную характеристику.")
            return redirect(reverse("report_characteristics"))

        try:
            from openpyxl import Workbook
        except ImportError:
            messages.error(request, "Для экспорта в XLSX установите пакет openpyxl.")
            return redirect(reverse("report_characteristics"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Характеристика"

        sheet["A1"] = "Отчет по характеристике техники"
        sheet["B1"] = selected_item.product_model.name
        sheet["A2"] = "РљРѕРґ"
        sheet["B2"] = selected_item.characteristic_type.code
        sheet["A3"] = "Характеристика"
        sheet["B3"] = selected_item.characteristic_type.name
        sheet["A4"] = "Р—РЅР°С‡РµРЅРёРµ"
        sheet["B4"] = selected_item.value
        sheet["A5"] = "Р‘СЂРµРЅРґ"
        sheet["B5"] = selected_item.product_model.brand.name if selected_item.product_model.brand_id else "-"
        sheet["A6"] = "Артикул техники"
        sheet["B6"] = selected_item.product_model.sku or "-"

        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 55

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"directory_characteristics_values_report_{selected_item.id}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _set_download_filename(response, filename)
        return response

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if per_page:
            params["per_page"] = per_page
        if tags_filter:
            params["tags"] = tags_filter
        if tags_mode:
            params["tags_mode"] = tags_mode
        if brand_filter:
            params["brand"] = brand_filter
        if category_filter:
            params["category"] = category_filter
        if model_filter:
            params["model"] = model_filter
        if type_filter:
            params["type"] = type_filter
        sort_links[key] = urlencode(params)

    list_params = {
        "sort": sort,
        "dir": direction,
        "per_page": per_page,
    }
    if query:
        list_params["q"] = query
    if tags_filter:
        list_params["tags"] = tags_filter
    if tags_mode:
        list_params["tags_mode"] = tags_mode
    if brand_filter:
        list_params["brand"] = brand_filter
    if category_filter:
        list_params["category"] = category_filter
    if model_filter:
        list_params["model"] = model_filter
    if type_filter:
        list_params["type"] = type_filter
    current_page = request.GET.get("page", "").strip()
    if current_page and current_page.isdigit():
        list_params["page"] = current_page
    list_url = reverse("report_characteristics")
    list_qs = urlencode(list_params)
    if list_qs:
        list_url = f"{list_url}?{list_qs}"

    context = {
        "items": paged_items,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "detail_view": detail_view,
        "list_url": list_url,
        "selected_item": selected_item,
        "query": query,
        "tags_filter": tags_filter,
        "tags_mode": tags_mode,
        "tag_tokens": tag_tokens,
        "brand_filter": brand_filter,
        "category_filter": category_filter,
        "model_filter": model_filter,
        "type_filter": type_filter,
        "all_brands": Brand.objects.order_by("name", "id"),
        "all_categories": ProductCategory.objects.order_by("group", "name", "id"),
        "all_models": ProductModel.objects.select_related("brand").order_by("name", "id"),
        "all_types": EquipmentCharacteristicType.objects.order_by("sort_order", "name", "id"),
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "report_characteristics.html", context)


@login_required(login_url="login")
def repair_document(request):
    query = request.GET.get("q", "").strip()
    query_lc = _normalize_search_term(query)
    organization_filter = request.GET.get("organization", "").strip()
    status_filter = request.GET.get("status", "").strip()
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    sort = request.GET.get("sort", "date")
    direction = request.GET.get("dir", "desc")

    sort_map = {
        "id": "id",
        "date": "date",
        "organization": "organization__name",
        "equipment": "client_equipment__product_model__name",
        "serviceman": "serviceman__full_name",
        "status": "status__name",
        "status_edited_at": "status_edited_at",
    }

    if sort not in sort_map:
        sort = "date"
    if direction not in {"asc", "desc"}:
        direction = "desc"

    documents_qs = RepairDocument.objects.select_related(
        "organization",
        "serviceman",
        "status",
        "client_equipment",
        "client_equipment__product_model",
    ).all()
    if query_lc:
        documents_qs = documents_qs.annotate(
            organization_name_lc=Lower("organization__name"),
            serviceman_name_lc=Lower("serviceman__full_name"),
            status_name_lc=Lower("status__name"),
            malfunction_lc=Lower("malfunction"),
            equipment_model_lc=Lower("client_equipment__product_model__name"),
            equipment_serial_lc=Lower("client_equipment__serial_number"),
            equipment_inventory_lc=Lower("client_equipment__inventory_number"),
        ).filter(
            Q(organization_name_lc__contains=query_lc)
            | Q(serviceman_name_lc__contains=query_lc)
            | Q(status_name_lc__contains=query_lc)
            | Q(malfunction_lc__contains=query_lc)
            | Q(equipment_model_lc__contains=query_lc)
            | Q(equipment_serial_lc__contains=query_lc)
            | Q(equipment_inventory_lc__contains=query_lc)
        )

    if organization_filter.isdigit():
        documents_qs = documents_qs.filter(organization_id=int(organization_filter))

    if status_filter.isdigit():
        documents_qs = documents_qs.filter(status_id=int(status_filter))

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"
    documents_qs = documents_qs.order_by(order_field, "-id")
    paged_documents, page_obj, per_page = _paginate_report_queryset(request, documents_qs, default_per_page=per_page)

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if organization_filter:
            params["organization"] = organization_filter
        if status_filter:
            params["status"] = status_filter
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    context = {
        "documents": paged_documents,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "query": query,
        "organizations": Organization.objects.all().order_by("name"),
        "statuses": StatusDirectory.objects.all().order_by("code", "name"),
        "organization_filter": organization_filter,
        "status_filter": status_filter,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "repair_document.html", context)


def _clone_repair_links(source_document, target_document):
    for link in source_document.work_links.all():
        RepairDocumentWork.objects.create(
            repair_document=target_document,
            work=link.work,
            quantity=link.quantity,
        )

    for link in source_document.part_links.all():
        RepairDocumentPart.objects.create(
            repair_document=target_document,
            part=link.part,
            manual_quantity=link.manual_quantity,
            work_quantity=link.work_quantity,
            quantity=link.quantity,
        )

    for link in source_document.consumable_links.all():
        RepairDocumentConsumable.objects.create(
            repair_document=target_document,
            consumable=link.consumable,
            manual_quantity=link.manual_quantity,
            work_quantity=link.work_quantity,
            quantity=link.quantity,
        )


def _repair_links_snapshot_key(document_id):
    return f"repair_links_snapshot_{document_id}"


def _build_repair_links_snapshot(document):
    return {
        "works": [
            {"work_id": link.work_id, "quantity": link.quantity}
            for link in document.work_links.all().order_by("id")
        ],
        "parts": [
            {
                "part_id": link.part_id,
                "manual_quantity": link.manual_quantity,
                "work_quantity": link.work_quantity,
                "quantity": link.quantity,
            }
            for link in document.part_links.all().order_by("id")
        ],
        "consumables": [
            {
                "consumable_id": link.consumable_id,
                "manual_quantity": link.manual_quantity,
                "work_quantity": link.work_quantity,
                "quantity": link.quantity,
            }
            for link in document.consumable_links.all().order_by("id")
        ],
    }


def _restore_repair_links_from_snapshot(document, snapshot):
    RepairDocumentWork.objects.filter(repair_document=document).delete()
    RepairDocumentPart.objects.filter(repair_document=document).delete()
    RepairDocumentConsumable.objects.filter(repair_document=document).delete()

    for item in snapshot.get("works", []):
        work_id = item.get("work_id")
        quantity = item.get("quantity", 1)
        if work_id:
            RepairDocumentWork.objects.create(
                repair_document=document,
                work_id=work_id,
                quantity=max(1, int(quantity or 1)),
            )

    for item in snapshot.get("parts", []):
        part_id = item.get("part_id")
        quantity = item.get("quantity", 1)
        manual_quantity = item.get("manual_quantity", quantity)
        work_quantity = item.get("work_quantity", 0)
        if part_id:
            RepairDocumentPart.objects.create(
                repair_document=document,
                part_id=part_id,
                manual_quantity=max(0, int(manual_quantity or 0)),
                work_quantity=max(0, int(work_quantity or 0)),
                quantity=max(1, int(quantity or 1)),
            )

    for item in snapshot.get("consumables", []):
        consumable_id = item.get("consumable_id")
        quantity = item.get("quantity", 1)
        manual_quantity = item.get("manual_quantity", quantity)
        work_quantity = item.get("work_quantity", 0)
        if consumable_id:
            RepairDocumentConsumable.objects.create(
                repair_document=document,
                consumable_id=consumable_id,
                manual_quantity=max(0, int(manual_quantity or 0)),
                work_quantity=max(0, int(work_quantity or 0)),
                quantity=max(1, int(quantity or 1)),
            )


def _to_positive_int(value, default=1):
    try:
        return max(1, int(value))
    except (TypeError, ValueError):
        return default


def _collect_work_component_ids(work):
    part_ids = set(work.part_links.values_list("part_id", flat=True))
    consumable_ids = set(work.consumable_links.values_list("consumable_id", flat=True))
    return part_ids, consumable_ids


def _recalculate_repair_components_from_works(document, affected_part_ids=None, affected_consumable_ids=None):
    work_links = RepairDocumentWork.objects.select_related("work").filter(repair_document=document)

    part_work_totals = {}
    consumable_work_totals = {}

    for work_link in work_links:
        work_qty = _to_positive_int(work_link.quantity, default=1)

        for part_link in work_link.work.part_links.all():
            calculated = _to_positive_int(part_link.quantity, default=1) * work_qty
            part_work_totals[part_link.part_id] = part_work_totals.get(part_link.part_id, 0) + calculated

        for consumable_link in work_link.work.consumable_links.all():
            calculated = _to_positive_int(consumable_link.quantity, default=1) * work_qty
            consumable_work_totals[consumable_link.consumable_id] = consumable_work_totals.get(consumable_link.consumable_id, 0) + calculated

    if affected_part_ids is None:
        affected_part_ids = set(part_work_totals.keys()) | set(
            RepairDocumentPart.objects.filter(repair_document=document).values_list("part_id", flat=True)
        )
    else:
        affected_part_ids = set(affected_part_ids)

    existing_parts = {
        link.part_id: link
        for link in RepairDocumentPart.objects.filter(repair_document=document, part_id__in=affected_part_ids)
    }
    for part_id in affected_part_ids:
        link = existing_parts.get(part_id)
        manual_quantity = link.manual_quantity if link else 0
        work_quantity = part_work_totals.get(part_id, 0)
        total_quantity = manual_quantity + work_quantity

        if total_quantity <= 0:
            if link:
                link.delete()
            continue

        if link:
            if (
                link.work_quantity != work_quantity
                or link.manual_quantity != manual_quantity
                or link.quantity != total_quantity
            ):
                link.work_quantity = work_quantity
                link.quantity = total_quantity
                link.save(update_fields=["work_quantity", "quantity"])
        else:
            RepairDocumentPart.objects.create(
                repair_document=document,
                part_id=part_id,
                manual_quantity=0,
                work_quantity=work_quantity,
                quantity=total_quantity,
            )

    if affected_consumable_ids is None:
        affected_consumable_ids = set(consumable_work_totals.keys()) | set(
            RepairDocumentConsumable.objects.filter(repair_document=document).values_list("consumable_id", flat=True)
        )
    else:
        affected_consumable_ids = set(affected_consumable_ids)

    existing_consumables = {
        link.consumable_id: link
        for link in RepairDocumentConsumable.objects.filter(repair_document=document, consumable_id__in=affected_consumable_ids)
    }
    for consumable_id in affected_consumable_ids:
        link = existing_consumables.get(consumable_id)
        manual_quantity = link.manual_quantity if link else 0
        work_quantity = consumable_work_totals.get(consumable_id, 0)
        total_quantity = manual_quantity + work_quantity

        if total_quantity <= 0:
            if link:
                link.delete()
            continue

        if link:
            if (
                link.work_quantity != work_quantity
                or link.manual_quantity != manual_quantity
                or link.quantity != total_quantity
            ):
                link.work_quantity = work_quantity
                link.quantity = total_quantity
                link.save(update_fields=["work_quantity", "quantity"])
        else:
            RepairDocumentConsumable.objects.create(
                repair_document=document,
                consumable_id=consumable_id,
                manual_quantity=0,
                work_quantity=work_quantity,
                quantity=total_quantity,
            )


@login_required(login_url="login")
def repair_document_edit(request, document_id=None):
    editing_document = get_object_or_404(RepairDocument, id=document_id) if document_id else None
    original_status_id = editing_document.status_id if editing_document else None
    snapshot_key = _repair_links_snapshot_key(editing_document.id) if editing_document else None

    parts_query = request.GET.get("parts_q", "").strip() if request.method == "GET" else request.POST.get("parts_q", "").strip()
    parts_query_lc = _normalize_search_term(parts_query)
    consumables_query = request.GET.get("cons_q", "").strip() if request.method == "GET" else request.POST.get("cons_q", "").strip()
    consumables_query_lc = _normalize_search_term(consumables_query)
    works_query = request.GET.get("works_q", "").strip() if request.method == "GET" else request.POST.get("works_q", "").strip()
    works_query_lc = _normalize_search_term(works_query)
    works_filter = request.GET.get("works_filter", "all").strip() if request.method == "GET" else request.POST.get("works_filter", "all").strip()
    if works_filter not in {"all", "with_parts", "with_consumables", "with_any"}:
        works_filter = "all"
    works_sort = request.GET.get("works_sort", "code").strip() if request.method == "GET" else request.POST.get("works_sort", "code").strip()
    if works_sort not in {"code", "name", "price"}:
        works_sort = "code"
    works_direction = request.GET.get("works_dir", "asc").strip() if request.method == "GET" else request.POST.get("works_dir", "asc").strip()
    if works_direction not in {"asc", "desc"}:
        works_direction = "asc"

    attachment_form = CatalogAttachmentForm()

    if request.method == "POST":
        action = request.POST.get("action", "save")

        if action == "save":
            form = RepairDocumentForm(request.POST, instance=editing_document)

            if form.is_valid():
                # РџСЂРѕРІРµСЂРєР° РёР·РјРµРЅРµРЅРёСЏ СЃС‚Р°С‚СѓСЃР°
                status_changed = False
                if editing_document and form.cleaned_data.get("status"):
                    old_status_id = original_status_id
                    new_status_id = form.cleaned_data["status"].id
                    status_changed = old_status_id != new_status_id
 
                if status_changed:
                    with transaction.atomic():
                        saved = RepairDocument(
                            date=form.cleaned_data["date"],
                            repair_place=form.cleaned_data["repair_place"],
                            service_center=form.cleaned_data.get("service_center"),
                            service_center_address=form.cleaned_data.get("service_center_address"),
                            organization=form.cleaned_data["organization"],
                            serviceman=form.cleaned_data["serviceman"],
                            status=form.cleaned_data["status"],
                            client_equipment=form.cleaned_data.get("client_equipment"),
                            malfunction=form.cleaned_data.get("malfunction", ""),
                            work_performed=form.cleaned_data.get("work_performed", ""),
                            note=form.cleaned_data.get("note", ""),
                            source_document=editing_document,
                        )
                        saved.save()
                        _clone_repair_links(editing_document, saved)

                        # РСЃС‚РѕСЂРёСЏ: Сѓ РЅРѕРІРѕРіРѕ СЃС‚Р°С‚СѓСЃР° РѕСЃС‚Р°СЋС‚СЃСЏ С‚РµРєСѓС‰РёРµ СЃРїРёСЃРєРё,
                        # Р° Сѓ РїСЂРµРґС‹РґСѓС‰РµРіРѕ РІРѕСЃСЃС‚Р°РЅР°РІР»РёРІР°РµС‚СЃСЏ СЃРЅРёРјРѕРє РїСЂРё РѕС‚РєСЂС‹С‚РёРё С„РѕСЂРјС‹.
                        snapshot = request.session.get(snapshot_key) if snapshot_key else None
                        if snapshot:
                            _restore_repair_links_from_snapshot(editing_document, snapshot)
                            request.session.pop(snapshot_key, None)
                            request.session.modified = True
                    messages.success(
                        request,
                        f"Статус изменен: создан новый документ ремонта #{saved.id}. История сохранена.",
                    )
                else:
                    saved = form.save()

                    # РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ СЃРѕС…СЂР°РЅРёР» РґРѕРєСѓРјРµРЅС‚ Р±РµР· СЃРјРµРЅС‹ СЃС‚Р°С‚СѓСЃР°:
                    # РѕР±РЅРѕРІР»СЏРµРј Р±Р°Р·РѕРІС‹Р№ СЃРЅРёРјРѕРє РґР»СЏ РґР°Р»СЊРЅРµР№С€РµР№ РёСЃС‚РѕСЂРёРё.
                    if editing_document and snapshot_key:
                        request.session[snapshot_key] = _build_repair_links_snapshot(saved)
                        request.session.modified = True

                    if editing_document:
                        messages.success(request, f"Документ ремонта #{saved.id} сохранен.")
                    else:
                        messages.success(request, f"Документ ремонта #{saved.id} создан.")

                return redirect(reverse("repair_document_edit", kwargs={"document_id": saved.id}))
        else:
            if not editing_document:
                messages.error(request, "Сначала сохраните документ, чтобы добавить связанные записи.")
                return redirect(reverse("repair_document_new"))

            if action == "add_part":
                part_id = request.POST.get("part")
                quantity = request.POST.get("part_quantity", "1").strip() or "1"
                quantity_value = _to_positive_int(quantity, default=1)
                if part_id:
                    part = get_object_or_404(Part, id=part_id)
                    link, created = RepairDocumentPart.objects.get_or_create(
                        repair_document=editing_document,
                        part=part,
                        defaults={"manual_quantity": quantity_value, "work_quantity": 0, "quantity": quantity_value},
                    )
                    if not created:
                        link.manual_quantity += quantity_value
                        link.quantity = link.manual_quantity + link.work_quantity
                        link.save(update_fields=["manual_quantity", "quantity"])
                    messages.success(request, f"Запчасть «{part.name}» добавлена в документ.")

            elif action == "remove_part":
                link_id = request.POST.get("link_id")
                if link_id:
                    link = get_object_or_404(RepairDocumentPart, id=link_id, repair_document=editing_document)
                    if link.work_quantity > 0 and link.manual_quantity <= 0:
                        messages.error(request, "Запчасть сформирована из блока «Работы». Измените или удалите связанную работу.")
                    elif link.work_quantity > 0 and link.manual_quantity > 0:
                        link.manual_quantity = 0
                        link.quantity = link.work_quantity
                        link.save(update_fields=["manual_quantity", "quantity"])
                        messages.success(request, "Ручная часть количества запчасти удалена. Автоматическое количество из работ сохранено.")
                    else:
                        link.delete()
                        messages.success(request, "Запчасть удалена из документа.")

            elif action == "add_consumable":
                consumable_id = request.POST.get("consumable")
                quantity = request.POST.get("consumable_quantity", "1").strip() or "1"
                quantity_value = _to_positive_int(quantity, default=1)
                if consumable_id:
                    consumable = get_object_or_404(Consumable, id=consumable_id)
                    link, created = RepairDocumentConsumable.objects.get_or_create(
                        repair_document=editing_document,
                        consumable=consumable,
                        defaults={"manual_quantity": quantity_value, "work_quantity": 0, "quantity": quantity_value},
                    )
                    if not created:
                        link.manual_quantity += quantity_value
                        link.quantity = link.manual_quantity + link.work_quantity
                        link.save(update_fields=["manual_quantity", "quantity"])
                    messages.success(request, f"Расходный материал «{consumable.name}» добавлен в документ.")

            elif action == "remove_consumable":
                link_id = request.POST.get("link_id")
                if link_id:
                    link = get_object_or_404(RepairDocumentConsumable, id=link_id, repair_document=editing_document)
                    if link.work_quantity > 0 and link.manual_quantity <= 0:
                        messages.error(request, "Расходный материал сформирован из блока «Работы». Измените или удалите связанную работу.")
                    elif link.work_quantity > 0 and link.manual_quantity > 0:
                        link.manual_quantity = 0
                        link.quantity = link.work_quantity
                        link.save(update_fields=["manual_quantity", "quantity"])
                        messages.success(request, "Ручная часть количества расходного материала удалена. Автоматическое количество из работ сохранено.")
                    else:
                        link.delete()
                        messages.success(request, "Расходный материал удален из документа.")

            elif action == "add_work":
                work_id = request.POST.get("work")
                quantity = request.POST.get("work_quantity", "1").strip() or "1"
                quantity_value = _to_positive_int(quantity, default=1)
                if work_id:
                    work = get_object_or_404(WorkDirectory, id=work_id)
                    affected_part_ids, affected_consumable_ids = _collect_work_component_ids(work)
                    link, created = RepairDocumentWork.objects.get_or_create(
                        repair_document=editing_document,
                        work=work,
                        defaults={"quantity": quantity_value},
                    )
                    if not created:
                        link.quantity += quantity_value
                        link.save(update_fields=["quantity"])
                    _recalculate_repair_components_from_works(
                        editing_document,
                        affected_part_ids=affected_part_ids,
                        affected_consumable_ids=affected_consumable_ids,
                    )
                    messages.success(request, f"Работа «{work.name}» добавлена в документ.")

            elif action == "update_work":
                link_id = request.POST.get("link_id")
                quantity = request.POST.get("work_quantity", "1").strip() or "1"
                quantity_value = _to_positive_int(quantity, default=1)
                if link_id:
                    link = get_object_or_404(RepairDocumentWork, id=link_id, repair_document=editing_document)
                    affected_part_ids, affected_consumable_ids = _collect_work_component_ids(link.work)
                    link.quantity = quantity_value
                    link.save(update_fields=["quantity"])
                    _recalculate_repair_components_from_works(
                        editing_document,
                        affected_part_ids=affected_part_ids,
                        affected_consumable_ids=affected_consumable_ids,
                    )
                    messages.success(request, "Количество работы обновлено.")

            elif action == "remove_work":
                link_id = request.POST.get("link_id")
                if link_id:
                    link = get_object_or_404(RepairDocumentWork, id=link_id, repair_document=editing_document)
                    affected_part_ids, affected_consumable_ids = _collect_work_component_ids(link.work)
                    work_name = link.work.name
                    link.delete()
                    _recalculate_repair_components_from_works(
                        editing_document,
                        affected_part_ids=affected_part_ids,
                        affected_consumable_ids=affected_consumable_ids,
                    )
                    messages.success(request, f"Работа «{work_name}» удалена из документа.")

            elif action == "add_attachment":
                attachment_form = CatalogAttachmentForm(request.POST, request.FILES)
                if attachment_form.is_valid():
                    _save_catalog_attachment("repair_document", editing_document, RepairDocumentAttachment, attachment_form)
                    messages.success(request, "Вложение добавлено.")
                else:
                    messages.error(request, _first_form_error(attachment_form))

            elif action == "remove_attachment":
                attachment_id = request.POST.get("attachment_id", "").strip()
                if attachment_id.isdigit():
                    attachment = get_object_or_404(
                        RepairDocumentAttachment,
                        id=int(attachment_id),
                        repair_document=editing_document,
                    )
                    attachment.delete()
                    messages.success(request, "Вложение удалено.")

            redirect_params = {}
            if works_query:
                redirect_params["works_q"] = works_query
            if works_filter != "all":
                redirect_params["works_filter"] = works_filter
            if works_sort != "code":
                redirect_params["works_sort"] = works_sort
            if works_direction != "asc":
                redirect_params["works_dir"] = works_direction
            if parts_query:
                redirect_params["parts_q"] = parts_query
            if consumables_query:
                redirect_params["cons_q"] = consumables_query
            qs = urlencode(redirect_params)
            edit_url = reverse("repair_document_edit", kwargs={"document_id": editing_document.id})
            return redirect(f"{edit_url}?{qs}" if qs else edit_url)
    else:
        initial = {"date": timezone.localdate()} if not editing_document else {}
        selected_org = request.GET.get("organization", "").strip()
        selected_service_center = request.GET.get("service_center", "").strip()
        if selected_org.isdigit():
            initial["organization"] = int(selected_org)
        if selected_service_center.isdigit():
            initial["service_center"] = int(selected_service_center)
        form = RepairDocumentForm(instance=editing_document, initial=initial)

        # РЎРЅРёРјРѕРє С„РёРєСЃРёСЂСѓРµС‚СЃСЏ РѕРґРёРЅ СЂР°Р· Р·Р° СЃРµСЃСЃРёСЋ СЂРµРґР°РєС‚РёСЂРѕРІР°РЅРёСЏ РґРѕРєСѓРјРµРЅС‚Р°.
        if editing_document and snapshot_key and snapshot_key not in request.session:
            request.session[snapshot_key] = _build_repair_links_snapshot(editing_document)
            request.session.modified = True

    linked_parts = RepairDocumentPart.objects.none()
    linked_consumables = RepairDocumentConsumable.objects.none()
    linked_works = RepairDocumentWork.objects.none()
    attachments = RepairDocumentAttachment.objects.none()
    part_form = RepairDocumentPartForm()
    consumable_form = RepairDocumentConsumableForm()
    work_form = RepairDocumentWorkForm()

    if editing_document:
        linked_works = RepairDocumentWork.objects.select_related("work").filter(
            repair_document=editing_document
        ).order_by("work__code", "work__name", "id")

        linked_parts = RepairDocumentPart.objects.select_related("part__brand", "part__category").filter(
            repair_document=editing_document
        )
        linked_consumables = RepairDocumentConsumable.objects.select_related("consumable__brand", "consumable__category").filter(
            repair_document=editing_document
        )
        attachments = editing_document.attachments.all()

        part_candidates = Part.objects.select_related("brand", "category").exclude(
            id__in=linked_parts.values_list("part_id", flat=True)
        )
        if parts_query_lc:
            part_candidates = part_candidates.annotate(
                name_lc=Lower("name"),
                sku_lc=Lower("sku"),
                brand_name_lc=Lower("brand__name"),
            ).filter(
                Q(name_lc__contains=parts_query_lc)
                | Q(sku_lc__contains=parts_query_lc)
                | Q(brand_name_lc__contains=parts_query_lc)
            )
        part_form.fields["part"].queryset = part_candidates.order_by("name", "id")[:200]

        consumable_candidates = Consumable.objects.select_related("brand", "category").exclude(
            id__in=linked_consumables.values_list("consumable_id", flat=True)
        )
        if consumables_query_lc:
            consumable_candidates = consumable_candidates.annotate(
                name_lc=Lower("name"),
                sku_lc=Lower("sku"),
                brand_name_lc=Lower("brand__name"),
            ).filter(
                Q(name_lc__contains=consumables_query_lc)
                | Q(sku_lc__contains=consumables_query_lc)
                | Q(brand_name_lc__contains=consumables_query_lc)
            )
        consumable_form.fields["consumable"].queryset = consumable_candidates.order_by("name", "id")[:200]

        work_candidates = WorkDirectory.objects.exclude(
            id__in=linked_works.values_list("work_id", flat=True)
        ).annotate(
            parts_count=Count("part_links"),
            consumables_count=Count("consumable_links"),
        )

        if works_query_lc:
            work_candidates = work_candidates.annotate(
                code_lc=Lower("code"),
                name_lc=Lower("name"),
            ).filter(
                Q(code_lc__contains=works_query_lc)
                | Q(name_lc__contains=works_query_lc)
            )

        if works_filter == "with_parts":
            work_candidates = work_candidates.filter(parts_count__gt=0)
        elif works_filter == "with_consumables":
            work_candidates = work_candidates.filter(consumables_count__gt=0)
        elif works_filter == "with_any":
            work_candidates = work_candidates.filter(Q(parts_count__gt=0) | Q(consumables_count__gt=0))

        works_sort_map = {
            "code": "code",
            "name": "name",
            "price": "unit_price",
        }
        sort_field = works_sort_map.get(works_sort, "code")
        if works_direction == "desc":
            sort_field = f"-{sort_field}"

        work_form.fields["work"].queryset = work_candidates.order_by(sort_field, "id")[:200]

    works_sort_links = {}
    for key in ("code", "name", "price"):
        next_dir = "desc" if works_sort == key and works_direction == "asc" else "asc"
        params = {
            "works_sort": key,
            "works_dir": next_dir,
        }
        if works_query:
            params["works_q"] = works_query
        if works_filter != "all":
            params["works_filter"] = works_filter
        if parts_query:
            params["parts_q"] = parts_query
        if consumables_query:
            params["cons_q"] = consumables_query
        works_sort_links[key] = urlencode(params)

    context = {
        "form": form,
        "editing_document": editing_document,
        "linked_works": linked_works,
        "linked_parts": linked_parts,
        "linked_consumables": linked_consumables,
        "work_form": work_form,
        "part_form": part_form,
        "consumable_form": consumable_form,
        "works_query": works_query,
        "works_filter": works_filter,
        "works_sort": works_sort,
        "works_direction": works_direction,
        "works_sort_links": works_sort_links,
        "parts_query": parts_query,
        "consumables_query": consumables_query,
        "attachments": attachments,
        "attachment_form": attachment_form,
        "back_url": reverse("repair_document"),
        "is_status_change_create": bool(editing_document),
    }
    return render(request, "repair_document_edit.html", context)


@login_required(login_url="login")
def repair_document_view(request, document_id):
    item = get_object_or_404(
        RepairDocument.objects.select_related(
            "organization",
            "serviceman",
            "status",
            "client_equipment",
            "client_equipment__product_model",
            "source_document",
        ),
        id=document_id,
    )

    equipment_history = RepairDocument.objects.none()
    if item.client_equipment_id:
        equipment_history = RepairDocument.objects.select_related("status").filter(
            client_equipment_id=item.client_equipment_id
        ).order_by("-status_edited_at", "-id")

    context = {
        "item": item,
        "linked_works": RepairDocumentWork.objects.select_related("work").filter(repair_document=item),
        "linked_parts": RepairDocumentPart.objects.select_related("part__brand").filter(repair_document=item),
        "linked_consumables": RepairDocumentConsumable.objects.select_related("consumable__brand").filter(repair_document=item),
        "equipment_history": equipment_history,
        "back_url": reverse("repair_document"),
    }
    return render(request, "repair_document_view.html", context)


@login_required(login_url="login")
def repair_document_equipment_history(request, equipment_id):
    equipment = get_object_or_404(
        ClientEquipment.objects.select_related("organization", "product_model"),
        id=equipment_id,
    )

    documents_qs = RepairDocument.objects.select_related(
        "organization",
        "serviceman",
        "status",
        "client_equipment",
        "client_equipment__product_model",
    ).filter(
        client_equipment_id=equipment.id
    ).order_by("-status_edited_at", "-id")

    context = {
        "equipment": equipment,
        "documents": documents_qs,
        "back_url": reverse("client_equipment"),
    }
    return render(request, "repair_document_equipment_history.html", context)


@login_required(login_url="login")
def repair_document_delete(request, document_id):
    item = get_object_or_404(RepairDocument, id=document_id)

    if request.method == "POST":
        item_id = item.id
        item.delete()
        messages.success(request, f"Документ ремонта #{item_id} удален.")
        return redirect(reverse("repair_document"))

    context = {
        "item": item,
        "back_url": reverse("repair_document"),
        "title": "Удаление документа ремонта",
        "object_label": "документ ремонта",
        "q": "",
        "sort": "",
        "dir": "",
    }
    return render(request, "dictionary_delete.html", context)


@login_required(login_url="login")
def client_equipment(request):
    equipment_id = request.GET.get("edit")
    query = request.GET.get("q", "").strip()
    normalized_query = _normalize_search_term(query)
    org_filter = request.GET.get("org", "").strip()
    per_page = request.GET.get("per_page", "10").strip().lower()
    if per_page not in {"10", "30", "50", "100", "all"}:
        per_page = "50"
    sort = request.GET.get("sort", "organization")
    direction = request.GET.get("dir", "asc")
    editing_equipment = None

    sort_map = {
        "organization": "organization__name",
        "model": "product_model__name",
        "serial": "serial_number",
        "inventory": "inventory_number",
    }

    if sort not in sort_map:
        sort = "organization"
    if direction not in {"asc", "desc"}:
        direction = "asc"

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"

    if equipment_id:
        editing_equipment = get_object_or_404(ClientEquipment, id=equipment_id)

    if request.method == "POST":
        action = request.POST.get("action", "save")
        target_id = request.POST.get("equipment_id")
        post_query = request.POST.get("q", query).strip()
        post_org = request.POST.get("org", org_filter).strip()
        post_per_page = request.POST.get("per_page", per_page).strip().lower()
        post_sort = request.POST.get("sort", sort).strip() or "organization"
        post_direction = request.POST.get("dir", direction).strip() or "asc"

        params = {}
        if post_query:
            params["q"] = post_query
        if post_org:
            params["org"] = post_org
        if post_per_page in {"10", "30", "50", "100", "all"}:
            params["per_page"] = post_per_page
        if post_sort in sort_map:
            params["sort"] = post_sort
        if post_direction in {"asc", "desc"}:
            params["dir"] = post_direction

        redirect_url = reverse("client_equipment")
        if params:
            redirect_url = f"{redirect_url}?{urlencode(params)}"

        if action == "delete" and target_id:
            target = get_object_or_404(ClientEquipment, id=target_id)
            target.delete()
            messages.success(request, "Техника удалена.")
            return redirect(redirect_url)

        if target_id:
            editing_equipment = get_object_or_404(ClientEquipment, id=target_id)
            form = ClientEquipmentForm(request.POST, instance=editing_equipment)
            is_new = False
        else:
            form = ClientEquipmentForm(request.POST)
            is_new = True

        if form.is_valid():
            equipment = form.save()
            if is_new:
                messages.success(request, f"Техника добавлена для {equipment.organization.name}.")
            else:
                messages.success(request, f"Техника обновлена.")
            params["edit"] = equipment.id
            return redirect(f"{reverse('client_equipment')}?{urlencode(params)}")
    else:
        form = ClientEquipmentForm(instance=editing_equipment) if not editing_equipment else ClientEquipmentForm(instance=editing_equipment)
        if not editing_equipment:
            org_id = request.GET.get("org_init", "").strip()
            if org_id and org_id.isdigit():
                form.initial["organization"] = int(org_id)

    equipment_qs = ClientEquipment.objects.select_related("organization", "product_model").all()
    
    if normalized_query:
        equipment_qs = equipment_qs.annotate(
            organization_name_lc=Lower("organization__name"),
            model_name_lc=Lower("product_model__name"),
            serial_lc=Lower("serial_number"),
            inventory_lc=Lower("inventory_number"),
        ).filter(
            Q(organization_name_lc__contains=normalized_query)
            | Q(model_name_lc__contains=normalized_query)
            | Q(serial_lc__contains=normalized_query)
            | Q(inventory_lc__contains=normalized_query)
        )
    
    if org_filter and org_filter.isdigit():
        equipment_qs = equipment_qs.filter(organization_id=int(org_filter))
    
    equipment_qs = equipment_qs.order_by(order_field, "id")
    paged_equipment, page_obj, per_page = _paginate_report_queryset(request, equipment_qs, default_per_page=per_page)

    sort_links = {}
    for key in sort_map:
        next_dir = "desc" if sort == key and direction == "asc" else "asc"
        params = {"sort": key, "dir": next_dir}
        if query:
            params["q"] = query
        if org_filter:
            params["org"] = org_filter
        if per_page:
            params["per_page"] = per_page
        sort_links[key] = urlencode(params)

    organizations_list = Organization.objects.all().order_by("name")

    context = {
        "form": form,
        "editing_equipment": editing_equipment,
        "equipment_list": paged_equipment,
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_options": ["10", "30", "50", "100", "all"],
        "query": query,
        "org_filter": org_filter,
        "organizations_list": organizations_list,
        "sort": sort,
        "direction": direction,
        "sort_links": sort_links,
    }
    return render(request, "client_equipment.html", context)


@login_required(login_url="login")
def client_equipment_delete(request, equipment_id):
    item = get_object_or_404(ClientEquipment, id=equipment_id)

    if request.method == "POST":
        item_id = item.id
        org_name = item.organization.name
        item.delete()
        messages.success(request, f"Техника удалена из {org_name}.")
        return redirect(reverse("client_equipment"))

    context = {
        "item": item,
        "back_url": reverse("client_equipment"),
        "title": "Удаление техники",
        "object_label": "технику",
        "q": "",
        "sort": "",
        "dir": "",
    }
    return render(request, "dictionary_delete.html", context)


